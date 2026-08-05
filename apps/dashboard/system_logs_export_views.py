from __future__ import annotations

from io import BytesIO
from typing import Any
from urllib.parse import quote

from django.contrib.auth.decorators import login_required
from django.db.models import Q, QuerySet
from django.http import HttpRequest, HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_GET

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.worksheet import Worksheet

from apps.core.permissions import require_staff
from apps.dashboard.activity_logger import log_activity
from apps.dashboard.models import SystemActivityLog


# ==========================================================
# الثوابت
# ==========================================================

EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)

EXCEL_HEADERS = [
    "م",
    "المستخدم",
    "القسم",
    "نوع العملية",
    "الوصف",
    "عنوان IP",
    "التاريخ",
    "الوقت",
]

EXCEL_COLUMN_WIDTHS = {
    "A": 10,
    "B": 28,
    "C": 22,
    "D": 20,
    "E": 60,
    "F": 20,
    "G": 16,
    "H": 16,
}


# ==========================================================
# أدوات قراءة الفلاتر
# ==========================================================

def _clean_query_value(
    request: HttpRequest,
    name: str,
) -> str:
    """
    قراءة قيمة من معاملات GET وإزالة المسافات الزائدة.
    """

    return str(
        request.GET.get(name) or ""
    ).strip()


def _filtered_system_logs(
    request: HttpRequest,
) -> QuerySet[SystemActivityLog]:
    """
    تطبيق فلاتر صفحة سجل نشاط النظام على بيانات التصدير.

    الفلاتر المدعومة:
    - البحث العام.
    - القسم.
    - نوع العملية.
    - المستخدم.
    - تاريخ البداية.
    - تاريخ النهاية.
    """

    query = _clean_query_value(
        request,
        "q",
    )

    module = _clean_query_value(
        request,
        "module",
    )

    action = _clean_query_value(
        request,
        "action",
    )

    user_id = _clean_query_value(
        request,
        "user",
    )

    date_from_value = _clean_query_value(
        request,
        "date_from",
    )

    date_to_value = _clean_query_value(
        request,
        "date_to",
    )

    logs = (
        SystemActivityLog.objects
        .select_related("user")
        .all()
    )

    if query:
        logs = logs.filter(
            Q(description__icontains=query)
            | Q(module__icontains=query)
            | Q(ip_address__icontains=query)
            | Q(user__username__icontains=query)
            | Q(user__first_name__icontains=query)
            | Q(user__last_name__icontains=query)
        )

    if module:
        logs = logs.filter(
            module=module,
        )

    valid_actions = {
        value
        for value, _label
        in SystemActivityLog.ActionType.choices
    }

    if action in valid_actions:
        logs = logs.filter(
            action=action,
        )

    if user_id.isdigit():
        logs = logs.filter(
            user_id=int(user_id),
        )

    date_from = parse_date(
        date_from_value
    )

    if date_from:
        logs = logs.filter(
            created_at__date__gte=date_from,
        )

    date_to = parse_date(
        date_to_value
    )

    if date_to:
        logs = logs.filter(
            created_at__date__lte=date_to,
        )

    return logs.order_by(
        "-created_at",
        "-id",
    )


# ==========================================================
# أدوات Excel
# ==========================================================

def _safe_excel_text(
    value: Any,
) -> str:
    """
    تحويل القيمة إلى نص آمن لملف Excel.

    تمنع هذه الدالة تفسير النصوص التي تبدأ بعلامات
    المعادلات على أنها صيغ قابلة للتنفيذ داخل Excel.
    """

    if value is None:
        return ""

    text = str(value)

    dangerous_prefixes = (
        "=",
        "+",
        "-",
        "@",
        "\t",
        "\r",
    )

    if text.startswith(dangerous_prefixes):
        return f"'{text}"

    return text


def _user_display_name(
    log: SystemActivityLog,
) -> str:
    """
    استخراج اسم المستخدم المناسب للعرض في ملف Excel.
    """

    if not log.user:
        return "عملية نظامية"

    full_name = (
        log.user.get_full_name() or ""
    ).strip()

    return (
        full_name
        or log.user.username
        or "مستخدم النظام"
    )


def _build_cell_border() -> Border:
    """
    إنشاء إطار موحد لخلايا ملف Excel.
    """

    border_side = Side(
        style="thin",
        color="D8B85F",
    )

    return Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=border_side,
    )


def _style_header(
    worksheet: Worksheet,
) -> None:
    """
    تنسيق صف رؤوس الأعمدة.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0B4A32",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
        size=12,
    )

    cell_border = _build_cell_border()

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[1].height = 30


def _style_data_rows(
    worksheet: Worksheet,
    *,
    last_row: int,
) -> None:
    """
    تنسيق صفوف البيانات.
    """

    if last_row < 2:
        return

    cell_border = _build_cell_border()

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=last_row,
        min_col=1,
        max_col=len(EXCEL_HEADERS),
    ):
        for cell in row:
            cell.border = cell_border
            cell.alignment = Alignment(
                horizontal="right",
                vertical="center",
                wrap_text=True,
            )


def _style_worksheet(
    worksheet: Worksheet,
    *,
    last_row: int,
    has_data: bool,
) -> None:
    """
    تطبيق التنسيق العام على ورقة Excel.
    """

    worksheet.sheet_view.rightToLeft = True
    worksheet.freeze_panes = "A2"

    _style_header(
        worksheet
    )

    _style_data_rows(
        worksheet,
        last_row=last_row,
    )

    for column_letter, width in (
        EXCEL_COLUMN_WIDTHS.items()
    ):
        worksheet.column_dimensions[
            column_letter
        ].width = width

    if has_data:
        worksheet.auto_filter.ref = (
            f"A1:H{last_row}"
        )
    else:
        worksheet.auto_filter.ref = "A1:H1"


def _append_empty_state(
    worksheet: Worksheet,
) -> None:
    """
    إضافة رسالة واضحة عندما لا توجد نتائج مطابقة.
    """

    worksheet.append(
        [
            "لا توجد سجلات مطابقة للفلاتر الحالية.",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=8,
    )

    empty_cell = worksheet["A2"]

    empty_cell.font = Font(
        bold=True,
        color="7A6330",
        size=12,
    )

    empty_cell.fill = PatternFill(
        fill_type="solid",
        fgColor="FFF6DA",
    )

    empty_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    empty_cell.border = _build_cell_border()

    worksheet.row_dimensions[2].height = 34


def _append_log_row(
    worksheet: Worksheet,
    *,
    index: int,
    log: SystemActivityLog,
) -> None:
    """
    إضافة سجل نشاط واحد إلى ورقة Excel.
    """

    local_datetime = timezone.localtime(
        log.created_at
    )

    worksheet.append(
        [
            index,
            _safe_excel_text(
                _user_display_name(log)
            ),
            _safe_excel_text(
                log.module
            ),
            _safe_excel_text(
                log.get_action_display()
            ),
            _safe_excel_text(
                log.description
            ),
            _safe_excel_text(
                log.ip_address
            ),
            local_datetime.strftime(
                "%Y-%m-%d"
            ),
            local_datetime.strftime(
                "%H:%M:%S"
            ),
        ]
    )


def _build_workbook(
    logs: list[SystemActivityLog],
) -> Workbook:
    """
    إنشاء ملف Excel كامل لسجل نشاط النظام.
    """

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "سجل نشاط النظام"

    worksheet.append(
        EXCEL_HEADERS
    )

    for index, log in enumerate(
        logs,
        start=1,
    ):
        _append_log_row(
            worksheet,
            index=index,
            log=log,
        )

    has_data = bool(logs)

    if not has_data:
        _append_empty_state(
            worksheet
        )

    _style_worksheet(
        worksheet,
        last_row=worksheet.max_row,
        has_data=has_data,
    )

    return workbook


def _workbook_to_bytes(
    workbook: Workbook,
) -> bytes:
    """
    تحويل ملف Excel إلى بيانات ثنائية.
    """

    output = BytesIO()

    try:
        workbook.save(output)
        output.seek(0)

        return output.getvalue()

    finally:
        output.close()


def _build_export_filename() -> str:
    """
    إنشاء اسم فريد للملف باستخدام التاريخ والوقت المحلي.
    """

    generated_at = timezone.localtime(
        timezone.now()
    )

    return (
        "system-activity-logs-"
        f"{generated_at:%Y%m%d-%H%M%S}.xlsx"
    )


def _build_excel_response(
    *,
    content: bytes,
    filename: str,
) -> HttpResponse:
    """
    إنشاء استجابة تنزيل آمنة لملف Excel.
    """

    response = HttpResponse(
        content,
        content_type=EXCEL_CONTENT_TYPE,
    )

    encoded_filename = quote(
        filename
    )

    response["Content-Disposition"] = (
        "attachment; "
        f'filename="{filename}"; '
        f"filename*=UTF-8''{encoded_filename}"
    )

    response["X-Content-Type-Options"] = (
        "nosniff"
    )

    response["Cache-Control"] = (
        "no-store, no-cache, must-revalidate, "
        "max-age=0"
    )

    response["Pragma"] = "no-cache"
    response["Expires"] = "0"

    return response


# ==========================================================
# تصدير سجل النشاط
# ==========================================================

@login_required
@require_GET
def export_system_logs_excel(
    request: HttpRequest,
) -> HttpResponse:
    """
    تصدير سجل نشاط النظام إلى ملف Excel.

    الخصائص:
    - تطبيق فلاتر الصفحة الحالية.
    - دعم النصوص العربية والاتجاه من اليمين لليسار.
    - نجاح التصدير حتى عند عدم وجود نتائج.
    - حماية الخلايا من Formula Injection.
    - تسجيل عملية التصدير في سجل نشاط النظام.
    """

    require_staff(
        request.user
    )

    queryset = _filtered_system_logs(
        request
    )

    # تحويل النتائج إلى قائمة قبل تسجيل عملية التصدير،
    # حتى لا تدخل عملية التصدير الحالية ضمن الملف.
    logs = list(
        queryset
    )

    workbook = _build_workbook(
        logs
    )

    excel_content = _workbook_to_bytes(
        workbook
    )

    filename = _build_export_filename()

    response = _build_excel_response(
        content=excel_content,
        filename=filename,
    )

    log_activity(
        user=request.user,
        module="سجل نشاط النظام",
        action=(
            SystemActivityLog
            .ActionType
            .EXPORT
        ),
        description=(
            "تم تصدير سجل نشاط النظام إلى Excel. "
            f"عدد السجلات: {len(logs)}"
        ),
        request=request,
    )

    return response
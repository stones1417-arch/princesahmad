from __future__ import annotations

from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from openpyxl import load_workbook

from apps.dashboard.models import SystemActivityLog


User = get_user_model()


class SystemLogsExcelExportTests(TestCase):
    """
    اختبارات تصدير سجل نشاط النظام إلى Excel.
    """

    @classmethod
    def setUpTestData(cls):
        cls.admin_user = User.objects.create_superuser(
            username="system_logs_admin",
            email="system-logs-admin@example.com",
            password="StrongAdminPassword123!",
        )

        cls.normal_user = User.objects.create_user(
            username="system_logs_user",
            email="system-logs-user@example.com",
            password="StrongUserPassword123!",
            is_staff=False,
            is_active=True,
        )

        cls.first_log = SystemActivityLog.objects.create(
            user=cls.admin_user,
            module="الموظفون",
            action=SystemActivityLog.ActionType.CREATE,
            description="تم إنشاء موظف جديد",
            ip_address="127.0.0.1",
        )

        cls.second_log = SystemActivityLog.objects.create(
            user=cls.admin_user,
            module="الراحات",
            action=SystemActivityLog.ActionType.UPDATE,
            description="تم تعديل راحة موظف",
            ip_address="127.0.0.1",
        )

    def setUp(self):
        self.export_url = reverse(
            "dashboard:system-logs-export-excel"
        )

    def login_admin(self):
        """
        تسجيل الدخول بحساب مدير النظام.
        """

        logged_in = self.client.login(
            username="system_logs_admin",
            password="StrongAdminPassword123!",
        )

        self.assertTrue(
            logged_in,
            msg="تعذر تسجيل الدخول بحساب مدير النظام.",
        )

    def _load_exported_workbook(self, response):
        """
        قراءة ملف Excel المرسل في الاستجابة.
        """

        stream = BytesIO(
            response.content
        )

        return load_workbook(
            stream,
            read_only=False,
            data_only=True,
        )

    def test_export_route_can_be_reversed(self):
        """
        يجب أن يكون مسار التصدير مسجلًا
        في namespace لوحة التحكم.
        """

        self.assertEqual(
            self.export_url,
            "/system-logs/export/excel/",
        )

    def test_anonymous_user_is_redirected_to_login(self):
        """
        المستخدم غير المسجل يجب تحويله
        إلى صفحة تسجيل الدخول.
        """

        response = self.client.get(
            self.export_url
        )

        self.assertEqual(
            response.status_code,
            302,
        )

        self.assertIn(
            "login",
            response.url,
        )

    def test_normal_user_cannot_export_system_logs(self):
        """
        المستخدم غير الإداري لا يجب أن يتمكن
        من تصدير سجل النشاط.
        """

        logged_in = self.client.login(
            username="system_logs_user",
            password="StrongUserPassword123!",
        )

        self.assertTrue(
            logged_in
        )

        response = self.client.get(
            self.export_url
        )

        self.assertIn(
            response.status_code,
            (
                302,
                403,
            ),
        )

    def test_admin_can_export_excel(self):
        """
        مدير النظام يستطيع تصدير سجل النشاط.
        """

        self.login_admin()

        response = self.client.get(
            self.export_url
        )

        self.assertEqual(
            response.status_code,
            200,
        )

        self.assertEqual(
            response["Content-Type"],
            (
                "application/"
                "vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

        self.assertIn(
            ".xlsx",
            response["Content-Disposition"],
        )

        self.assertEqual(
            response["X-Content-Type-Options"],
            "nosniff",
        )

    def test_export_contains_arabic_headers(self):
        """
        ملف Excel يجب أن يحتوي على رؤوس عربية.
        """

        self.login_admin()

        response = self.client.get(
            self.export_url
        )

        workbook = self._load_exported_workbook(
            response
        )

        worksheet = workbook.active

        headers = [
            worksheet.cell(
                row=1,
                column=column,
            ).value
            for column in range(1, 9)
        ]

        self.assertEqual(
            headers,
            [
                "م",
                "المستخدم",
                "القسم",
                "نوع العملية",
                "الوصف",
                "عنوان IP",
                "التاريخ",
                "الوقت",
            ],
        )

    def test_export_sheet_is_right_to_left(self):
        """
        ورقة Excel يجب أن تكون مهيأة للعربية.
        """

        self.login_admin()

        response = self.client.get(
            self.export_url
        )

        workbook = self._load_exported_workbook(
            response
        )

        worksheet = workbook.active

        self.assertTrue(
            worksheet.sheet_view.rightToLeft
        )

    def test_export_freezes_header_row(self):
        """
        يجب تثبيت صف العناوين أثناء التمرير.
        """

        self.login_admin()

        response = self.client.get(
            self.export_url
        )

        workbook = self._load_exported_workbook(
            response
        )

        worksheet = workbook.active

        self.assertEqual(
            worksheet.freeze_panes,
            "A2",
        )

    def test_export_row_count_matches_logs_before_export_activity(self):
        """
        عدد صفوف البيانات يجب أن يطابق السجلات
        الموجودة قبل تسجيل عملية التصدير الحالية.
        """

        self.login_admin()

        logs_count_before_export = (
            SystemActivityLog.objects.count()
        )

        response = self.client.get(
            self.export_url
        )

        workbook = self._load_exported_workbook(
            response
        )

        worksheet = workbook.active

        exported_data_rows = (
            worksheet.max_row - 1
        )

        self.assertEqual(
            exported_data_rows,
            logs_count_before_export,
        )

    def test_export_registers_export_activity(self):
        """
        يجب تسجيل عملية التصدير داخل سجل النشاط
        بعد نجاحها.
        """

        self.login_admin()

        previous_export_count = (
            SystemActivityLog.objects
            .filter(
                action=(
                    SystemActivityLog
                    .ActionType
                    .EXPORT
                )
            )
            .count()
        )

        response = self.client.get(
            self.export_url
        )

        self.assertEqual(
            response.status_code,
            200,
        )

        current_export_count = (
            SystemActivityLog.objects
            .filter(
                action=(
                    SystemActivityLog
                    .ActionType
                    .EXPORT
                )
            )
            .count()
        )

        self.assertEqual(
            current_export_count,
            previous_export_count + 1,
        )

        export_log = (
            SystemActivityLog.objects
            .filter(
                action=(
                    SystemActivityLog
                    .ActionType
                    .EXPORT
                )
            )
            .latest("created_at")
        )

        self.assertEqual(
            export_log.module,
            "سجل نشاط النظام",
        )

        self.assertIn(
            "تم تصدير سجل نشاط النظام",
            export_log.description,
        )

    def test_export_respects_module_filter(self):
        """
        فلتر القسم يجب أن يطبق على ملف التصدير.
        """

        self.login_admin()

        response = self.client.get(
            self.export_url,
            {
                "module": "الراحات",
            },
        )

        workbook = self._load_exported_workbook(
            response
        )

        worksheet = workbook.active

        self.assertEqual(
            worksheet.max_row,
            2,
        )

        self.assertEqual(
            worksheet["C2"].value,
            "الراحات",
        )

        self.assertEqual(
            worksheet["E2"].value,
            "تم تعديل راحة موظف",
        )

    def test_export_respects_action_filter(self):
        """
        فلتر نوع العملية يجب أن يطبق
        على ملف التصدير.
        """

        self.login_admin()

        response = self.client.get(
            self.export_url,
            {
                "action": (
                    SystemActivityLog
                    .ActionType
                    .CREATE
                ),
            },
        )

        workbook = self._load_exported_workbook(
            response
        )

        worksheet = workbook.active

        self.assertEqual(
            worksheet.max_row,
            2,
        )

        self.assertEqual(
            worksheet["D2"].value,
            "إنشاء",
        )

        self.assertEqual(
            worksheet["C2"].value,
            "الموظفون",
        )

    def test_export_respects_search_filter(self):
        """
        البحث العام يجب أن يطبق على التصدير.
        """

        self.login_admin()

        response = self.client.get(
            self.export_url,
            {
                "q": "راحة موظف",
            },
        )

        workbook = self._load_exported_workbook(
            response
        )

        worksheet = workbook.active

        self.assertEqual(
            worksheet.max_row,
            2,
        )

        self.assertEqual(
            worksheet["E2"].value,
            "تم تعديل راحة موظف",
        )

    def test_export_succeeds_without_matching_data(self):
        """
        يجب نجاح التصدير حتى عند عدم وجود
        نتائج مطابقة للفلاتر.
        """

        self.login_admin()

        response = self.client.get(
            self.export_url,
            {
                "module": "قسم غير موجود",
            },
        )

        self.assertEqual(
            response.status_code,
            200,
        )

        workbook = self._load_exported_workbook(
            response
        )

        worksheet = workbook.active

        self.assertEqual(
            worksheet.max_row,
            2,
        )

        self.assertEqual(
            worksheet["A2"].value,
            "لا توجد سجلات مطابقة للفلاتر الحالية.",
        )

        merged_ranges = {
            str(cell_range)
            for cell_range in worksheet.merged_cells.ranges
        }

        self.assertIn(
            "A2:H2",
            merged_ranges,
        )

    def test_export_uses_safe_text_for_formula_prefix(self):
        """
        يجب حماية النصوص التي تبدأ بعلامات
        قد يفسرها Excel كصيغة.
        """

        SystemActivityLog.objects.create(
            user=self.admin_user,
            module="الاختبارات",
            action=SystemActivityLog.ActionType.OTHER,
            description="=HYPERLINK('https://example.com')",
            ip_address="127.0.0.1",
        )

        self.login_admin()

        response = self.client.get(
            self.export_url,
            {
                "module": "الاختبارات",
            },
        )

        workbook = self._load_exported_workbook(
            response
        )

        worksheet = workbook.active

        self.assertTrue(
            str(
                worksheet["E2"].value
            ).startswith("'=")
        )
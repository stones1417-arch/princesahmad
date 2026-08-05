from __future__ import annotations

from datetime import date
from urllib.parse import quote

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Count, Q, QuerySet
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .activity_logger import log_activity
from .models import SystemActivityLog


ALLOWED_PAGE_SIZES = (25, 50, 100)


def _require_admin(user):
    """
    السماح لمدير النظام فقط بعرض سجل النشاط وتصديره.

    يسمح للحساب إذا كان Superuser أو كان اسم المستخدم admin.
    """
    is_admin = (
        user.is_authenticated
        and (
            user.is_superuser
            or user.username.lower() == "admin"
        )
    )

    if not is_admin:
        raise PermissionDenied(
            "هذه الصفحة متاحة لمدير النظام فقط."
        )


def _parse_iso_date(value: str) -> date | None:
    value = (value or "").strip()

    if not value:
        return None

    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _parse_positive_int(value: str) -> int | None:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None

    return parsed if parsed > 0 else None


def _get_filters(request: HttpRequest) -> dict:
    q = (request.GET.get("q") or "").strip()
    module = (request.GET.get("module") or "").strip()
    action = (request.GET.get("action") or "").strip()
    user_id = _parse_positive_int(request.GET.get("user") or "")

    date_from_value = (request.GET.get("date_from") or "").strip()
    date_to_value = (request.GET.get("date_to") or "").strip()

    date_from = _parse_iso_date(date_from_value)
    date_to = _parse_iso_date(date_to_value)

    try:
        per_page = int(request.GET.get("per_page") or 25)
    except (TypeError, ValueError):
        per_page = 25

    if per_page not in ALLOWED_PAGE_SIZES:
        per_page = 25

    return {
        "q": q,
        "module": module,
        "action": action,
        "user_id": user_id,
        "date_from": date_from,
        "date_to": date_to,
        "date_from_value": date_from_value if date_from else "",
        "date_to_value": date_to_value if date_to else "",
        "per_page": per_page,
    }


def _build_logs_queryset(filters: dict) -> QuerySet[SystemActivityLog]:
    logs = (
        SystemActivityLog.objects
        .select_related("user")
        .order_by("-created_at")
    )

    q = filters["q"]

    if q:
        logs = logs.filter(
            Q(description__icontains=q)
            | Q(module__icontains=q)
            | Q(user__username__icontains=q)
            | Q(user__first_name__icontains=q)
            | Q(user__last_name__icontains=q)
            | Q(ip_address__icontains=q)
        )

    if filters["module"]:
        logs = logs.filter(module=filters["module"])

    if filters["action"]:
        logs = logs.filter(action=filters["action"])

    if filters["user_id"]:
        logs = logs.filter(user_id=filters["user_id"])

    if filters["date_from"]:
        logs = logs.filter(created_at__date__gte=filters["date_from"])

    if filters["date_to"]:
        logs = logs.filter(created_at__date__lte=filters["date_to"])

    return logs


def _build_filter_query(request: HttpRequest) -> str:
    query = request.GET.copy()
    query.pop("page", None)
    return query.urlencode()


@login_required
def system_activity_logs_view(request: HttpRequest) -> HttpResponse:
    _require_admin(request.user)

    filters = _get_filters(request)

    if (
        filters["date_from"]
        and filters["date_to"]
        and filters["date_from"] > filters["date_to"]
    ):
        messages.warning(
            request,
            "تاريخ البداية يجب أن يكون قبل تاريخ النهاية.",
        )
        filters["date_from"] = None
        filters["date_to"] = None
        filters["date_from_value"] = ""
        filters["date_to_value"] = ""

    logs = _build_logs_queryset(filters)
    today = timezone.localdate()

    action_totals = logs.aggregate(
        create_count=Count(
            "id",
            filter=Q(action=SystemActivityLog.ActionType.CREATE),
        ),
        update_count=Count(
            "id",
            filter=Q(action=SystemActivityLog.ActionType.UPDATE),
        ),
        delete_count=Count(
            "id",
            filter=Q(action=SystemActivityLog.ActionType.DELETE),
        ),
        export_count=Count(
            "id",
            filter=Q(action=SystemActivityLog.ActionType.EXPORT),
        ),
        approve_count=Count(
            "id",
            filter=Q(action=SystemActivityLog.ActionType.APPROVE),
        ),
    )

    paginator = Paginator(logs, filters["per_page"])
    page_obj = paginator.get_page(request.GET.get("page"))

    User = get_user_model()

    users = (
        User.objects
        .filter(system_activity_logs__isnull=False)
        .distinct()
        .order_by("username")
    )

    modules = (
        SystemActivityLog.objects
        .exclude(module="")
        .values_list("module", flat=True)
        .distinct()
        .order_by("module")
    )

    context = {
        "logs": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,

        "logs_count": paginator.count,
        "today_count": logs.filter(created_at__date=today).count(),

        "create_count": action_totals["create_count"] or 0,
        "update_count": action_totals["update_count"] or 0,
        "delete_count": action_totals["delete_count"] or 0,
        "export_count": action_totals["export_count"] or 0,
        "approve_count": action_totals["approve_count"] or 0,

        "modules": modules,
        "users": users,
        "actions": SystemActivityLog.ActionType.choices,
        "page_sizes": ALLOWED_PAGE_SIZES,

        "selected_q": filters["q"],
        "selected_module": filters["module"],
        "selected_action": filters["action"],
        "selected_user": str(filters["user_id"] or ""),
        "selected_date_from": filters["date_from_value"],
        "selected_date_to": filters["date_to_value"],
        "selected_per_page": filters["per_page"],

        "filter_query": _build_filter_query(request),
    }

    return render(
        request,
        "dashboard/system_activity_logs.html",
        context,
    )


def _auto_width(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        sheet.column_dimensions[column_letter].width = min(
            max(max_length + 4, 12),
            48,
        )


@login_required
def export_system_activity_logs_excel(
    request: HttpRequest,
) -> HttpResponse:
    _require_admin(request.user)

    filters = _get_filters(request)
    logs = _build_logs_queryset(filters)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "سجل نشاط النظام"
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = "A6"

    primary_fill = PatternFill(
        fill_type="solid",
        fgColor="0F6F55",
    )
    gold_fill = PatternFill(
        fill_type="solid",
        fgColor="C9A23F",
    )
    soft_fill = PatternFill(
        fill_type="solid",
        fgColor="EAF6F1",
    )
    white_font = Font(
        color="FFFFFF",
        bold=True,
        size=12,
    )
    title_font = Font(
        color="FFFFFF",
        bold=True,
        size=18,
    )
    meta_font = Font(
        color="15201D",
        bold=True,
        size=10,
    )
    thin_border = Border(
        left=Side(style="thin", color="D8E2DD"),
        right=Side(style="thin", color="D8E2DD"),
        top=Side(style="thin", color="D8E2DD"),
        bottom=Side(style="thin", color="D8E2DD"),
    )
    centered = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    sheet.merge_cells("A1:G1")
    title_cell = sheet["A1"]
    title_cell.value = "منصة أبواب - سجل نشاط النظام"
    title_cell.fill = primary_fill
    title_cell.font = title_font
    title_cell.alignment = centered
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells("A2:G2")
    sheet["A2"] = (
        f"تاريخ التصدير: "
        f"{timezone.localtime().strftime('%Y-%m-%d %H:%M')}"
    )
    sheet["A2"].fill = gold_fill
    sheet["A2"].font = meta_font
    sheet["A2"].alignment = centered

    filter_parts = []

    if filters["q"]:
        filter_parts.append(f"البحث: {filters['q']}")

    if filters["module"]:
        filter_parts.append(f"القسم: {filters['module']}")

    if filters["action"]:
        action_label = dict(SystemActivityLog.ActionType.choices).get(
            filters["action"],
            filters["action"],
        )
        filter_parts.append(f"العملية: {action_label}")

    if filters["date_from"]:
        filter_parts.append(
            f"من: {filters['date_from'].isoformat()}"
        )

    if filters["date_to"]:
        filter_parts.append(
            f"إلى: {filters['date_to'].isoformat()}"
        )

    sheet.merge_cells("A3:G3")
    sheet["A3"] = (
        "الفلاتر: " + " | ".join(filter_parts)
        if filter_parts
        else "الفلاتر: جميع السجلات"
    )
    sheet["A3"].fill = soft_fill
    sheet["A3"].font = meta_font
    sheet["A3"].alignment = centered

    headers = [
        "المستخدم",
        "القسم",
        "نوع العملية",
        "الوصف",
        "عنوان IP",
        "التاريخ",
        "الوقت",
    ]

    header_row = 5

    for column_number, title in enumerate(headers, start=1):
        cell = sheet.cell(
            row=header_row,
            column=column_number,
            value=title,
        )
        cell.fill = primary_fill
        cell.font = white_font
        cell.alignment = centered
        cell.border = thin_border

    first_data_row = header_row + 1
    records_count = 0

    for row_number, log in enumerate(logs.iterator(), start=first_data_row):
        created_at = timezone.localtime(log.created_at)

        user_display = "مستخدم النظام"

        if log.user:
            user_display = (
                log.user.get_full_name()
                or log.user.username
            )

        values = [
            user_display,
            log.module,
            log.get_action_display(),
            log.description,
            log.ip_address or "—",
            created_at.strftime("%Y-%m-%d"),
            created_at.strftime("%H:%M:%S"),
        ]

        for column_number, value in enumerate(values, start=1):
            cell = sheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )
            cell.alignment = centered
            cell.border = thin_border

        records_count += 1

    if records_count:
        table_ref = f"A{header_row}:G{header_row + records_count}"
        table = Table(
            displayName="SystemActivityLogsTable",
            ref=table_ref,
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    sheet.auto_filter.ref = (
        f"A{header_row}:G{max(header_row, header_row + records_count)}"
    )
    sheet.print_title_rows = f"1:{header_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    _auto_width(sheet)
    sheet.column_dimensions["D"].width = 55

    file_name = (
        f"system_activity_logs_"
        f"{timezone.localdate().isoformat()}.xlsx"
    )

    log_activity(
        user=request.user,
        module="سجل نشاط النظام",
        action=SystemActivityLog.ActionType.EXPORT,
        description=(
            f"تم تصدير سجل نشاط النظام إلى Excel "
            f"- عدد السجلات: {records_count}"
        ),
        request=request,
    )

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        f"attachment; filename*=UTF-8''{quote(file_name)}"
    )

    workbook.save(response)
    return response
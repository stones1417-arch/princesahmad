from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping
from urllib.parse import quote

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)

from apps.exports_center.registry import (
    ExportColumn,
    ExportReportDefinition,
    FORMAT_EXCEL,
    get_report_definition,
)
from apps.exports_center.selectors import (
    build_report_indicators,
    select_report_queryset,
)
from apps.exports_center.services.column_selector import (
    select_export_columns,
)


# ==================================================
# الهوية الرسمية
# ==================================================

AUTHORITY_NAME = (
    "الهيئة العامة للعناية بشؤون "
    "المسجد الحرام والمسجد النبوي"
)

PLATFORM_NAME = "منصة أبواب"

PLATFORM_SUBTITLE = (
    "نظام إدارة وتشغيل أبواب المسجد الحرام"
)


# ==================================================
# الألوان
# ==================================================

BRAND_GREEN = "0F7B5C"
BRAND_DARK_GREEN = "064E3B"
BRAND_GOLD = "D4AF37"

BRAND_LIGHT_GREEN = "ECFDF5"
BRAND_LIGHT_GOLD = "FFF8E1"

WHITE = "FFFFFF"
TEXT_DARK = "17211E"
TEXT_MUTED = "66736E"

BORDER_COLOR = "DDE6E2"
ROW_ALTERNATE = "F7FAF8"
INDICATOR_FILL = "E8F5EF"


# ==================================================
# نتيجة إنشاء ملف Excel
# ==================================================

@dataclass(frozen=True)
class ExcelExportResult:
    """
    نتيجة بناء ملف Excel.

    تستخدم لاحقًا في:
    - حفظ الملف داخل ExportLog.
    - إرسال الملف للمستخدم.
    - تسجيل عدد السجلات والحجم.
    """

    content: bytes
    file_name: str
    records_count: int
    indicators: dict[str, Any]
    report_key: str
    report_title: str

    @property
    def file_size(self) -> int:
        """
        حجم الملف بالبايت.
        """

        return len(self.content)


# ==================================================
# محرك Excel الموحد
# ==================================================

class ExcelExportEngine:
    """
    محرك مؤسسي موحد لإنشاء ملفات Excel الرسمية.

    يدعم:
    - شعار الهيئة.
    - شعار منصة أبواب.
    - الهوية الرسمية.
    - اتجاه من اليمين إلى اليسار.
    - ورقة مؤشرات.
    - ورقة بيانات.
    - اختيار الأعمدة قبل التصدير.
    - الحفاظ على ترتيب الأعمدة الذي اختاره المستخدم.
    - رفض الأعمدة غير الموجودة أو غير المصرح بها.
    - رأس جدول أخضر.
    - خط أبيض عريض.
    - حدود للخلايا.
    - صفوف متبادلة.
    - فلاتر Excel.
    - تجميد رؤوس الأعمدة.
    - ضبط عرض الأعمدة.
    - التفاف النصوص.
    - إعدادات الطباعة.
    - أرقام الصفحات.
    - بيانات المستخدم والفلاتر.
    """

    def __init__(
        self,
        *,
        authority_name: str = AUTHORITY_NAME,
        platform_name: str = PLATFORM_NAME,
        platform_subtitle: str = PLATFORM_SUBTITLE,
    ) -> None:
        self.authority_name = authority_name
        self.platform_name = platform_name
        self.platform_subtitle = platform_subtitle

    # ==================================================
    # الواجهة العامة
    # ==================================================

    def build(
        self,
        *,
        report_key: str,
        queryset=None,
        filters: Mapping[str, Any] | None = None,
        user=None,
        indicators: dict[str, Any] | None = None,
        file_name: str | None = None,
        selected_columns: (
            str
            | Iterable[str]
            | None
        ) = None,
    ) -> ExcelExportResult:
        """
        إنشاء ملف Excel كامل.

        عند عدم تمرير queryset يتم جلبه تلقائيًا
        من selectors.py.

        عند عدم تمرير selected_columns يتم استخدام
        جميع أعمدة التقرير المتاحة لصيغة Excel.
        """

        report = get_report_definition(
            report_key
        )

        if not report.supports_format(
            FORMAT_EXCEL
        ):
            raise ValueError(
                f"التقرير {report.title} "
                "لا يدعم صيغة Excel."
            )

        normalized_filters = (
            self._normalize_filters(
                filters or {}
            )
        )

        columns = select_export_columns(
            report=report,
            export_format=FORMAT_EXCEL,
            selected_columns=selected_columns,
            require_at_least_one=True,
            reject_unknown=True,
        )

        if queryset is None:
            queryset = select_report_queryset(
                report_key,
                normalized_filters,
            )

        records_count = queryset.count()

        if indicators is None:
            indicators = build_report_indicators(
                report_key,
                queryset,
            )

        workbook = self._build_workbook(
            report=report,
            queryset=queryset,
            columns=columns,
            filters=normalized_filters,
            user=user,
            indicators=indicators,
            records_count=records_count,
        )

        stream = BytesIO()

        try:
            workbook.save(
                stream
            )

            content = stream.getvalue()

        finally:
            stream.close()

        resolved_file_name = (
            file_name
            or self._build_file_name(
                report
            )
        )

        resolved_file_name = (
            self._ensure_excel_extension(
                resolved_file_name
            )
        )

        return ExcelExportResult(
            content=content,
            file_name=resolved_file_name,
            records_count=records_count,
            indicators=indicators,
            report_key=report.key,
            report_title=report.title,
        )

    def build_response(
        self,
        *,
        report_key: str,
        queryset=None,
        filters: Mapping[str, Any] | None = None,
        user=None,
        indicators: dict[str, Any] | None = None,
        file_name: str | None = None,
        selected_columns: (
            str
            | Iterable[str]
            | None
        ) = None,
    ) -> HttpResponse:
        """
        إنشاء استجابة HTTP مباشرة لتنزيل Excel.
        """

        result = self.build(
            report_key=report_key,
            queryset=queryset,
            filters=filters,
            user=user,
            indicators=indicators,
            file_name=file_name,
            selected_columns=selected_columns,
        )

        response = HttpResponse(
            result.content,
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

        encoded_name = quote(
            result.file_name
        )

        response[
            "Content-Disposition"
        ] = (
            f'attachment; filename="{result.file_name}"; '
            f"filename*=UTF-8''{encoded_name}"
        )

        response[
            "Content-Length"
        ] = str(
            result.file_size
        )

        response[
            "X-Content-Type-Options"
        ] = "nosniff"

        response[
            "Cache-Control"
        ] = (
            "private, no-store, "
            "max-age=0"
        )

        return response

    # ==================================================
    # إنشاء المصنف
    # ==================================================

    def _build_workbook(
        self,
        *,
        report: ExportReportDefinition,
        queryset,
        columns: tuple[ExportColumn, ...],
        filters: dict[str, Any],
        user,
        indicators: dict[str, Any],
        records_count: int,
    ) -> Workbook:
        """
        بناء مصنف Excel باستخدام الأعمدة المختارة فقط.
        """

        workbook = Workbook()

        default_sheet = workbook.active

        if default_sheet is not None:
            workbook.remove(
                default_sheet
            )

        if report.include_indicators:
            self._create_indicators_sheet(
                workbook=workbook,
                report=report,
                indicators=indicators,
                filters=filters,
                user=user,
            )

        self._create_data_sheet(
            workbook=workbook,
            report=report,
            queryset=queryset,
            columns=columns,
            filters=filters,
            user=user,
            records_count=records_count,
        )

        if not workbook.worksheets:
            workbook.create_sheet(
                title="التقرير"
            )

        workbook.active = 0

        return workbook

    # ==================================================
    # ورقة البيانات
    # ==================================================

    def _create_data_sheet(
        self,
        *,
        workbook: Workbook,
        report: ExportReportDefinition,
        queryset,
        columns: tuple[ExportColumn, ...],
        filters: dict[str, Any],
        user,
        records_count: int,
    ) -> None:
        """
        إنشاء ورقة البيانات باستخدام الأعمدة المعتمدة.
        """

        sheet = workbook.create_sheet(
            title="البيانات"
        )

        visible_columns = max(
            len(columns),
            8,
        )

        self._configure_sheet(
            sheet=sheet,
            landscape=report.landscape,
        )

        header_row = self._build_official_header(
            sheet=sheet,
            report=report,
            filters=filters,
            user=user,
            records_count=records_count,
            visible_columns=visible_columns,
        )

        table_header_row = (
            header_row + 2
        )

        data_start_row = (
            table_header_row + 1
        )

        self._build_table_headers(
            sheet=sheet,
            columns=columns,
            header_row=table_header_row,
        )

        current_row = data_start_row

        for record in queryset.iterator(
            chunk_size=1000
        ):
            self._write_record_row(
                sheet=sheet,
                row_number=current_row,
                record=record,
                columns=columns,
                alternate=(
                    (
                        current_row
                        - data_start_row
                    )
                    % 2
                    == 1
                ),
            )

            current_row += 1

        last_data_row = (
            current_row - 1
        )

        if records_count == 0:
            self._write_empty_message(
                sheet=sheet,
                row_number=data_start_row,
                columns_count=max(
                    len(columns),
                    1,
                ),
            )

        elif columns:
            self._add_excel_table(
                sheet=sheet,
                header_row=table_header_row,
                last_data_row=last_data_row,
                columns_count=len(columns),
                report_key=report.key,
            )

            self._add_summary_row(
                sheet=sheet,
                row_number=(
                    last_data_row + 2
                ),
                columns_count=len(columns),
                records_count=records_count,
            )

        if columns:
            last_column = get_column_letter(
                len(columns)
            )

            sheet.auto_filter.ref = (
                f"A{table_header_row}:"
                f"{last_column}{table_header_row}"
            )

        sheet.freeze_panes = (
            f"A{data_start_row}"
        )

        sheet.print_title_rows = (
            f"1:{table_header_row}"
        )

        sheet.sheet_properties.pageSetUpPr.fitToPage = True

        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        print_last_row = max(
            last_data_row,
            data_start_row,
        )

        if columns:
            print_last_column = get_column_letter(
                len(columns)
            )

            sheet.print_area = (
                f"A1:"
                f"{print_last_column}"
                f"{print_last_row}"
            )

    # ==================================================
    # ورقة المؤشرات
    # ==================================================

    def _create_indicators_sheet(
        self,
        *,
        workbook: Workbook,
        report: ExportReportDefinition,
        indicators: dict[str, Any],
        filters: dict[str, Any],
        user,
    ) -> None:
        sheet = workbook.create_sheet(
            title="المؤشرات"
        )

        self._configure_sheet(
            sheet=sheet,
            landscape=False,
        )

        visible_columns = 6

        self._build_official_header(
            sheet=sheet,
            report=report,
            filters=filters,
            user=user,
            records_count=int(
                indicators.get(
                    "records_count",
                    0,
                )
                or 0
            ),
            visible_columns=visible_columns,
        )

        start_row = 7

        sheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=visible_columns,
        )

        title_cell = sheet.cell(
            row=start_row,
            column=1,
            value="مؤشرات الأداء",
        )

        self._style_cell(
            title_cell,
            fill=BRAND_GREEN,
            font_color=WHITE,
            bold=True,
            font_size=14,
            border=True,
        )

        current_row = (
            start_row + 2
        )

        flattened_indicators = (
            self._flatten_indicators(
                indicators
            )
        )

        if not flattened_indicators:
            flattened_indicators = [
                (
                    "عدد السجلات",
                    indicators.get(
                        "records_count",
                        0,
                    ),
                )
            ]

        for index, (
            label,
            value,
        ) in enumerate(
            flattened_indicators,
            start=0,
        ):
            label_row = (
                current_row + index
            )

            sheet.merge_cells(
                start_row=label_row,
                start_column=1,
                end_row=label_row,
                end_column=3,
            )

            label_cell = sheet.cell(
                row=label_row,
                column=1,
                value=label,
            )

            self._style_cell(
                label_cell,
                fill=INDICATOR_FILL,
                font_color=BRAND_DARK_GREEN,
                bold=True,
                font_size=11,
                horizontal="right",
            )

            sheet.merge_cells(
                start_row=label_row,
                start_column=4,
                end_row=label_row,
                end_column=6,
            )

            value_cell = sheet.cell(
                row=label_row,
                column=4,
                value=self._safe_excel_value(
                    value
                ),
            )

            self._style_cell(
                value_cell,
                fill=WHITE,
                font_color=TEXT_DARK,
                bold=True,
                font_size=11,
            )

            sheet.row_dimensions[
                label_row
            ].height = 26

        for column_index in range(
            1,
            visible_columns + 1,
        ):
            sheet.column_dimensions[
                get_column_letter(
                    column_index
                )
            ].width = 18

        sheet.freeze_panes = (
            f"A{start_row + 2}"
        )

        sheet.print_area = (
            f"A1:F"
            f"{current_row + len(flattened_indicators)}"
        )

    # ==================================================
    # إعدادات الورقة
    # ==================================================

    def _configure_sheet(
        self,
        *,
        sheet,
        landscape: bool,
    ) -> None:
        sheet.sheet_view.rightToLeft = True

        sheet.sheet_properties.pageSetUpPr.fitToPage = True

        sheet.page_setup.orientation = (
            "landscape"
            if landscape
            else "portrait"
        )

        sheet.page_setup.paperSize = (
            sheet.PAPERSIZE_A4
        )

        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.45
        sheet.page_margins.bottom = 0.55
        sheet.page_margins.header = 0.2
        sheet.page_margins.footer = 0.25

        sheet.oddFooter.center.text = (
            f"{self.platform_name} | "
            f"{self.authority_name}"
        )

        sheet.oddFooter.left.text = (
            "صفحة &P من &N"
        )

        sheet.oddFooter.right.text = (
            "تاريخ الطباعة: &D"
        )

    # ==================================================
    # الرأس الرسمي
    # ==================================================

    def _build_official_header(
        self,
        *,
        sheet,
        report: ExportReportDefinition,
        filters: dict[str, Any],
        user,
        records_count: int,
        visible_columns: int,
    ) -> int:
        title_start_column = 3

        title_end_column = max(
            visible_columns - 2,
            title_start_column,
        )

        for row_number in (
            1,
            2,
            3,
        ):
            sheet.merge_cells(
                start_row=row_number,
                start_column=title_start_column,
                end_row=row_number,
                end_column=title_end_column,
            )

        authority_cell = sheet.cell(
            row=1,
            column=title_start_column,
            value=self.authority_name,
        )

        self._style_cell(
            authority_cell,
            font_color=BRAND_DARK_GREEN,
            bold=True,
            font_size=14,
            border=False,
        )

        report_cell = sheet.cell(
            row=2,
            column=title_start_column,
            value=(
                f"{self.platform_name} — "
                f"{report.title}"
            ),
        )

        self._style_cell(
            report_cell,
            font_color=BRAND_GREEN,
            bold=True,
            font_size=18,
            border=False,
        )

        subtitle_cell = sheet.cell(
            row=3,
            column=title_start_column,
            value=self.platform_subtitle,
        )

        self._style_cell(
            subtitle_cell,
            font_color=TEXT_MUTED,
            bold=True,
            font_size=10,
            border=False,
        )

        self._add_brand_logos(
            sheet=sheet,
            visible_columns=visible_columns,
        )

        sheet.row_dimensions[1].height = 25
        sheet.row_dimensions[2].height = 30
        sheet.row_dimensions[3].height = 22

        metadata_row = 4

        sheet.merge_cells(
            start_row=metadata_row,
            start_column=1,
            end_row=metadata_row,
            end_column=visible_columns,
        )

        metadata_text = self._build_metadata_text(
            user=user,
            records_count=records_count,
            filters=filters,
        )

        metadata_cell = sheet.cell(
            row=metadata_row,
            column=1,
            value=metadata_text,
        )

        self._style_cell(
            metadata_cell,
            fill=BRAND_LIGHT_GOLD,
            font_color=TEXT_DARK,
            bold=True,
            font_size=9,
            border=True,
            wrap_text=True,
            horizontal="right",
        )

        sheet.row_dimensions[
            metadata_row
        ].height = 32

        return metadata_row

    # ==================================================
    # رؤوس الجدول
    # ==================================================

    def _build_table_headers(
        self,
        *,
        sheet,
        columns: tuple[ExportColumn, ...],
        header_row: int,
    ) -> None:
        for column_index, column in enumerate(
            columns,
            start=1,
        ):
            cell = sheet.cell(
                row=header_row,
                column=column_index,
                value=column.header,
            )

            self._style_cell(
                cell,
                fill=BRAND_GREEN,
                font_color=WHITE,
                bold=True,
                font_size=10,
                border=True,
                wrap_text=True,
            )

            column_width = getattr(
                column,
                "width",
                18,
            )

            try:
                resolved_width = float(
                    column_width
                )

            except (
                TypeError,
                ValueError,
            ):
                resolved_width = 18

            sheet.column_dimensions[
                get_column_letter(
                    column_index
                )
            ].width = max(
                8,
                min(
                    resolved_width,
                    80,
                ),
            )

        sheet.row_dimensions[
            header_row
        ].height = 32

    # ==================================================
    # كتابة سجل
    # ==================================================

    def _write_record_row(
        self,
        *,
        sheet,
        row_number: int,
        record,
        columns: tuple[ExportColumn, ...],
        alternate: bool,
    ) -> None:
        row_fill = (
            ROW_ALTERNATE
            if alternate
            else WHITE
        )

        for column_index, column in enumerate(
            columns,
            start=1,
        ):
            try:
                raw_value = column.get_value(
                    record
                )

            except Exception:
                raw_value = ""

            value = self._safe_excel_value(
                raw_value
            )

            cell = sheet.cell(
                row=row_number,
                column=column_index,
                value=value,
            )

            wrap_text = bool(
                getattr(
                    column,
                    "wrap_text",
                    False,
                )
            )

            self._style_cell(
                cell,
                fill=row_fill,
                font_color=TEXT_DARK,
                bold=False,
                font_size=9,
                horizontal=(
                    "right"
                    if wrap_text
                    else "center"
                ),
                wrap_text=wrap_text,
            )

            number_format = getattr(
                column,
                "number_format",
                None,
            )

            if number_format:
                cell.number_format = str(
                    number_format
                )

        sheet.row_dimensions[
            row_number
        ].height = 25

    # ==================================================
    # رسالة عدم وجود بيانات
    # ==================================================

    def _write_empty_message(
        self,
        *,
        sheet,
        row_number: int,
        columns_count: int,
    ) -> None:
        sheet.merge_cells(
            start_row=row_number,
            start_column=1,
            end_row=row_number,
            end_column=max(
                columns_count,
                1,
            ),
        )

        empty_cell = sheet.cell(
            row=row_number,
            column=1,
            value=(
                "لا توجد بيانات مطابقة "
                "للفلاتر المحددة."
            ),
        )

        self._style_cell(
            empty_cell,
            fill=BRAND_LIGHT_GREEN,
            font_color=TEXT_MUTED,
            bold=True,
            font_size=11,
            border=True,
        )

        sheet.row_dimensions[
            row_number
        ].height = 34

    # ==================================================
    # جدول Excel
    # ==================================================

    def _add_excel_table(
        self,
        *,
        sheet,
        header_row: int,
        last_data_row: int,
        columns_count: int,
        report_key: str,
    ) -> None:
        if (
            columns_count <= 0
            or last_data_row <= header_row
        ):
            return

        last_column = get_column_letter(
            columns_count
        )

        table_reference = (
            f"A{header_row}:"
            f"{last_column}{last_data_row}"
        )

        safe_report_key = "".join(
            character
            for character in report_key.title()
            if character.isalnum()
        )

        timestamp = timezone.now().strftime(
            "%H%M%S%f"
        )

        table_name = (
            f"Export{safe_report_key}"
            f"{timestamp}"
        )[:255]

        table = Table(
            displayName=table_name,
            ref=table_reference,
        )

        table.tableStyleInfo = (
            TableStyleInfo(
                name="TableStyleMedium4",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )
        )

        sheet.add_table(
            table
        )

    # ==================================================
    # صف إجمالي السجلات
    # ==================================================

    def _add_summary_row(
        self,
        *,
        sheet,
        row_number: int,
        columns_count: int,
        records_count: int,
    ) -> None:
        if columns_count <= 0:
            return

        sheet.merge_cells(
            start_row=row_number,
            start_column=1,
            end_row=row_number,
            end_column=columns_count,
        )

        summary_cell = sheet.cell(
            row=row_number,
            column=1,
            value=(
                "إجمالي السجلات المصدرة: "
                f"{records_count}"
            ),
        )

        self._style_cell(
            summary_cell,
            fill=BRAND_LIGHT_GREEN,
            font_color=BRAND_DARK_GREEN,
            bold=True,
            font_size=11,
            border=True,
        )

        sheet.row_dimensions[
            row_number
        ].height = 28

    # ==================================================
    # تنسيق الخلايا
    # ==================================================

    def _style_cell(
        self,
        cell,
        *,
        fill: str | None = None,
        font_color: str = TEXT_DARK,
        bold: bool = False,
        font_size: int = 10,
        horizontal: str = "center",
        vertical: str = "center",
        border: bool = True,
        wrap_text: bool = False,
    ) -> None:
        if fill:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=fill,
            )

        cell.font = Font(
            name="Arial",
            size=font_size,
            bold=bold,
            color=font_color,
        )

        cell.alignment = Alignment(
            horizontal=horizontal,
            vertical=vertical,
            wrap_text=wrap_text,
        )

        if border:
            side = Side(
                style="thin",
                color=BORDER_COLOR,
            )

            cell.border = Border(
                left=side,
                right=side,
                top=side,
                bottom=side,
            )

    # ==================================================
    # الشعارات
    # ==================================================

    def _add_brand_logos(
        self,
        *,
        sheet,
        visible_columns: int,
    ) -> None:
        haramain_path = self._first_existing_path(
            self._haramain_logo_candidates()
        )

        abwaab_path = self._first_existing_path(
            self._abwaab_logo_candidates()
        )

        self._add_logo(
            sheet=sheet,
            path=haramain_path,
            anchor="A1",
            width=95,
            height=62,
        )

        right_logo_column = get_column_letter(
            max(
                visible_columns - 1,
                1,
            )
        )

        self._add_logo(
            sheet=sheet,
            path=abwaab_path,
            anchor=(
                f"{right_logo_column}1"
            ),
            width=95,
            height=55,
        )

    def _haramain_logo_candidates(
        self,
    ) -> tuple[Path, ...]:
        base_dir = Path(
            settings.BASE_DIR
        )

        media_root = Path(
            settings.MEDIA_ROOT
        )

        return (
            media_root / "aharamain_logo.png",
            media_root / "aharamaian_logo.png",
            media_root / "alharamain_logo.png",
            media_root / "haramain_logo.png",
            media_root / "شعار الحرمين.png",

            base_dir
            / "static"
            / "images"
            / "aharamain_logo.png",

            base_dir
            / "static"
            / "images"
            / "haramain_logo.png",

            base_dir
            / "static"
            / "img"
            / "aharamain_logo.png",

            base_dir
            / "static"
            / "img"
            / "haramain_logo.png",

            base_dir
            / "static"
            / "img"
            / "شعار الحرمين.png",
        )

    def _abwaab_logo_candidates(
        self,
    ) -> tuple[Path, ...]:
        base_dir = Path(
            settings.BASE_DIR
        )

        media_root = Path(
            settings.MEDIA_ROOT
        )

        return (
            media_root / "abwaab-logo.jpeg",
            media_root / "abwaab-logo.jpg",
            media_root / "abwaab-logo.png",
            media_root / "abwaab_logo.png",

            base_dir
            / "static"
            / "images"
            / "abwaab-logo.jpeg",

            base_dir
            / "static"
            / "images"
            / "abwaab-logo.png",

            base_dir
            / "static"
            / "images"
            / "abwaab_logo.png",

            base_dir
            / "static"
            / "img"
            / "abwaab-logo.jpeg",

            base_dir
            / "static"
            / "img"
            / "abwaab-logo.png",
        )

    @staticmethod
    def _first_existing_path(
        paths: Iterable[Path],
    ) -> Path | None:
        for path in paths:
            if (
                path.exists()
                and path.is_file()
            ):
                return path

        return None

    @staticmethod
    def _add_logo(
        *,
        sheet,
        path: Path | None,
        anchor: str,
        width: int,
        height: int,
    ) -> None:
        if not path:
            return

        try:
            image = ExcelImage(
                str(path)
            )

            image.width = width
            image.height = height

            sheet.add_image(
                image,
                anchor,
            )

        except Exception:
            # غياب Pillow أو تلف الصورة
            # لا يجب أن يوقف التصدير.
            return

    # ==================================================
    # بيانات المستخدم والفلاتر
    # ==================================================

    def _build_metadata_text(
        self,
        *,
        user,
        records_count: int,
        filters: dict[str, Any],
    ) -> str:
        exported_at = (
            timezone.localtime()
            .strftime(
                "%Y-%m-%d %H:%M"
            )
        )

        exported_by = self._display_user(
            user
        )

        metadata_parts = [
            (
                "المستخدم: "
                f"{exported_by or '—'}"
            ),
            (
                "تاريخ التصدير: "
                f"{exported_at}"
            ),
            (
                "عدد السجلات: "
                f"{records_count}"
            ),
        ]

        filter_text = self._format_filters(
            filters
        )

        if filter_text:
            metadata_parts.append(
                f"الفلاتر: {filter_text}"
            )

        return "   |   ".join(
            metadata_parts
        )

    def _format_filters(
        self,
        filters: dict[str, Any],
    ) -> str:
        if not filters:
            return ""

        ignored_keys = {
            "csrfmiddlewaretoken",
            "page",
            "page_size",
            "preview",
            "preview_limit",
            "export_format",
            "format",
            "report_key",
            "selected_columns",
            "columns",
            "search",
            "sort",
            "direction",
            "action",
            "submit",
        }

        formatted_parts: list[str] = []

        for key, value in filters.items():
            if key in ignored_keys:
                continue

            if value in (
                None,
                "",
                [],
                (),
                {},
            ):
                continue

            label = self._filter_label(
                key
            )

            formatted_value = (
                self._display_filter_value(
                    value
                )
            )

            formatted_parts.append(
                f"{label}: {formatted_value}"
            )

        return " | ".join(
            formatted_parts
        )

    @staticmethod
    def _filter_label(
        key: str,
    ) -> str:
        labels = {
            "date_from": "من تاريخ",
            "date_to": "إلى تاريخ",
            "shift_plan": "الوردية",
            "shift_plan_id": "الوردية",
            "zone": "المنطقة",
            "zone_id": "المنطقة",
            "door_direction": "جهة الأبواب",
            "door_state": "حالة الباب",
            "incident_status": "حالة البلاغ",
            "maintenance_status": "حالة الصيانة",
            "report_status": "حالة التقرير",
            "employee": "الموظف",
            "employee_id": "الموظف",
            "technician": "الفني",
            "technician_id": "الفني",
            "priority": "الأولوية",
            "job_title": "المسمى الوظيفي",
            "work_status": "حالة الموظف",
            "is_active": "حالة التفعيل",
            "is_confirmed": "حالة التأكيد",
            "report_type": "نوع التقرير",
            "role": "الدور",
            "shift_type": "نوع الوردية",
            "shift_type_id": "نوع الوردية",
            "rest_days": "أيام الراحة",
            "incident_type": "نوع البلاغ",
            "q": "البحث",
            "search": "البحث",
        }

        return labels.get(
            key,
            key.replace(
                "_",
                " ",
            ),
        )

    @staticmethod
    def _display_filter_value(
        value: Any,
    ) -> str:
        if isinstance(
            value,
            bool,
        ):
            return (
                "نعم"
                if value
                else "لا"
            )

        if isinstance(
            value,
            (
                list,
                tuple,
                set,
            ),
        ):
            return "، ".join(
                str(item)
                for item in value
            )

        return str(
            value
        )

    @staticmethod
    def _display_user(
        user,
    ) -> str:
        if not user:
            return ""

        get_full_name = getattr(
            user,
            "get_full_name",
            None,
        )

        if callable(
            get_full_name
        ):
            full_name = (
                get_full_name()
                or ""
            ).strip()

            if full_name:
                return full_name

        employee = getattr(
            user,
            "employee",
            None,
        )

        if employee:
            employee_name = getattr(
                employee,
                "full_name",
                "",
            )

            if employee_name:
                return str(
                    employee_name
                )

        return str(
            getattr(
                user,
                "username",
                "",
            )
            or ""
        )

    # ==================================================
    # معالجة المؤشرات
    # ==================================================

    def _flatten_indicators(
        self,
        indicators: dict[str, Any],
    ) -> list[tuple[str, Any]]:
        labels = {
            "records_count": "عدد السجلات",
            "active_count": "السجلات النشطة",
            "inactive_count": "السجلات غير النشطة",
            "confirmed_count": "التسكينات المؤكدة",
            "unconfirmed_count": "التسكينات غير المؤكدة",
            "employees_count": "عدد الموظفين",
            "doors_count": "عدد الأبواب",
            "shifts_count": "عدد الورديات",
            "active_doors": "الأبواب النشطة",
            "inactive_doors": "الأبواب غير النشطة",
            "zones_count": "عدد المناطق",
            "total_doors": "إجمالي الأبواب",
            "open_doors": "الأبواب المفتوحة",
            "maintenance_requests": "طلبات الصيانة",
            "generated_at": "وقت إنشاء المؤشرات",
            "status_totals": "الإجمالي حسب الحالة",
            "priority_totals": "الإجمالي حسب الأولوية",
        }

        rows: list[
            tuple[str, Any]
        ] = []

        for key, value in indicators.items():
            label = labels.get(
                key,
                key.replace(
                    "_",
                    " ",
                ),
            )

            if isinstance(
                value,
                list,
            ):
                if not value:
                    continue

                for item in value:
                    if not isinstance(
                        item,
                        dict,
                    ):
                        rows.append(
                            (
                                label,
                                item,
                            )
                        )

                        continue

                    item_label = (
                        item.get("status")
                        or item.get("priority")
                        or item.get("label")
                        or "غير محدد"
                    )

                    item_total = (
                        item.get("total")
                        or item.get("count")
                        or 0
                    )

                    rows.append(
                        (
                            f"{label} - "
                            f"{item_label}",
                            item_total,
                        )
                    )

            elif isinstance(
                value,
                dict,
            ):
                for (
                    child_key,
                    child_value,
                ) in value.items():
                    rows.append(
                        (
                            f"{label} - "
                            f"{child_key}",
                            child_value,
                        )
                    )

            else:
                rows.append(
                    (
                        label,
                        value,
                    )
                )

        return rows

    # ==================================================
    # القيم الآمنة
    # ==================================================

    @staticmethod
    def _safe_excel_value(
        value: Any,
    ) -> Any:
        """
        تحويل القيمة إلى قيمة آمنة لملف Excel.

        يمنع Formula Injection عند بدء النص بأحد
        رموز الصيغ المعروفة.
        """

        if value is None:
            return ""

        if isinstance(
            value,
            datetime,
        ):
            if timezone.is_aware(
                value
            ):
                value = (
                    timezone.localtime(
                        value
                    )
                    .replace(
                        tzinfo=None
                    )
                )

            return value

        if isinstance(
            value,
            str,
        ):
            cleaned = (
                value
                .replace(
                    "\x00",
                    "",
                )
                .replace(
                    "\r\n",
                    "\n",
                )
                .replace(
                    "\r",
                    "\n",
                )
            )

            if cleaned.startswith(
                (
                    "=",
                    "+",
                    "-",
                    "@",
                    "\t",
                )
            ):
                return f"'{cleaned}"

            return cleaned

        if isinstance(
            value,
            (
                int,
                float,
                bool,
                date,
            ),
        ):
            return value

        if isinstance(
            value,
            (
                list,
                tuple,
                set,
            ),
        ):
            return "، ".join(
                str(item)
                for item in value
            )

        if isinstance(
            value,
            dict,
        ):
            return " | ".join(
                f"{key}: {item}"
                for key, item
                in value.items()
            )

        return str(
            value
        )

    @staticmethod
    def _normalize_filters(
        filters: Mapping[str, Any],
    ) -> dict[str, Any]:
        """
        تنظيف الفلاتر قبل تمريرها إلى selectors.
        """

        normalized: dict[str, Any] = {}

        lists_method = getattr(
            filters,
            "lists",
            None,
        )

        if callable(
            lists_method
        ):
            for key, values in lists_method():
                cleaned_values = [
                    (
                        value.strip()
                        if isinstance(
                            value,
                            str,
                        )
                        else value
                    )
                    for value in values
                    if value not in (
                        None,
                        "",
                    )
                ]

                if not cleaned_values:
                    continue

                normalized[key] = (
                    cleaned_values[0]
                    if len(
                        cleaned_values
                    ) == 1
                    else cleaned_values
                )

            return normalized

        for key, value in filters.items():
            if isinstance(
                value,
                str,
            ):
                value = value.strip()

            if value in (
                None,
                "",
                [],
                (),
                {},
            ):
                continue

            normalized[key] = value

        return normalized

    # ==================================================
    # اسم الملف
    # ==================================================

    @staticmethod
    def _build_file_name(
        report: ExportReportDefinition,
    ) -> str:
        timestamp = (
            timezone.localtime()
            .strftime(
                "%Y%m%d_%H%M%S"
            )
        )

        return (
            f"{report.filename_prefix}_"
            f"{timestamp}.xlsx"
        )

    @staticmethod
    def _ensure_excel_extension(
        file_name: str,
    ) -> str:
        """
        تنظيف اسم الملف وضمان امتداد xlsx.
        """

        normalized_name = (
            str(
                file_name
                or "export"
            )
            .replace(
                '"',
                "",
            )
            .replace(
                "\r",
                "",
            )
            .replace(
                "\n",
                "",
            )
            .strip()
        )

        if not normalized_name:
            normalized_name = "export"

        if not normalized_name.lower().endswith(
            ".xlsx"
        ):
            normalized_name += ".xlsx"

        return normalized_name


# ==================================================
# نسخة افتراضية جاهزة للاستخدام
# ==================================================

excel_export_engine = ExcelExportEngine()


# ==================================================
# دوال مختصرة
# ==================================================

def build_excel_export(
    *,
    report_key: str,
    queryset=None,
    filters: Mapping[str, Any] | None = None,
    user=None,
    indicators: dict[str, Any] | None = None,
    file_name: str | None = None,
    selected_columns: (
        str
        | Iterable[str]
        | None
    ) = None,
) -> ExcelExportResult:
    """
    إنشاء ملف Excel باستخدام المحرك الافتراضي.
    """

    return excel_export_engine.build(
        report_key=report_key,
        queryset=queryset,
        filters=filters,
        user=user,
        indicators=indicators,
        file_name=file_name,
        selected_columns=selected_columns,
    )


def build_excel_response(
    *,
    report_key: str,
    queryset=None,
    filters: Mapping[str, Any] | None = None,
    user=None,
    indicators: dict[str, Any] | None = None,
    file_name: str | None = None,
    selected_columns: (
        str
        | Iterable[str]
        | None
    ) = None,
) -> HttpResponse:
    """
    إنشاء استجابة تنزيل Excel مباشرة.
    """

    return excel_export_engine.build_response(
        report_key=report_key,
        queryset=queryset,
        filters=filters,
        user=user,
        indicators=indicators,
        file_name=file_name,
        selected_columns=selected_columns,
    )
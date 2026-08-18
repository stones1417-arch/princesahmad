from __future__ import annotations

import csv
import io
import shutil
import tempfile

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from apps.exports_center.registry import REPORT_REGISTRY
from apps.roles.models import Role, UserRole


class RegisteredReportWorkflowTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp(prefix="registered-exports-")
        self.settings_override = override_settings(MEDIA_ROOT=self.media_root)
        self.settings_override.enable()

        self.user = get_user_model().objects.create_user(
            username="registered-report-exporter"
        )
        group = Group.objects.create(name="registered-report-exporters")
        group.permissions.add(
            Permission.objects.get(
                content_type__app_label="roles",
                codename="export_report",
            )
        )
        role = Role.objects.create(
            code="registered-report-exporter",
            name="Registered report exporter",
            group=group,
            operational_section=Role.OperationalSection.ALL,
        )
        UserRole.objects.create(user=self.user, role=role)
        self.client.force_login(self.user)

    def tearDown(self):
        self.settings_override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def test_filters_and_preview_render_for_every_registered_report(self):
        for report_key in REPORT_REGISTRY:
            with self.subTest(report_key=report_key, page="filters"):
                response = self.client.get(
                    reverse(
                        "exports_center:filters",
                        kwargs={"report_key": report_key},
                    )
                )
                self.assertEqual(response.status_code, 200)

            with self.subTest(report_key=report_key, page="preview"):
                response = self.client.get(
                    reverse(
                        "exports_center:preview",
                        kwargs={"report_key": report_key},
                    )
                )
                self.assertEqual(response.status_code, 200)

    def test_all_supported_exports_return_valid_files(self):
        for report_key, report in REPORT_REGISTRY.items():
            for export_format in ("excel", "pdf", "csv"):
                if not report.supports_format(export_format):
                    continue

                with self.subTest(
                    report_key=report_key,
                    export_format=export_format,
                ):
                    response = self.client.post(
                        reverse(
                            "exports_center:export",
                            kwargs={
                                "report_key": report_key,
                                "export_format": export_format,
                            },
                        )
                    )
                    self.assertEqual(response.status_code, 200)
                    self.assertGreater(len(response.content), 0)

                    if export_format == "pdf":
                        self.assertEqual(
                            response["Content-Type"],
                            "application/pdf",
                        )
                        self.assertTrue(response.content.startswith(b"%PDF"))
                    elif export_format == "excel":
                        workbook = load_workbook(
                            io.BytesIO(response.content),
                            read_only=True,
                        )
                        self.assertTrue(workbook.sheetnames)
                    else:
                        self.assertIn("text/csv", response["Content-Type"])
                        rows = list(
                            csv.reader(
                                io.StringIO(
                                    response.content.decode("utf-8-sig")
                                )
                            )
                        )
                        self.assertTrue(rows)

    def test_rendered_export_actions_are_post_forms_without_get_links(self):
        for report_key in REPORT_REGISTRY:
            response = self.client.get(
                reverse(
                    "exports_center:filters",
                    kwargs={"report_key": report_key},
                )
            )
            html = response.content.decode()

            for export_format in ("excel", "pdf", "csv"):
                export_path = reverse(
                    "exports_center:export",
                    kwargs={
                        "report_key": report_key,
                        "export_format": export_format,
                    },
                )
                self.assertIn('method="post"', html)
                self.assertIn(f'action="{export_path}"', html)
                self.assertIn("csrfmiddlewaretoken", html)
                self.assertNotIn(f'href="{export_path}"', html)

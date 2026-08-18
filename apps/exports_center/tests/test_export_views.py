from __future__ import annotations

import shutil
import tempfile
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.test import TestCase, override_settings
from django.urls import reverse
from django.http import QueryDict
from openpyxl import load_workbook

from apps.exports_center.models import ExportLog
from apps.hr.models import Employee
from apps.roles.models import Role, UserRole
from apps.exports_center.services.export_service import (
    ExportService,
    ExportServiceError,
)


class ExportWorkflowTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp(prefix="exports-center-tests-")
        self.override = override_settings(MEDIA_ROOT=self.media_root)
        self.override.enable()
        self.user = get_user_model().objects.create_user(
            username="exports_workflow_user",
            password="StrongPassword123!",
        )
        self._grant_export_access(self.user, "exports-workflow-role")
        self.client.force_login(self.user)

    def tearDown(self):
        self.override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def test_csv_export_is_logged_and_stored(self):
        response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "csv"},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        export_log = ExportLog.objects.get()
        self.assertEqual(export_log.status, ExportLog.ExportStatus.SUCCESS)
        self.assertEqual(response["X-Export-Log-ID"], str(export_log.pk))
        self.assertTrue(export_log.file.storage.exists(export_log.file.name))

    def test_pdf_ui_post_contract_returns_valid_pdf(self):
        response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "pdf"},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertGreater(len(response.content), 4)
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_excel_ui_post_contract_returns_openable_workbook(self):
        response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "excel"},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertGreaterEqual(len(workbook.sheetnames), 1)

    def test_pdf_get_is_rejected_because_ui_uses_post(self):
        response = self.client.get(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "pdf"},
            )
        )

        self.assertEqual(response.status_code, 405)
        self.assertEqual(response["Allow"], "POST")

    def test_stored_export_can_be_downloaded_again(self):
        first_response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "csv"},
            )
        )
        export_log = ExportLog.objects.get(pk=first_response["X-Export-Log-ID"])

        response = self.client.get(
            reverse(
                "exports_center:download-export",
                kwargs={"export_log_id": export_log.pk},
            )
        )

        self.assertEqual(response.status_code, 200)
        export_log.refresh_from_db()
        self.assertEqual(export_log.download_count, 1)

    def test_section_filter_is_persisted_in_export_log(self):
        response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "door_distribution", "export_format": "csv"},
            ),
            {
                "section": "female",
            },
        )

        self.assertEqual(response.status_code, 200)
        export_log = ExportLog.objects.get()
        self.assertEqual(
            export_log.filters.get("section"),
            "female",
        )

    def test_employees_operational_section_filter_is_canonicalized_to_section(self):
        response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "csv"},
            ),
            {
                "operational_section": "female",
            },
        )

        self.assertEqual(response.status_code, 200)
        export_log = ExportLog.objects.get()
        self.assertEqual(
            export_log.filters.get("operational_section"),
            "female",
        )
        self.assertEqual(export_log.filters.get("section"), "female")

    def test_csv_export_uses_selected_columns_without_logging_them_as_filters(self):
        response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "csv"},
            ),
                {"selected_columns": ["full_name"]},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.content.decode("utf-8-sig").splitlines()[0],
            "الاسم الكامل",
        )

        export_log = ExportLog.objects.get()
        self.assertNotIn("selected_columns", export_log.filters)

    def test_submitted_csv_export_preserves_section_and_selected_columns(self):
        response = self.client.post(
            reverse("exports_center:export-submit"),
            {
                "report_key": "employees",
                "export_format": "csv",
                "section": "female",
                "selected_columns": ["full_name"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.content.decode("utf-8-sig").splitlines()[0],
            "الاسم الكامل",
        )

        export_log = ExportLog.objects.get()
        self.assertEqual(export_log.filters.get("section"), "female")
        self.assertNotIn("selected_columns", export_log.filters)

    @override_settings(ASYNC_EXPORTS_ENABLED=True)
    def test_async_export_is_persisted_then_dispatched_after_commit(self):
        with patch(
            "apps.exports_center.views.build_export_file_task.delay"
        ) as mocked_delay:
            with self.captureOnCommitCallbacks(execute=True):
                response = self.client.post(
                    reverse(
                        "exports_center:export",
                        kwargs={
                            "report_key": "employees",
                            "export_format": "csv",
                        },
                    ),
                        {"selected_columns": ["full_name"]},
                )

        export_log = ExportLog.objects.get()
        self.assertRedirects(response, reverse("exports_center:logs"))
        self.assertEqual(export_log.status, ExportLog.ExportStatus.PENDING)
        self.assertEqual(export_log.metadata["selected_columns"], ["full_name"])
        mocked_delay.assert_called_once_with(export_log.pk)

    def test_normalize_filters_canonicalizes_querydict_section(self):
        filters = ExportService().normalize_filters(
            QueryDict("section=FEMALE&status=open&status=closed")
        )

        self.assertEqual(filters["section"], "female")
        self.assertEqual(filters["status"], ["open", "closed"])

    def test_normalize_filters_rejects_invalid_section(self):
        with self.assertRaises(ExportServiceError):
            ExportService().normalize_filters(
                {"section": "shared"}
            )

    def test_unsupported_format_creates_failed_audit_entry(self):
        response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "random"},
            )
        )

        self.assertEqual(response.status_code, 302)
        export_log = ExportLog.objects.get()
        self.assertEqual(export_log.status, ExportLog.ExportStatus.FAILED)
        self.assertTrue(export_log.error_message)

    def test_get_export_generation_is_rejected(self):
        response = self.client.get(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "csv"},
            )
        )
        self.assertEqual(response.status_code, 405)

    def test_export_requires_institutional_permission(self):
        user = get_user_model().objects.create_user(username="no-export-permission")
        self.client.force_login(user)
        response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "csv"},
            )
        )
        self.assertEqual(response.status_code, 403)

    def test_user_cannot_download_another_users_export(self):
        response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "csv"},
            )
        )
        export_log = ExportLog.objects.get(pk=response["X-Export-Log-ID"])
        other_user = get_user_model().objects.create_user(username="another-export-user")
        self._grant_export_access(other_user, "another-exports-role")
        self.client.force_login(other_user)
        response = self.client.get(
            reverse("exports_center:download-export", kwargs={"export_log_id": export_log.pk})
        )
        self.assertEqual(response.status_code, 404)

    def test_section_filter_prevents_cross_section_employee_leakage(self):
        Employee.objects.create(full_name="Male Employee", employee_number="export-male", operational_section="male")
        Employee.objects.create(full_name="Female Employee", employee_number="export-female", operational_section="female")
        response = self.client.post(
            reverse(
                "exports_center:export",
                kwargs={"report_key": "employees", "export_format": "csv"},
            ),
            {"section": "male"},
        )
        content = response.content.decode("utf-8-sig")
        self.assertIn("Male Employee", content)
        self.assertNotIn("Female Employee", content)

    def _grant_export_access(self, user, role_code):
        group = Group.objects.create(name=role_code)
        group.permissions.add(
            Permission.objects.get(
                content_type__app_label="roles",
                codename="export_report",
            )
        )
        role = Role.objects.create(
            code=role_code,
            name=role_code,
            group=group,
            operational_section=Role.OperationalSection.ALL,
        )
        UserRole.objects.create(user=user, role=role)

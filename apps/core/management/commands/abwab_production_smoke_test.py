from __future__ import annotations

import importlib
import uuid
from contextlib import ExitStack
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction
from django.db.migrations.executor import MigrationExecutor
from django.test import RequestFactory
from openpyxl import load_workbook

from apps.breaks.models import Break
from apps.distribution.models import DoorAssignment
from apps.distribution.services import DistributionService
from apps.exports_center.models import ExportLog
from apps.exports_center.services.shift_excel_exporter import export_shift_excel_response
from apps.exports_center.services.shift_pdf_exporter import export_shift_pdf_response
from apps.hr.models import Employee
from apps.locations.models import Door
from apps.ops.models import DoorShift
from apps.reporting.models import ShiftReport
from apps.reporting.services import ReportService
from apps.scheduling.models import ShiftAssignment, ShiftPlan, ShiftType
from apps.scheduling.services import activate_shift, finish_shift

MARKER = f"RENDER_SMOKE_{uuid.uuid4().hex[:7].upper()}"
REQUIRED_CODES = {"5", "6A", "6B", "7", "8", "9"}


class SmokeFailure(RuntimeError):
    pass


class Command(BaseCommand):
    help = "Production-safe smoke test for the shift lifecycle, exports, and rollback integrity."

    def add_arguments(self, parser):
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--execute", action="store_true")
        parser.add_argument("--strict", action="store_true")

    @staticmethod
    def _safe_message(error: object) -> str:
        return " ".join(str(error).split())[:500]

    @staticmethod
    def _database_ready() -> tuple[bool, str]:
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
                cursor.fetchone()
        except Exception as error:  # pragma: no cover - defensive production guard
            return False, f"database connection failed: {error}"
        return True, "database reachable"

    @staticmethod
    def _migrations_ready() -> tuple[bool, str]:
        try:
            executor = MigrationExecutor(connection)
            pending = executor.migration_plan(executor.loader.graph.leaf_nodes())
        except Exception as error:  # pragma: no cover - defensive production guard
            return False, f"migration check failed: {error}"
        if pending:
            return False, f"pending migrations: {len(pending)}"
        return True, "all migrations applied"

    @staticmethod
    def _snapshot_baseline() -> dict[str, int]:
        return {
            "User": get_user_model().objects.count(),
            "Employee": Employee.objects.count(),
            "ShiftType": ShiftType.objects.count(),
            "ShiftPlan": ShiftPlan.objects.count(),
            "DoorShift": DoorShift.objects.count(),
            "DoorAssignment": DoorAssignment.objects.count(),
            "ShiftAssignment": ShiftAssignment.objects.count(),
            "Break": Break.objects.count(),
            "ShiftReport": ShiftReport.objects.count(),
            "ExportLog": ExportLog.objects.count(),
            "active_shift_plan": ShiftPlan.objects.filter(is_active=True).count(),
            "active_door_shift": DoorShift.objects.filter(is_active=True).count(),
        }

    @staticmethod
    def _marker_residue(marker: str) -> int:
        return sum(
            (
                get_user_model().objects.filter(username__contains=marker).count(),
                Employee.objects.filter(employee_number__contains=marker).count(),
                ShiftType.objects.filter(name__contains=marker).count(),
                ShiftPlan.objects.filter(notes__contains=marker).count(),
                DoorShift.objects.filter(notes__contains=marker).count(),
                DoorAssignment.objects.filter(notes__contains=marker).count(),
                ShiftAssignment.objects.filter(notes__contains=marker).count(),
                Break.objects.filter(notes__contains=marker).count(),
                ShiftReport.objects.filter(summary__contains=marker).count(),
                ExportLog.objects.filter(report_name__contains=marker).count(),
            )
        )

    @staticmethod
    def _master_doors() -> list[Door]:
        return list(
            Door.objects.filter(is_active=True).order_by("sort_order", "door_number")
        )

    @classmethod
    def _validate_master_doors(cls, master_doors) -> tuple[bool, str]:
        codes = [door.door_number for door in master_doors]
        problems = []
        if len(codes) != 42:
            problems.append(f"active count={len(codes)}")
        if len(codes) != len(set(codes)):
            problems.append("duplicate codes")
        missing_required = sorted(REQUIRED_CODES - set(codes))
        if missing_required:
            problems.append(f"missing required={missing_required}")
        if "6" in codes:
            problems.append("plain door 6 exists")
        return not problems, "; ".join(problems) or "42-door master catalog valid"

    @staticmethod
    def _import_readiness() -> bool:
        for module_name in (
            "apps.distribution.services",
            "apps.reporting.services",
            "apps.exports_center.services.shift_excel_exporter",
            "apps.exports_center.services.shift_pdf_exporter",
            "apps.scheduling.services",
        ):
            importlib.import_module(module_name)
        return True

    def _write_stage(self, stage: str, passed: bool, message: str = "") -> None:
        suffix = f" {message}" if message else ""
        self.stdout.write(f"{stage}={'PASS' if passed else 'FAIL'}{suffix}")

    def _require(self, stage: str, condition: bool, message: str) -> None:
        self.current_stage = stage
        if not condition:
            raise SmokeFailure(message)

    def _write_failure(self, failure: dict[str, str]) -> None:
        self.stdout.write("FULL_SHIFT_LIFECYCLE=FAIL")
        self.stdout.write(f"FAILED_STAGE={failure['stage']}")
        self.stdout.write(f"ERROR_TYPE={failure['type']}")
        self.stdout.write(f"ERROR={failure['error']}")

    def _audit_rollback(self, baseline: dict[str, int]) -> tuple[bool, int]:
        after = self._snapshot_baseline()
        baseline_ok = after == baseline
        self._write_stage(
            "POST_ROLLBACK_BASELINE",
            baseline_ok,
            f"before={baseline}; after={after}",
        )
        residue = self._marker_residue(MARKER)
        self._write_stage(
            "POST_ROLLBACK_RESIDUE",
            residue == 0,
            f"marker residue count={residue}",
        )
        self.stdout.write(f"DATABASE_RESIDUE={residue}")
        if baseline_ok and residue == 0:
            self.stdout.write("ROLLBACK_INTEGRITY=PASS")
        return baseline_ok, residue

    def _dry_run(self) -> None:
        self.stdout.write("ABWAB_PRODUCTION_SMOKE_TEST")
        self.stdout.write(f"MARKER={MARKER}")
        self.stdout.write("SAFE_MODE=READ_ONLY")
        database_ok, message = self._database_ready()
        self._write_stage("PRECHECK", database_ok, message)
        if not database_ok:
            return
        migrations_ok, message = self._migrations_ready()
        self._write_stage("MIGRATIONS", migrations_ok, message)
        if not migrations_ok:
            return
        baseline = self._snapshot_baseline()
        self._write_stage("BASELINE", True, f"counts={baseline}")
        doors_ok, message = self._validate_master_doors(self._master_doors())
        self._write_stage("MASTER_DOORS", doors_ok, message)
        active_ok = not baseline["active_shift_plan"] and not baseline["active_door_shift"]
        self._write_stage("ACTIVE_SHIFT_GATE", active_ok, "read-only active-state check")
        self._write_stage("IMPORT_READY", self._import_readiness())
        self.stdout.write("DRY_RUN_READ_ONLY=PASS")
        self.stdout.write(
            f"SMOKE_TEST_RESULT={'PASS' if doors_ok and active_ok else 'FAIL'}"
        )

    @staticmethod
    def _mock_file_save(field_file, name, content, save=False):
        del content, save
        field_file.name = f"smoke-only/{name}"

    def _external_mocks(self) -> ExitStack:
        stack = ExitStack()
        self.storage_save_mock = stack.enter_context(
            patch(
                "django.db.models.fields.files.FieldFile.save",
                autospec=True,
                side_effect=self._mock_file_save,
            )
        )
        stack.enter_context(patch("django.core.mail.send_mail"))
        stack.enter_context(patch("django.core.mail.backends.smtp.EmailBackend.send_messages"))
        stack.enter_context(patch("apps.distribution.services._schedule_assignment_notification"))
        stack.enter_context(patch("apps.distribution.services.dispatch_assignment_message"))
        stack.enter_context(
            patch(
                "apps.communications.services.assignment_message_service."
                "AssignmentMessageService.dispatch_assignment_message"
            )
        )
        stack.enter_context(patch("apps.core.tasks.send_sms_task.delay"))
        stack.enter_context(patch("apps.core.tasks.send_email_task.delay"))
        stack.enter_context(
            patch(
                "apps.communications.providers.authentica."
                "AuthenticaProvider.send_operational_sms"
            )
        )
        stack.enter_context(
            patch(
                "apps.communications.providers.authentica."
                "AuthenticaProvider.send_operational_whatsapp"
            )
        )
        return stack

    @staticmethod
    def _employee(suffix: str, section: str, phone_suffix: str) -> Employee:
        return Employee.objects.create(
            full_name=f"{MARKER}_{suffix}",
            employee_number=f"{MARKER[:18]}{phone_suffix}",
            operational_section=section,
            work_status=Employee.WorkStatus.ACTIVE,
            can_work_on_doors=True,
            is_active=True,
            phone_number=f"05000000{phone_suffix}",
        )

    @staticmethod
    def _build_test_shift(operator) -> ShiftPlan:
        shift_type = ShiftType.objects.create(
            name=f"{MARKER}_SHIFT_TYPE",
            start_time="08:00:00",
            end_time="16:00:00",
            is_active=True,
            ordering=1,
        )
        return ShiftPlan.objects.create(
            shift_type=shift_type,
            date=__import__("django.utils.timezone").utils.timezone.localdate(),
            start_time="08:00:00",
            end_time="16:00:00",
            is_active=False,
            is_finished=False,
            created_by=operator,
            notes=MARKER,
        )

    def _assignment(self, shift, employee, door, section, operator):
        return DistributionService.create_assignment(
            shift_plan=shift,
            employee=employee,
            door=door,
            role=DoorAssignment.Role.MONITOR,
            section=section,
            assigned_by=operator,
            notes=MARKER,
        )

    def _run_lifecycle(self, operator, master_doors) -> None:
        self.current_stage = "TEST_SHIFT"
        shift = self._build_test_shift(operator)
        self._write_stage("TEST_SHIFT", True)

        self.current_stage = "ACTIVATE_SHIFT"
        shift = activate_shift(shift)
        self._require("ACTIVATE_SHIFT", shift.is_active and not shift.is_finished, "shift activation failed")
        self._write_stage("SHIFT_ACTIVATE", True)

        self.current_stage = "42_DOORS"
        door_shifts = list(
            DoorShift.objects.filter(shift_plan=shift, is_active=True).order_by(
                "sort_order", "door_number"
            )
        )
        expected_codes = [door.door_number for door in master_doors]
        actual_codes = [door_shift.door_number for door_shift in door_shifts]
        missing = sorted(set(expected_codes) - set(actual_codes))
        extra = sorted(set(actual_codes) - set(expected_codes))
        duplicates = len(actual_codes) - len(set(actual_codes))
        self._require(
            "42_DOORS",
            len(door_shifts) == 42
            and actual_codes == expected_codes
            and not missing
            and not extra
            and duplicates == 0,
            f"missing={missing}; extra={extra}; duplicates={duplicates}",
        )
        master_by_code = {door.door_number: door for door in master_doors}
        self._require(
            "42_DOORS",
            all(
                door_shift.sort_order
                == master_by_code[door_shift.door_number].sort_order
                for door_shift in door_shifts
            ),
            "DoorShift sort_order differs from master Door",
        )
        self._write_stage("42_DOORS", True)
        self._write_stage("6A_6B", True)

        male_door = Door.objects.get(door_number="5", is_active=True)
        female_door = Door.objects.get(door_number="12", is_active=True)
        self.current_stage = "DISTRIBUTION"
        male = self._employee("male", Employee.OperationalSection.MALE, "01")
        female = self._employee("female", Employee.OperationalSection.FEMALE, "02")
        male_assignment = self._assignment(
            shift, male, male_door, DoorAssignment.AssignmentSection.MALE, operator
        )
        self._require("DISTRIBUTION", male_assignment.section == "male", "male assignment failed")
        self._write_stage("MALE_ASSIGNMENT", True)
        female_assignment = self._assignment(
            shift,
            female,
            female_door,
            DoorAssignment.AssignmentSection.FEMALE,
            operator,
        )
        self._require("DISTRIBUTION", female_assignment.section == "female", "female assignment failed")
        self._write_stage("FEMALE_ASSIGNMENT", True)

        self.current_stage = "CROSS_SECTION"
        cross = self._employee("cross", Employee.OperationalSection.MALE, "03")
        try:
            self._assignment(
                shift,
                cross,
                female_door,
                DoorAssignment.AssignmentSection.MALE,
                operator,
            )
        except ValidationError:
            self._write_stage("CROSS_SECTION_REJECTED", True)
        else:
            raise SmokeFailure("cross-section assignment unexpectedly succeeded")

        self.current_stage = "DUPLICATE"
        duplicate = self._employee("duplicate", Employee.OperationalSection.MALE, "04")
        self._assignment(
            shift, duplicate, male_door, DoorAssignment.AssignmentSection.MALE, operator
        )
        try:
            self._assignment(
                shift,
                duplicate,
                Door.objects.get(door_number="7", is_active=True),
                DoorAssignment.AssignmentSection.MALE,
                operator,
            )
        except ValidationError:
            self._write_stage("DUPLICATE_PROTECTION", True)
        else:
            raise SmokeFailure("duplicate assignment unexpectedly succeeded")

        self.current_stage = "BREAK"
        break_employee = self._employee("break", Employee.OperationalSection.MALE, "05")
        rest_days = {
            0: Break.RestDays.MONDAY_TUESDAY,
            1: Break.RestDays.TUESDAY_WEDNESDAY,
            2: Break.RestDays.WEDNESDAY_THURSDAY,
            3: Break.RestDays.THURSDAY_FRIDAY,
            4: Break.RestDays.FRIDAY_SATURDAY,
            5: Break.RestDays.SATURDAY_SUNDAY,
            6: Break.RestDays.SUNDAY_MONDAY,
        }[shift.date.weekday()]
        break_record = Break.objects.create(
            employee=break_employee,
            shift_type=shift.shift_type,
            job_title=Break.BreakJobTitle.MONITOR,
            rest_days=rest_days,
            is_active=True,
            notes=MARKER,
        )
        self._require(
            "BREAK",
            DistributionService.employee_is_on_break(
                employee=break_employee, shift_plan=shift
            ),
            "break did not start",
        )
        self._write_stage("BREAK_START", True)
        try:
            self._assignment(
                shift,
                break_employee,
                male_door,
                DoorAssignment.AssignmentSection.MALE,
                operator,
            )
        except ValidationError:
            self._write_stage("BREAK_ASSIGNMENT_BLOCK", True)
        else:
            raise SmokeFailure("assignment during break unexpectedly succeeded")
        break_record.is_active = False
        break_record.save(update_fields=["is_active", "notes"])
        self._require(
            "BREAK",
            not DistributionService.employee_is_on_break(
                employee=break_employee, shift_plan=shift
            ),
            "break did not end",
        )
        final_assignment = self._assignment(
            shift,
            break_employee,
            male_door,
            DoorAssignment.AssignmentSection.MALE,
            operator,
        )
        self._require(
            "BREAK", final_assignment.employee_id == break_employee.pk, "employee changed"
        )
        self._write_stage("BREAK_END", True)

        self.current_stage = "SHIFT_FINISH"
        finished = finish_shift(shift)
        self._require(
            "SHIFT_FINISH",
            not finished.is_active
            and finished.is_finished
            and finished.finished_at is not None
            and not DoorShift.objects.filter(shift_plan=finished, is_active=True).exists(),
            "shift did not finish cleanly",
        )
        self._write_stage("SHIFT_FINISH", True)

        self.current_stage = "REPORT"
        report = ReportService.generate_shift_report(shift_plan=finished, user=operator)
        self._require(
            "REPORT",
            report.shift_plan_id == finished.pk and report.total_doors == 42,
            "report does not match test shift",
        )
        self._write_stage("REPORT", True)

        request = RequestFactory().get("/exports/shift/")
        request.user = operator
        self.current_stage = "EXCEL_EXPORT"
        excel = export_shift_excel_response(request, finished.pk, "all")
        self._require(
            "EXCEL_EXPORT", excel.status_code == 200 and bool(excel.content), "empty Excel"
        )
        workbook = load_workbook(BytesIO(excel.content))
        self._require(
            "EXCEL_EXPORT",
            bool(workbook.sheetnames) and workbook.active.max_row > 0,
            "invalid Excel workbook",
        )
        self._write_stage("EXCEL_EXPORT", True)

        self.current_stage = "PDF_EXPORT"
        pdf = export_shift_pdf_response(request, finished.pk, "all")
        self._require(
            "PDF_EXPORT",
            pdf.status_code == 200 and bool(pdf.content) and pdf.content[:4] == b"%PDF",
            "invalid PDF export",
        )
        self._require(
            "PDF_EXPORT",
            self.storage_save_mock.call_count >= 2,
            "export archive path did not pass through mocked storage",
        )
        self._write_stage("PDF_EXPORT", True)

    def _execute_lifecycle(self) -> None:
        self.stdout.write("ABWAB_PRODUCTION_SMOKE_TEST")
        self.stdout.write(f"MARKER={MARKER}")
        self.stdout.write("SAFE_MODE=EXECUTE")
        self.current_stage = "PRECHECK"
        database_ok, message = self._database_ready()
        self._write_stage("PRECHECK", database_ok, message)
        if not database_ok:
            self._write_failure(
                {"stage": "PRECHECK", "type": "DATABASE_UNAVAILABLE", "error": message}
            )
            raise CommandError(message)

        self.current_stage = "MIGRATIONS"
        migrations_ok, message = self._migrations_ready()
        self._write_stage("MIGRATIONS", migrations_ok, message)
        if not migrations_ok:
            self._write_failure(
                {"stage": "MIGRATIONS", "type": "PENDING_MIGRATIONS", "error": message}
            )
            raise CommandError(message)

        baseline = self._snapshot_baseline()
        self._write_stage("BASELINE", True, f"counts={baseline}")
        self.current_stage = "MASTER_DOORS"
        master_doors = self._master_doors()
        doors_ok, message = self._validate_master_doors(master_doors)
        self._write_stage("MASTER_DOORS", doors_ok, message)
        if not doors_ok:
            self._write_failure(
                {"stage": "MASTER_DOORS", "type": "MASTER_DOOR_INTEGRITY", "error": message}
            )
            self._audit_rollback(baseline)
            raise CommandError(message)

        self.current_stage = "BASELINE"
        if baseline["active_shift_plan"] or baseline["active_door_shift"]:
            message = (
                f"active ShiftPlan={baseline['active_shift_plan']}; "
                f"active DoorShift={baseline['active_door_shift']}"
            )
            self._write_stage("BASELINE", False, message)
            self._write_failure(
                {"stage": "BASELINE", "type": "ACTIVE_PRODUCTION_STATE", "error": message}
            )
            self._audit_rollback(baseline)
            raise CommandError(message)

        failure = None
        lifecycle_success = False
        with self._external_mocks():
            with transaction.atomic():
                try:
                    operator = get_user_model().objects.create_user(
                        username=f"{MARKER}_operator",
                        email=f"{MARKER.lower()}@example.test",
                        password="SmokeOnlyPassword123!",
                        is_staff=True,
                        is_active=True,
                    )
                    self.current_stage = "TEST_SHIFT"
                    self._run_lifecycle(operator, master_doors)
                    lifecycle_success = True
                except Exception as error:  # audited after forced rollback
                    failure = {
                        "stage": self.current_stage,
                        "type": type(error).__name__,
                        "error": self._safe_message(error),
                    }
                finally:
                    transaction.set_rollback(True)

        baseline_ok, residue = self._audit_rollback(baseline)
        if not baseline_ok:
            failure = {
                "stage": "POST_ROLLBACK_BASELINE",
                "type": "BASELINE_DRIFT",
                "error": "database counts differ after rollback",
            }
        elif residue:
            failure = {
                "stage": "POST_ROLLBACK_RESIDUE",
                "type": "MARKER_RESIDUE",
                "error": f"marker residue count={residue}",
            }

        if failure or not lifecycle_success:
            self._write_failure(
                failure
                or {
                    "stage": self.current_stage,
                    "type": "SmokeFailure",
                    "error": "lifecycle did not complete",
                }
            )
            raise CommandError("Production smoke lifecycle failed safely.")

        self.stdout.write("NO_REAL_EMAIL=PASS")
        self.stdout.write("NO_REAL_SMS=PASS")
        self.stdout.write("REAL_STORAGE_UPLOAD=NO")
        self.stdout.write("FULL_SHIFT_LIFECYCLE=PASS")
        self.stdout.write("READY_FOR_NEXT_RELEASE_STAGE=YES")
        self.stdout.write("SMOKE_TEST_RESULT=PASS")

    def handle(self, *args, **options):
        dry_run = bool(options.get("dry_run"))
        execute = bool(options.get("execute"))
        if dry_run and execute:
            raise CommandError("Choose either --dry-run or --execute, not both.")
        if dry_run or not execute:
            self._dry_run()
            return
        self._execute_lifecycle()

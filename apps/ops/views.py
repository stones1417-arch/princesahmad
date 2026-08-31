from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Case, Count, IntegerField, Prefetch, Q, Value, When
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.views.decorators.http import require_GET, require_POST
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.locations.models import Door

from apps.dashboard.models import SystemActivityLog
from apps.audit.models import DoorStateHistory
from apps.scheduling.models import ShiftAssignment, ShiftOperationalLeadership, ShiftPlan
from apps.scheduling.operational_leadership_service import leadership_for_shift
from apps.breaks.models import Break
from apps.distribution.models import DoorAssignment
from apps.roles.services.section_access import (
    filter_doors_for_user,
    get_allowed_sections,
    has_institutional_scope,
)
from apps.roles.services.access_control import user_has_permission, user_has_role
from apps.roles.services.permission_registry import PlatformPermissions

from .command_center_service import CommandCenterService
from .activity_logger import record_live_operation
from .door_service import DoorService
from .engineering_center_service import EngineeringCenterService
from .engineering_incident_followup_service import EngineeringIncidentFollowupService
from .incident_service import IncidentService
from .incident_routing_service import IncidentRoutingService
from .maintenance_service import MaintenanceService
from .models import (
    DoorCurrentState,
    DoorOperationalProfile,
    DoorShift,
    Incident,
    IncidentSupervisoryAction,
    MaintenanceRequest,
    LeadershipDelegation,
)
from .supervisory_leadership_service import SupervisoryLeadershipService
from .operations_center_service import OperationsCenterService


LIVE_OPERATION_MODULES = (
    "الأبواب",
    "الصيانة",
    "البلاغات",
    "توزيع الأبواب",
)


@login_required
def supervisory_command_center_view(request, center="department", pk=None):
    role_centers = []
    if request.user.is_superuser or user_has_role(request.user, SupervisoryLeadershipService.HEAD):
        role_centers.append(("department", "قيادة قسم الأبواب", "ops:department-command-center"))
    if user_has_role(request.user, SupervisoryLeadershipService.DEPUTY):
        role_centers.append(("department", "قيادة قسم الأبواب", "ops:department-command-center"))
    if user_has_role(request.user, SupervisoryLeadershipService.SENIOR_ADMIN):
        role_centers.append(("administrative", "المتابعة الإدارية", "ops:administrative-command-center"))
    if user_has_role(request.user, SupervisoryLeadershipService.GENERAL_MANAGER):
        role_centers.append(("executive", "القيادة التنفيذية", "ops:executive-command-center"))
    role_centers = list(dict.fromkeys(role_centers))
    if not role_centers:
        raise PermissionDenied
    requested_detail_center = str(request.GET.get("center") or "").strip()
    active_center = (
        requested_detail_center
        if center == "detail" and requested_detail_center in {item[0] for item in role_centers}
        else role_centers[0][0] if center == "detail" else center
    )
    if active_center not in {item[0] for item in role_centers}:
        raise PermissionDenied

    incidents = SupervisoryLeadershipService.visible_incidents(request.user)
    selected_section = str(request.GET.get("section") or "").strip()
    if selected_section in {"male", "female"}:
        incidents = incidents.filter(section=selected_section)
    selected_status = str(request.GET.get("status") or "open").strip()
    if selected_status == "open":
        incidents = incidents.exclude(status__in=(Incident.Status.RESOLVED, Incident.Status.CLOSED))
    elif selected_status in Incident.Status.values:
        incidents = incidents.filter(status=selected_status)
    selected_incident = None
    if pk is not None:
        selected_incident = get_object_or_404(incidents, pk=pk)
    ordered_incidents = incidents.order_by("-updated_at", "-pk")
    page = Paginator(ordered_incidents, 30).get_page(request.GET.get("page"))
    action_choices = []
    probe_section = selected_incident.section if selected_incident else selected_section
    if not probe_section:
        probe_section = next(iter(sorted(get_allowed_sections(request.user))), "")
    role, delegation = SupervisoryLeadershipService.authority(request.user, probe_section)
    if role:
        allowed = SupervisoryLeadershipService.ROLE_ACTIONS.get(role, set())
        action_choices = [
            choice for choice in IncidentSupervisoryAction.ActionType.choices
            if choice[0] in allowed
        ]
    display_role = next((code for code in (
        SupervisoryLeadershipService.GENERAL_MANAGER,
        SupervisoryLeadershipService.HEAD,
        SupervisoryLeadershipService.SENIOR_ADMIN,
        SupervisoryLeadershipService.DEPUTY,
    ) if user_has_role(request.user, code)), SupervisoryLeadershipService.HEAD if request.user.is_superuser else "")
    role_labels = {
        SupervisoryLeadershipService.HEAD: "رئيس قسم الأبواب",
        SupervisoryLeadershipService.DEPUTY: "وكيل رئيس قسم الأبواب",
        SupervisoryLeadershipService.SENIOR_ADMIN: "كبير الإداريين",
        SupervisoryLeadershipService.GENERAL_MANAGER: "المدير العام",
    }
    natural_centers = {
        SupervisoryLeadershipService.HEAD: "department",
        SupervisoryLeadershipService.DEPUTY: "department",
        SupervisoryLeadershipService.SENIOR_ADMIN: "administrative",
        SupervisoryLeadershipService.GENERAL_MANAGER: "executive",
    }
    actual_role_label = role_labels.get(display_role, "عرض مؤسسي")
    is_deputy = display_role == SupervisoryLeadershipService.DEPUTY
    is_delegated_acting = is_deputy and delegation is not None
    is_cross_center_oversight = (
        natural_centers.get(display_role) != active_center
        or (is_deputy and not is_delegated_acting)
    )
    center_config = {
        "department": {
            "title": "مركز قيادة قسم الأبواب",
            "subtitle": "متابعة الحالات المصعّدة واتخاذ القرارات والتوجيهات الإشرافية.",
            "queue_title": "تحتاج قرارك",
            "empty_title": "لا توجد حالات تحتاج قرارك حاليًا.",
            "empty_text": "ستظهر هنا الحالات المصعّدة أو التي تنتظر إجراءً إشرافيًا.",
        },
        "administrative": {
            "title": "مركز المتابعة الإدارية",
            "subtitle": "متابعة الحالات والتحديثات والتنبيهات والتوجيهات الإدارية.",
            "queue_title": "المتابعات المفتوحة",
            "empty_title": "لا توجد متابعات إدارية مفتوحة.",
            "empty_text": "ستظهر هنا طلبات التحديث والتنبيهات الإدارية المفتوحة.",
        },
        "executive": {
            "title": "مركز القيادة التنفيذية",
            "subtitle": "مراجعة الحالات المصعّدة واتخاذ القرارات والتوجيهات التنفيذية.",
            "queue_title": "قرارات تنتظر المدير العام",
            "empty_title": "لا توجد قرارات تنفيذية معلّقة.",
            "empty_text": "ستظهر هنا الحالات المصعّدة إلى القيادة التنفيذية.",
        },
    }[active_center]
    attention_queue = SupervisoryLeadershipService.center_attention_queue(
        request.user, active_center
    )
    action_rows = IncidentSupervisoryAction.objects.filter(
        incident__in=incidents,
    )
    open_action_rows = action_rows.filter(status__in=(
        IncidentSupervisoryAction.Status.OPEN,
        IncidentSupervisoryAction.Status.ANSWERED,
        IncidentSupervisoryAction.Status.ACKNOWLEDGED,
    ))
    if active_center == "department":
        kpis = [
            ("قرار", len(attention_queue), "تحتاج قرارك", "حالات تنتظر تدخلك"),
            ("تصعيد", incidents.filter(escalation_level=Incident.EscalationLevel.DEPARTMENT_HEAD).count(), "مصعّدة للقسم", "ضمن النطاق الإشرافي"),
            ("تحديث", open_action_rows.filter(action_type=IncidentSupervisoryAction.ActionType.REQUEST_UPDATE).count(), "طلبات تحديث مفتوحة", "بانتظار رد أو مراجعة"),
            ("رد", open_action_rows.filter(action_type=IncidentSupervisoryAction.ActionType.REQUEST_UPDATE, status=IncidentSupervisoryAction.Status.ANSWERED).count(), "ردود تنتظر المراجعة", "ردود مرتبطة بطلبات"),
            ("توجيه", open_action_rows.filter(action_type=IncidentSupervisoryAction.ActionType.SUPERVISORY_DIRECTIVE).count(), "توجيهات مفتوحة", "بانتظار الاستلام أو الإنجاز"),
            ("تنفيذي", incidents.filter(escalation_level=Incident.EscalationLevel.GENERAL_MANAGER).count(), "مصعّدة للمدير العام", "ضمن القيادة التنفيذية"),
        ]
    elif active_center == "administrative":
        kpis = [
            ("متابعة", len(attention_queue), "متابعات مفتوحة", "تحتاج متابعة إدارية"),
            ("تحديث", open_action_rows.filter(action_type=IncidentSupervisoryAction.ActionType.REQUEST_UPDATE).count(), "طلبات تحديث", "بانتظار رد أو مراجعة"),
            ("تنبيه", open_action_rows.filter(action_type=IncidentSupervisoryAction.ActionType.ADMINISTRATIVE_ALERT).count(), "تنبيهات إدارية", "تنبيهات ما زالت مفتوحة"),
            ("توجيه", open_action_rows.filter(action_type__in=(IncidentSupervisoryAction.ActionType.SUPERVISORY_DIRECTIVE, IncidentSupervisoryAction.ActionType.EXECUTIVE_DIRECTIVE)).count(), "توجيهات مفتوحة", "بانتظار الإنجاز"),
            ("صيانة", incidents.filter(maintenance_request__status__in=(MaintenanceRequest.Status.DONE, MaintenanceRequest.Status.CLOSED)).count(), "صيانة مكتملة", "تنتظر المتابعة"),
        ]
    else:
        kpis = [
            ("قرار", len(attention_queue), "تحتاج قرارًا", "حالات تنتظر تدخلك"),
            ("تصعيد", incidents.filter(escalation_level=Incident.EscalationLevel.GENERAL_MANAGER).count(), "مصعّدة إليك", "ضمن النطاق التنفيذي"),
            ("توجيه", open_action_rows.filter(action_type=IncidentSupervisoryAction.ActionType.EXECUTIVE_DIRECTIVE).count(), "توجيهات تنفيذية مفتوحة", "بانتظار الإنجاز"),
            ("إعادة", open_action_rows.filter(action_type=IncidentSupervisoryAction.ActionType.RETURN_TO_FOLLOWUP).count(), "أعيدت للقسم", "للمتابعة الإشرافية"),
            ("مكتمل", action_rows.filter(action_type=IncidentSupervisoryAction.ActionType.SUPERVISORY_RESOLVED).count(), "مكتملة إشرافيًا", "متابعة منتهية"),
        ]
    active_delegations = LeadershipDelegation.objects.none()
    deputies = []
    if user_has_role(request.user, SupervisoryLeadershipService.HEAD) or request.user.is_superuser:
        active_delegations = LeadershipDelegation.objects.filter(
            principal=request.user, revoked_at__isnull=True, ends_at__gt=timezone.now(),
        ).select_related("delegate")
        from apps.roles.models import UserRole
        deputies = UserRole.objects.filter(
            role__code=SupervisoryLeadershipService.DEPUTY, role__is_active=True,
            is_active=True, user__is_active=True,
        ).select_related("user", "role")
    return render(request, "ops/supervisory_command_center.html", {
        "center": active_center, "page": page, "selected_incident": selected_incident,
        "selected_section": selected_section, "selected_status": selected_status,
        "action_choices": action_choices, "authority_role": role,
        "active_delegation": delegation, "active_delegations": active_delegations,
        "deputies": deputies,
        "attention_queue": attention_queue,
        "center_config": center_config, "center_links": role_centers,
        "actual_role_label": actual_role_label,
        "center_label": center_config["title"].removeprefix("مركز "),
        "effective_capacity_label": actual_role_label,
        "is_cross_center_oversight": is_cross_center_oversight,
        "is_delegated_acting": is_delegated_acting,
        "is_deputy": is_deputy,
        "delegated_principal_label": "رئيس قسم الأبواب" if is_delegated_acting else "",
        "kpis": kpis, "allowed_sections": sorted(get_allowed_sections(request.user)),
    })


@login_required
@require_POST
def create_supervisory_action_view(request, pk):
    incident = get_object_or_404(Incident, pk=pk)
    try:
        action = SupervisoryLeadershipService.create_action(
            incident=incident, actor=request.user,
            action_type=str(request.POST.get("action_type") or ""),
            subject=request.POST.get("subject", ""), note=request.POST.get("note", ""),
        )
    except ValidationError as error:
        return JsonResponse({"ok": False, "message": _validation_error_message(error)}, status=400)
    except PermissionDenied:
        return JsonResponse({"ok": False, "message": "لا تملك صلاحية هذا الإجراء."}, status=403)
    return JsonResponse({"ok": True, "action_id": action.pk, "message": "تم تسجيل الإجراء الإشرافي."})


def _supervisory_transition_response(callback):
    try:
        action = callback()
    except ValidationError as error:
        return JsonResponse(
            {"ok": False, "message": _validation_error_message(error)}, status=400,
        )
    except PermissionDenied:
        return JsonResponse(
            {"ok": False, "message": "لا تملك صلاحية هذا الانتقال."}, status=403,
        )
    return JsonResponse({"ok": True, "action_id": action.pk, "message": "تم تحديث الإجراء الإشرافي."})


@login_required
@require_POST
def respond_to_update_request_view(request, pk):
    action = get_object_or_404(IncidentSupervisoryAction, pk=pk)
    return _supervisory_transition_response(lambda: (
        SupervisoryLeadershipService.respond_to_update_request(
            action, request.user, request.POST.get("note", "")
        )
    ))


@login_required
@require_POST
def resolve_update_request_view(request, pk):
    action = get_object_or_404(IncidentSupervisoryAction, pk=pk)
    return _supervisory_transition_response(lambda: (
        SupervisoryLeadershipService.resolve_update_request(action, request.user)
    ))


@login_required
@require_POST
def acknowledge_directive_view(request, pk):
    action = get_object_or_404(IncidentSupervisoryAction, pk=pk)
    return _supervisory_transition_response(lambda: (
        SupervisoryLeadershipService.acknowledge_directive(action, request.user)
    ))


@login_required
@require_POST
def complete_directive_view(request, pk):
    action = get_object_or_404(IncidentSupervisoryAction, pk=pk)
    return _supervisory_transition_response(lambda: (
        SupervisoryLeadershipService.complete_directive(
            action, request.user, request.POST.get("note", "")
        )
    ))


@login_required
@require_POST
def create_leadership_delegation_view(request):
    from django.contrib.auth import get_user_model
    try:
        delegation = SupervisoryLeadershipService.create_delegation(
            principal=request.user,
            delegate=get_object_or_404(get_user_model(), pk=request.POST.get("delegate")),
            section=str(request.POST.get("section") or ""),
            starts_at=_parse_planned_datetime(request.POST.get("starts_at")),
            ends_at=_parse_planned_datetime(request.POST.get("ends_at")),
            reason=request.POST.get("reason", ""),
        )
    except ValidationError as error:
        return JsonResponse({"ok": False, "message": _validation_error_message(error)}, status=400)
    except PermissionDenied:
        return JsonResponse({"ok": False, "message": "لا تملك صلاحية إنشاء التفويض."}, status=403)
    return JsonResponse({"ok": True, "delegation_id": delegation.pk, "message": "تم إنشاء التفويض."})


@login_required
@require_POST
def revoke_leadership_delegation_view(request, pk):
    delegation = get_object_or_404(LeadershipDelegation, pk=pk)
    try:
        SupervisoryLeadershipService.revoke_delegation(delegation, request.user)
    except (ValidationError, PermissionDenied) as error:
        status = 403 if isinstance(error, PermissionDenied) else 400
        return JsonResponse({"ok": False, "message": _validation_error_message(error)}, status=status)
    return JsonResponse({"ok": True, "message": "تم إلغاء التفويض."})


def _parse_planned_datetime(value):
    parsed = parse_datetime(str(value or "").strip())
    if parsed and timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


# =========================================================
# أدوات مساعدة
# =========================================================


def _get_active_shift():
    """
    إرجاع الوردية النشطة الحالية.
    """

    return (
        ShiftPlan.objects
        .select_related("shift_type")
        .filter(is_active=True)
        .first()
    )


def _validation_error_message(error):
    """
    تحويل ValidationError إلى رسالة نصية مناسبة.
    """

    if hasattr(error, "message_dict"):
        messages = []

        for field_messages in error.message_dict.values():
            messages.extend(field_messages)

        return " ".join(
            str(message)
            for message in messages
        )

    if hasattr(error, "messages"):
        return " ".join(
            str(message)
            for message in error.messages
        )

    return str(error)


def _require_ops_permission(request, permission):
    """Require both a platform permission and an institutional role scope."""
    if request.user.is_superuser:
        return
    if (
        not user_has_permission(request.user, permission)
        or not has_institutional_scope(request.user)
    ):
        raise PermissionDenied("لا تملك صلاحية الوصول إلى هذه العملية.")


def _scoped_by_section(queryset, user, field_name="section"):
    if user.is_superuser:
        return queryset
    return queryset.filter(
        **{f"{field_name}__in": get_allowed_sections(user)}
    )


def _incident_assignees(active_shift, *, section="", user=None):
    if not active_shift:
        return ShiftAssignment.objects.none()

    queryset = (
        ShiftAssignment.objects
        .filter(
            shift_plan=active_shift,
            is_confirmed=True,
            employee__is_active=True,
            employee__user__is_active=True,
        )
        .filter(
            Q(
                role=ShiftAssignment.OperationalRole.SHIFT_HEAD,
                employee__user__platform_role_assignments__is_active=True,
                employee__user__platform_role_assignments__role__is_active=True,
                employee__user__platform_role_assignments__role__code="shift_supervisor",
            )
            | Q(
                role=ShiftAssignment.OperationalRole.SHIFT_DEPUTY,
                employee__user__platform_role_assignments__is_active=True,
                employee__user__platform_role_assignments__role__is_active=True,
                employee__user__platform_role_assignments__role__code="shift_deputy",
            )
        )
        .select_related("employee", "employee__user")
        .annotate(
            incident_role_order=Case(
                When(
                    role=ShiftAssignment.OperationalRole.SHIFT_HEAD,
                    then=Value(0),
                ),
                default=Value(1),
                output_field=IntegerField(),
            )
        )
        .order_by("incident_role_order", "employee__full_name")
        .distinct()
    )

    allowed_sections = get_allowed_sections(user) if user else set()
    if section:
        queryset = queryset.filter(employee__operational_section=section)
    elif user and has_institutional_scope(user):
        queryset = queryset.filter(employee__operational_section__in=allowed_sections)

    return queryset


# =========================================================
# مركز العمليات المباشرة
# =========================================================


@login_required
def operations_center_view(request):
    """
    الصفحة الرئيسية لمركز العمليات المباشرة.
    """

    context = OperationsCenterService.build()
    operational_leadership = leadership_for_shift(context.get("active_shift"))

    metrics = context["metrics"]
    total_doors = metrics.total_doors or 0
    readiness_rate = round(
        ((metrics.open_doors + metrics.secured_doors) / total_doors) * 100, 1
    ) if total_doors else 0
    supervisor_coverage = round(
        ((total_doors - metrics.doors_without_supervisor) / total_doors) * 100, 1
    ) if total_doors else 0
    monitor_coverage = round(
        ((total_doors - metrics.doors_without_monitor) / total_doors) * 100, 1
    ) if total_doors else 0
    if metrics.critical_incidents or readiness_rate < 70:
        operations_status = {"key": "critical", "label": "تدخل فوري مطلوب"}
    elif metrics.open_maintenance or readiness_rate < 90:
        operations_status = {"key": "warning", "label": "مستقر مع ملاحظات"}
    else:
        operations_status = {"key": "stable", "label": "تشغيل مستقر"}

    incident_queue = Incident.objects.select_related(
        "door_shift", "shift_plan"
    ).filter(status__in=[
        Incident.Status.NEW, Incident.Status.IN_PROGRESS, Incident.Status.FORWARDED,
    ]).annotate(priority_order=Case(
        When(priority=Incident.Priority.CRITICAL, then=Value(0)),
        When(priority=Incident.Priority.HIGH, then=Value(1)),
        When(priority=Incident.Priority.MEDIUM, then=Value(2)),
        default=Value(3), output_field=IntegerField(),
    )).order_by("priority_order", "created_at")[:6]
    maintenance_queue = _scoped_by_section(
        MaintenanceRequest.objects, request.user
    ).select_related(
        "door_shift", "technician", "created_by"
    ).filter(status=MaintenanceRequest.Status.NEW).annotate(priority_order=Case(
        When(priority=MaintenanceRequest.Priority.URGENT, then=Value(0)),
        When(priority=MaintenanceRequest.Priority.HIGH, then=Value(1)),
        When(priority=MaintenanceRequest.Priority.MEDIUM, then=Value(2)),
        default=Value(3), output_field=IntegerField(),
    )).order_by("priority_order", "created_at")[:6]

    context.update(
        {
            "state_choices": (
                DoorShift.DoorState.choices
            ),
            "maintenance_priority_choices": (
                MaintenanceRequest.Priority.choices
            ),
            "incident_type_choices": (
                Incident.IncidentType.choices
            ),
            "incident_priority_choices": (
                Incident.Priority.choices
            ),
            "active_assignments_count": DoorAssignment.objects.filter(
                shift_plan=context.get("active_shift"), is_active=True
            ).count() if context.get("active_shift") else 0,
            "active_breaks_count": Break.objects.filter(
                is_active=True
            ).count(),
            "total_incidents_count": Incident.objects.count(),
            "total_maintenance_count": MaintenanceRequest.objects.count(),
            "readiness_rate": readiness_rate,
            "supervisor_coverage": supervisor_coverage,
            "monitor_coverage": monitor_coverage,
            "operations_status": operations_status,
            "incident_queue": incident_queue,
            "maintenance_queue": maintenance_queue,
            "operations_supervisor": operational_leadership.get(
                ShiftOperationalLeadership.Responsibility.OPERATIONS_SUPERVISOR
            ),
        }
    )

    return render(
        request,
        "ops/operations_center.html",
        context,
    )


# =========================================================
# غرفة القيادة التشغيلية
# =========================================================


@login_required
def command_center_view(request):
    """
    عرض غرفة القيادة التشغيلية.
    """

    context = CommandCenterService.build()

    return render(
        request,
        "ops/command_center.html",
        context,
    )


@login_required
def command_center_data_ajax(request):
    """
    تزويد غرفة القيادة بالبيانات اللحظية بصيغة JSON.
    """

    return JsonResponse(
        CommandCenterService.build_json()
    )


# =========================================================
# حالات الأبواب
# =========================================================


@login_required
def door_status_view(request):
    """
    عرض الحالات الحالية للأبواب.
    """

    _require_ops_permission(request, PlatformPermissions.VIEW_DOORS)
    active_shift = _get_active_shift()

    can_view_distribution = request.user.is_superuser or user_has_permission(
        request.user, PlatformPermissions.VIEW_DISTRIBUTION
    )
    snapshot = EngineeringCenterService.build(
        active_shift=active_shift,
        include_employee_names=can_view_distribution,
        allowed_sections=None if request.user.is_superuser else get_allowed_sections(request.user),
    )

    official_doors = list(
        Door.objects.filter(is_active=True)
        .order_by("sort_order", "door_number")
        .select_related("zone")
    )

    door_shift_map = {}
    if active_shift:
        door_shift_map = {
            item.door_number: item
            for item in (
                _scoped_by_section(DoorShift.objects, request.user)
                .filter(
                    shift_plan=active_shift,
                    is_active=True,
                )
                .select_related(
                    "supervisor",
                    "shift_plan",
                )
                .annotate(
                    open_incidents_count=Count(
                        "incidents",
                        filter=Q(incidents__status__in=[
                            Incident.Status.NEW,
                            Incident.Status.IN_PROGRESS,
                            Incident.Status.FORWARDED,
                        ]),
                        distinct=True,
                    ),
                    open_maintenance_count=Count(
                        "maintenance_requests",
                        filter=Q(maintenance_requests__status__in=[
                            MaintenanceRequest.Status.NEW,
                            MaintenanceRequest.Status.APPROVED,
                            MaintenanceRequest.Status.ASSIGNED,
                            MaintenanceRequest.Status.IN_PROGRESS,
                            MaintenanceRequest.Status.OPEN,
                        ]),
                        distinct=True,
                    ),
                )
                .order_by("door_number")
            )
        }

    current_states = (
        DoorCurrentState.objects
        .select_related(
            "door",
            "current_shift",
            "current_shift__shift_plan",
            "updated_by",
        )
        .filter(
            door__is_active=True,
        )
        .order_by(
            "door__door_number",
        )
    )
    current_state_map = {item.door_id: item for item in current_states}

    visible_door_rows = []
    for door in official_doors:
        active_door_shift = door_shift_map.get(door.door_number)
        current_state = current_state_map.get(door.id)
        state = OperationsCenterService._resolve_state(
            door_shift=active_door_shift,
            current_state=current_state,
        )
        notes = OperationsCenterService._resolve_notes(
            door_shift=active_door_shift,
            current_state=current_state,
        )
        updated_at = OperationsCenterService._resolve_updated_at(
            door_shift=active_door_shift,
            current_state=current_state,
        )

        display_shift = active_door_shift or DoorShift(
            door_number=door.door_number,
            state=state,
            notes=notes,
            sort_order=door.sort_order,
            is_active=True,
        )
        display_shift.direction_key, display_shift.direction_label = (
            OperationsCenterService.direction_for_number(door.door_number)
        )
        display_shift.supervisor = getattr(active_door_shift, "supervisor", None)
        display_shift.updated_at = updated_at
        display_shift.open_incidents_count = 0
        display_shift.open_maintenance_count = 0
        if active_door_shift is not None:
            display_shift.open_incidents_count = getattr(active_door_shift, "open_incidents_count", 0)
            display_shift.open_maintenance_count = getattr(active_door_shift, "open_maintenance_count", 0)
        visible_door_rows.append(display_shift)

    open_count = sum(1 for row in visible_door_rows if row.state == DoorShift.DoorState.OPEN)
    closed_count = sum(1 for row in visible_door_rows if row.state == DoorShift.DoorState.CLOSED)
    maintenance_count = sum(1 for row in visible_door_rows if row.state == DoorShift.DoorState.MAINTENANCE)
    secured_count = sum(1 for row in visible_door_rows if row.state == DoorShift.DoorState.SECURED)
    total_count = len(visible_door_rows)
    readiness_rate = round(
        ((open_count + secured_count) / total_count * 100),
        1,
    ) if total_count else 0

    has_scope = request.user.is_superuser or has_institutional_scope(
        request.user
    )
    can_update_doors = request.user.is_superuser or (
        has_scope
        and any(
            user_has_permission(request.user, permission)
            for permission in (
                PlatformPermissions.OPEN_DOOR,
                PlatformPermissions.CLOSE_DOOR,
                PlatformPermissions.MOVE_DOOR_TO_MAINTENANCE,
            )
        )
    )
    can_create_maintenance = request.user.is_superuser or (
        has_scope
        and user_has_permission(
            request.user,
            PlatformPermissions.CREATE_MAINTENANCE_REQUEST,
        )
    )

    context = {
        "active_shift": active_shift,
        "active_door_shifts": visible_door_rows,
        "current_door_states": current_states,
        "open_count": open_count,
        "closed_count": closed_count,
        "maintenance_count": maintenance_count,
        "secured_count": secured_count,
        "total_count": total_count,
        "readiness_rate": readiness_rate,
        "can_update_doors": can_update_doors,
        "can_create_maintenance": can_create_maintenance,
        "recent_state_changes": DoorStateHistory.objects.select_related(
            "door_shift", "changed_by"
        ).filter(
            door_shift__shift_plan=active_shift
        ).order_by("-changed_at")[:8] if active_shift else DoorStateHistory.objects.none(),
        "engineering_doors": snapshot["doors"],
        "engineering_summary": snapshot["summary"],
        "can_manage_staff_targets": request.user.is_superuser or user_has_permission(
            request.user, PlatformPermissions.CHANGE_SYSTEM_SETTINGS
        ) or user_has_permission(
            request.user, PlatformPermissions.CHANGE_DOOR_COVERAGE_SETTINGS
        ),
        "can_view_staff_targets": user_has_permission(
            request.user, PlatformPermissions.VIEW_SYSTEM_SETTINGS
        ) or user_has_permission(
            request.user, PlatformPermissions.VIEW_DOOR_COVERAGE_SETTINGS
        ),
        "can_create_incident": request.user.is_superuser or user_has_permission(
            request.user, PlatformPermissions.CREATE_INCIDENT
        ),
        "can_view_incidents": True,
        "can_view_maintenance": request.user.is_superuser or user_has_permission(
            request.user, PlatformPermissions.VIEW_MAINTENANCE_REQUESTS
        ),
        "can_view_distribution": can_view_distribution,
        "incident_supervisor": leadership_for_shift(active_shift).get(
            ShiftOperationalLeadership.Responsibility.INCIDENT_SUPERVISOR
        ),
    }

    return render(
        request,
        "ops/door_status.html",
        context,
    )


@login_required
@require_GET
def door_incident_followup_ajax(request, pk):
    _require_ops_permission(request, PlatformPermissions.VIEW_DOORS)
    door = get_object_or_404(
        filter_doors_for_user(
            Door.objects.filter(is_active=True).select_related("current_state"), request.user
        ),
        pk=pk,
    )
    active_shift = _get_active_shift()
    supervisor = leadership_for_shift(active_shift).get(
        ShiftOperationalLeadership.Responsibility.INCIDENT_SUPERVISOR
    )
    allowed_sections = None if request.user.is_superuser else get_allowed_sections(request.user)
    section_label = (
        "جميع الأقسام" if allowed_sections is None
        else "رجالي" if allowed_sections == {"male"}
        else "نسائي" if allowed_sections == {"female"}
        else "نطاق مشترك"
    )
    current_state = getattr(door, "current_state", None)
    return JsonResponse(EngineeringIncidentFollowupService.build(
        door=door,
        user=request.user,
        active_shift=active_shift,
        incident_supervisor=supervisor,
        section_label=section_label,
        door_status_label=current_state.get_state_display() if current_state else "غير محدد",
        can_view_maintenance_details=user_has_permission(
            request.user, PlatformPermissions.VIEW_MAINTENANCE_REQUESTS
        ),
    ))


@login_required
@require_POST
def update_door_staff_targets(request):
    return door_coverage_settings_view(request)


def _coverage_settings_context(request, *, posted=None, errors=None):
    snapshot = EngineeringCenterService.build(
        active_shift=_get_active_shift(),
        allowed_sections=None if request.user.is_superuser else get_allowed_sections(request.user),
    )
    scoped_ids = set(filter_doors_for_user(
        Door.objects.filter(is_active=True), request.user
    ).values_list("pk", flat=True))
    rows = [row for row in snapshot["doors"] if row.door.pk in scoped_ids]
    for row in rows:
        field = f"target_{row.door.pk}"
        row.form_value = posted.get(field, "") if posted is not None else (
            row.target_staff_count if row.target_staff_count is not None else ""
        )
        row.form_error = (errors or {}).get(row.door.pk, "")
    configured = [row for row in rows if row.target_staff_count is not None]
    return {
        "coverage_rows": rows,
        "configured_count": len(configured),
        "unconfigured_count": len(rows) - len(configured),
        "total_target": sum(row.target_staff_count or 0 for row in rows),
        "can_change_settings": user_has_permission(
            request.user, PlatformPermissions.CHANGE_DOOR_COVERAGE_SETTINGS
        ) or user_has_permission(request.user, PlatformPermissions.CHANGE_SYSTEM_SETTINGS),
        "updated_count": request.GET.get("updated", ""),
    }


@login_required
def door_coverage_settings_view(request):
    """Manage approved per-door staffing targets without touching assignments."""
    if request.method == "GET":
        if not (
            user_has_permission(request.user, PlatformPermissions.VIEW_DOOR_COVERAGE_SETTINGS)
            or user_has_permission(request.user, PlatformPermissions.VIEW_SYSTEM_SETTINGS)
        ):
            raise PermissionDenied("لا تملك صلاحية عرض إعدادات التغطية التشغيلية.")
        return render(request, "ops/door_coverage_settings.html", _coverage_settings_context(request))

    if not (
        user_has_permission(request.user, PlatformPermissions.CHANGE_DOOR_COVERAGE_SETTINGS)
        or user_has_permission(request.user, PlatformPermissions.CHANGE_SYSTEM_SETTINGS)
    ):
        raise PermissionDenied("لا تملك صلاحية تعديل إعدادات التغطية التشغيلية.")
    doors = list(filter_doors_for_user(
        Door.objects.filter(is_active=True), request.user
    ).order_by("sort_order", "door_number"))
    door_by_id = {door.pk: door for door in doors}
    submitted_ids = set()
    malformed = False
    for key in request.POST:
        if key.startswith("target_"):
            try:
                submitted_ids.add(int(key.removeprefix("target_")))
            except ValueError:
                malformed = True
    if malformed or not submitted_ids.issubset(door_by_id):
        raise PermissionDenied("باب غير مصرح به أو خارج النطاق التشغيلي.")

    submitted = {}
    errors = {}
    for door_id in submitted_ids:
        raw_value = str(request.POST.get(f"target_{door_id}", "") or "").strip()
        if not raw_value:
            submitted[door_id] = None
        elif not raw_value.isdigit() or not 1 <= int(raw_value) <= 999:
            errors[door_id] = "أدخل عددًا صحيحًا من 1 إلى 999، أو اترك الحقل فارغًا."
        else:
            submitted[door_id] = int(raw_value)
    if errors:
        return render(
            request, "ops/door_coverage_settings.html",
            _coverage_settings_context(request, posted=request.POST, errors=errors),
            status=400,
        )

    existing = {
        profile.door_id: profile
        for profile in DoorOperationalProfile.objects.filter(door_id__in=submitted)
    }
    changes, to_create, to_update = [], [], []
    now = timezone.now()
    for door_id, target in submitted.items():
        profile = existing.get(door_id)
        previous = profile.target_staff_count if profile else None
        if previous == target:
            continue
        changes.append((door_by_id[door_id], previous, target))
        if profile is None:
            if target is not None:
                to_create.append(DoorOperationalProfile(door_id=door_id, target_staff_count=target))
        else:
            profile.target_staff_count = target
            profile.updated_at = now
            to_update.append(profile)

    with transaction.atomic():
        DoorOperationalProfile.objects.bulk_create(to_create)
        if to_update:
            DoorOperationalProfile.objects.bulk_update(to_update, ["target_staff_count", "updated_at"])
        for door, previous, target in changes:
            record_live_operation(
                module="doors", action=SystemActivityLog.ActionType.UPDATE,
                description=(f"تعديل التغطية التشغيلية للباب {door.door_number}: "
                             f"{previous if previous is not None else 'غير مهيأة'} ← "
                             f"{target if target is not None else 'غير مهيأة'}"),
                request=request,
            )
    return redirect(f"{reverse('ops:door-coverage-settings')}?updated={len(changes)}")


@login_required
def door_status_data_ajax(request):
    """Return the engineering-center card snapshot for lightweight polling."""
    _require_ops_permission(request, PlatformPermissions.VIEW_DOORS)
    snapshot = EngineeringCenterService.build(
        active_shift=_get_active_shift(),
        allowed_sections=None if request.user.is_superuser else get_allowed_sections(request.user),
    )
    return JsonResponse({
        "doors": [
            {
                "id": item.door.pk,
                "number": item.door.door_number,
                "status": item.status,
                "status_label": item.status_label,
                "employee_count": item.employee_count,
                "open_incident_count": item.open_incident_count,
                "today_incident_count": item.today_incident_count,
                "active_maintenance_count": item.active_maintenance_count,
                "target_staff_count": item.target_staff_count,
                "staff_coverage_percent": item.staff_coverage_percent,
                "staff_coverage_level": item.staff_coverage_level,
                "staff_coverage_label": item.staff_coverage_label,
                "staff_coverage_detail": item.staff_coverage_detail,
                "coverage_applicable": item.staff_coverage_applicable,
                "coverage_status": item.staff_coverage_level,
                "coverage_reason": item.staff_coverage_reason,
                "staff_delta": item.staff_delta,
                "last_activity": item.last_activity.isoformat() if item.last_activity else None,
            }
            for item in snapshot["doors"]
        ],
        "summary": snapshot["summary"],
    })


@login_required
@require_POST
def update_door_status_ajax(request, pk):
    """
    تحديث حالة باب ضمن الوردية النشطة.
    """

    state = (
        request.POST.get("state", request.POST.get("status", "")) or ""
    ).strip()
    permission = {
        DoorShift.DoorState.OPEN: PlatformPermissions.OPEN_DOOR,
        DoorShift.DoorState.CLOSED: PlatformPermissions.CLOSE_DOOR,
        DoorShift.DoorState.MAINTENANCE: PlatformPermissions.MOVE_DOOR_TO_MAINTENANCE,
    }.get(state, PlatformPermissions.VIEW_DOORS)
    _require_ops_permission(request, permission)

    door = get_object_or_404(
        _scoped_by_section(DoorShift.objects, request.user).select_related(
            "shift_plan",
            "supervisor",
        ),
        pk=pk,
        shift_plan__is_active=True,
        is_active=True,
    )

    notes = (
        request.POST.get(
            "notes",
            request.POST.get(
                "reason",
                "",
            ),
        )
        or ""
    ).strip()

    maintenance = None
    try:
        if state == DoorShift.DoorState.MAINTENANCE:
            _require_ops_permission(
                request,
                PlatformPermissions.CREATE_MAINTENANCE_REQUEST,
            )
            maintenance = MaintenanceService.create_request(
                request=request,
                door=door,
                description=notes,
                priority=MaintenanceRequest.Priority.MEDIUM,
                planned_start_at=_parse_planned_datetime(request.POST.get("planned_start_at")),
                planned_end_at=_parse_planned_datetime(request.POST.get("planned_end_at")),
            )
            door.refresh_from_db()
            changed = False
        else:
            door, changed = DoorService.update_state(
                request=request,
                door_shift=door,
                new_state=state,
                reason=notes,
            )

    except ValidationError as error:
        return JsonResponse(
            {
                "success": False,
                "error": (
                    _validation_error_message(
                        error
                    )
                ),
            },
            status=400,
        )

    master_door = Door.objects.filter(door_number=door.door_number).first()
    current_state = (
        DoorCurrentState.objects.filter(door=master_door).first()
        if master_door else None
    )
    resolved_state = OperationsCenterService._resolve_state(
        door_shift=door,
        current_state=current_state,
    )
    resolved_notes = OperationsCenterService._resolve_notes(
        door_shift=door,
        current_state=current_state,
    )

    return JsonResponse(
        {
            "success": True,
            "changed": changed,
            "maintenance_request_id": (
                maintenance.id if maintenance else None
            ),
            "maintenance_status": (
                maintenance.status if maintenance else None
            ),
            "door": {
                "id": door.id,
                "door_number": door.door_number,
                "state": resolved_state,
                "state_label": (
                    dict(DoorShift.DoorState.choices).get(resolved_state, resolved_state)
                ),
                "notes": resolved_notes,
                "updated_at": (
                    door.updated_at.strftime(
                        "%Y-%m-%d %H:%M"
                    )
                    if door.updated_at
                    else None
                ),
            },
        }
    )


# =========================================================
# طلبات الصيانة
# =========================================================


@login_required
@require_POST
def create_maintenance_request_ajax(
    request,
    pk,
):
    """
    إنشاء طلب صيانة مرتبط بباب في الوردية النشطة.
    """

    _require_ops_permission(
        request, PlatformPermissions.CREATE_MAINTENANCE_REQUEST
    )
    door = get_object_or_404(
        _scoped_by_section(DoorShift.objects, request.user).select_related(
            "shift_plan",
        ),
        pk=pk,
        shift_plan__is_active=True,
        is_active=True,
    )

    description = (
        request.POST.get(
            "description",
            "",
        )
        or ""
    ).strip()

    priority = (
        request.POST.get(
            "priority",
            MaintenanceRequest.Priority.MEDIUM,
        )
        or MaintenanceRequest.Priority.MEDIUM
    ).strip()

    technician_name = (
        request.POST.get(
            "technician_name",
            "",
        )
        or ""
    ).strip()

    technician_phone = (request.POST.get("technician_phone", "") or "").strip()
    planned_start_at = _parse_planned_datetime(request.POST.get("planned_start_at"))
    planned_end_at = _parse_planned_datetime(request.POST.get("planned_end_at"))

    section = str(
        request.POST.get("section", "") or ""
    ).strip().lower()
    assignment = None
    assignment_id = str(
        request.POST.get("assignment_id", "") or ""
    ).strip()

    if assignment_id.isdigit():
        assignment = get_object_or_404(
            _scoped_by_section(DoorAssignment.objects, request.user).select_related(
                "door",
            ),
            pk=int(assignment_id),
            shift_plan=door.shift_plan,
            door__door_number=door.door_number,
            is_active=True,
        )

    try:
        maintenance = (
            MaintenanceService.create_request(
                request=request,
                door=door,
                description=description,
                priority=priority,
                technician_name=technician_name,
                technician_phone=technician_phone,
                planned_start_at=planned_start_at,
                planned_end_at=planned_end_at,
                section=section,
                assignment=assignment,
            )
        )

    except ValidationError as error:
        return JsonResponse(
            {
                "success": False,
                "error": (
                    _validation_error_message(
                        error
                    )
                ),
            },
            status=400,
        )

    return JsonResponse(
        {
            "success": True,
            "maintenance": {
                "id": maintenance.id,
                "request_number": (
                    maintenance.request_number
                ),
                "status": maintenance.status,
                "status_label": (
                    maintenance.get_status_display()
                ),
                "priority": maintenance.priority,
                "priority_label": (
                    maintenance.get_priority_display()
                ),
                "technician_name": (
                    maintenance.technician_name
                    or "غير محدد"
                ),
                "technician_phone": maintenance.technician_phone,
                "planned_start_at": maintenance.planned_start_at.isoformat(),
                "planned_end_at": maintenance.planned_end_at.isoformat(),
                "planned_duration_minutes": maintenance.planned_duration_minutes,
                "section": maintenance.section,
                "created_at": (
                    maintenance.created_at.strftime(
                        "%Y-%m-%d %H:%M"
                    )
                    if maintenance.created_at
                    else None
                ),
                "door_id": maintenance.door_shift_id,
                "door_number": maintenance.door_shift.door_number,
                "state": maintenance.door_shift.state,
                "state_label": maintenance.door_shift.get_state_display(),
            },
        }
    )


@login_required
def maintenance_requests_view(request):
    """
    عرض طلبات الصيانة مع البحث والتصفية.
    """

    _require_ops_permission(
        request, PlatformPermissions.VIEW_MAINTENANCE_REQUESTS
    )
    active_shift = _get_active_shift()

    status_filter = (
        request.GET.get(
            "status",
            "",
        )
        or ""
    ).strip()

    priority_filter = (
        request.GET.get(
            "priority",
            "",
        )
        or ""
    ).strip()

    section_filter = str(
        request.GET.get("section", "") or ""
    ).strip().lower()

    query = (
        request.GET.get(
            "q",
            "",
        )
        or ""
    ).strip()

    maintenance_requests = (
        _scoped_by_section(MaintenanceRequest.objects, request.user)
        .filter(status__in=(
            MaintenanceRequest.Status.APPROVED,
            MaintenanceRequest.Status.ASSIGNED,
            MaintenanceRequest.Status.IN_PROGRESS,
            MaintenanceRequest.Status.FIXED,
            MaintenanceRequest.Status.OPEN,
            MaintenanceRequest.Status.DONE,
        ))
        .select_related(
            "door_shift",
            "door_shift__shift_plan",
            "created_by",
            "technician",
            "approved_by",
            "assigned_by",
        )
        .order_by("-created_at")
    )
    specialist_responsibility = None
    if not request.user.is_superuser and user_has_role(request.user, "operations_supervisor"):
        specialist_responsibility = ShiftOperationalLeadership.Responsibility.OPERATIONS_SUPERVISOR
    elif not request.user.is_superuser and user_has_role(request.user, "maintenance_shift_supervisor"):
        specialist_responsibility = ShiftOperationalLeadership.Responsibility.MAINTENANCE_SUPERVISOR
    specialist_shift_ids = None
    if specialist_responsibility:
        specialist_shift_ids = ShiftOperationalLeadership.objects.filter(
            responsibility=specialist_responsibility,
            employee__user=request.user,
        ).values_list("shift_plan_id", flat=True)
        maintenance_requests = maintenance_requests.filter(
            door_shift__shift_plan_id__in=specialist_shift_ids
        )

    valid_statuses = {
        value
        for value, _label
        in MaintenanceRequest.Status.choices
    }

    valid_priorities = {
        value
        for value, _label
        in MaintenanceRequest.Priority.choices
    }

    if (
        status_filter
        and status_filter in valid_statuses
    ):
        maintenance_requests = (
            maintenance_requests.filter(
                status=status_filter,
            )
        )

    if (
        priority_filter
        and priority_filter in valid_priorities
    ):
        maintenance_requests = (
            maintenance_requests.filter(
                priority=priority_filter,
            )
        )

    if section_filter in {"male", "female"}:
        maintenance_requests = maintenance_requests.filter(
            section=section_filter,
        )

    if query:
        maintenance_requests = (
            maintenance_requests.filter(
                Q(
                    request_number__icontains=query
                )
                | Q(
                    description__icontains=query
                )
                | Q(
                    door_shift__door_number__icontains=query
                )
                | Q(
                    technician_name__icontains=query
                )
                | Q(
                    created_by__username__icontains=query
                )
                | Q(
                    created_by__first_name__icontains=query
                )
                | Q(
                    created_by__last_name__icontains=query
                )
            )
        )

    today = timezone.localdate()

    all_requests = _scoped_by_section(
        MaintenanceRequest.objects,
        request.user,
    )
    if specialist_shift_ids is not None:
        all_requests = all_requests.filter(
            door_shift__shift_plan_id__in=specialist_shift_ids
        )
    closed_statuses = (
        MaintenanceRequest.Status.CLOSED,
        MaintenanceRequest.Status.DONE,
    )

    context = {
        "requests": maintenance_requests,
        "active_shift": active_shift,
        "maintenance_shift_supervisor": leadership_for_shift(active_shift).get(
            ShiftOperationalLeadership.Responsibility.MAINTENANCE_SUPERVISOR
        ),

        "status_choices": (
            (
                MaintenanceRequest.Status.IN_PROGRESS,
                "بدء المعالجة",
            ),
            (
                MaintenanceRequest.Status.DONE,
                "إنهاء الصيانة",
            ),
        ),
        "priority_choices": (
            MaintenanceRequest.Priority.choices
        ),

        "selected_status": status_filter,
        "selected_priority": priority_filter,
        "selected_section": section_filter,
        "q": query,

        "total_requests": (
            all_requests.count()
        ),

        "today_requests": (
            all_requests.filter(
                created_at__date=today,
            ).count()
        ),

        "open_requests": (
            all_requests.exclude(
                status__in=closed_statuses,
            ).count()
        ),

        "closed_requests": (
            all_requests.filter(
                status__in=closed_statuses,
            ).count()
        ),

        "urgent_requests": (
            all_requests.filter(
                priority=(
                    MaintenanceRequest
                    .Priority
                    .URGENT
                )
            )
            .exclude(
                status__in=closed_statuses,
            )
            .count()
        ),
    }

    return render(
        request,
        "ops/maintenance_requests.html",
        context,
    )


@login_required
@require_POST
def update_maintenance_status_ajax(
    request,
    pk,
):
    """
    تحديث حالة طلب الصيانة.
    """

    new_status = (request.POST.get("status", "") or "").strip()
    permission = {
        MaintenanceRequest.Status.APPROVED: PlatformPermissions.APPROVE_MAINTENANCE_REQUEST,
        MaintenanceRequest.Status.CLOSED: PlatformPermissions.APPROVE_MAINTENANCE_REQUEST,
        MaintenanceRequest.Status.ASSIGNED: PlatformPermissions.ASSIGN_MAINTENANCE_TECHNICIAN,
        MaintenanceRequest.Status.IN_PROGRESS: PlatformPermissions.ASSIGN_MAINTENANCE_TECHNICIAN,
        MaintenanceRequest.Status.FIXED: PlatformPermissions.CLOSE_MAINTENANCE_REQUEST,
        MaintenanceRequest.Status.DONE: PlatformPermissions.CLOSE_MAINTENANCE_REQUEST,
    }.get(new_status, PlatformPermissions.VIEW_MAINTENANCE_REQUESTS)
    _require_ops_permission(request, permission)
    maintenance = get_object_or_404(
        _scoped_by_section(MaintenanceRequest.objects, request.user)
        .select_related(
            "door_shift",
            "door_shift__shift_plan",
        ),
        pk=pk,
    )

    closing_notes = (
        request.POST.get(
            "closing_notes",
            "",
        )
        or ""
    ).strip()

    try:
        maintenance = (
            MaintenanceService.update_status(
                request=request,
                maintenance=maintenance,
                new_status=new_status,
                closing_notes=closing_notes,
            )
        )

    except ValidationError as error:
        return JsonResponse(
            {
                "success": False,
                "error": (
                    _validation_error_message(
                        error
                    )
                ),
            },
            status=400,
        )

    maintenance.door_shift.refresh_from_db(
        fields=[
            "state",
            "notes",
            "updated_at",
        ]
    )

    return JsonResponse(
        {
            "success": True,

            "maintenance": {
                "id": maintenance.id,
                "request_number": (
                    maintenance.request_number
                ),
                "status": maintenance.status,
                "status_label": (
                    maintenance.get_status_display()
                ),
                "priority": maintenance.priority,
                "priority_label": (
                    maintenance.get_priority_display()
                ),
                "technician_name": (
                    maintenance.technician_name
                    or "غير محدد"
                ),
                "closing_notes": (
                    maintenance.closing_notes
                    or ""
                ),
                "closed_at": (
                    maintenance.closed_at.strftime(
                        "%Y-%m-%d %H:%M"
                    )
                    if maintenance.closed_at
                    else None
                ),
            },

            "door": {
                "id": maintenance.door_shift.id,
                "door_number": (
                    maintenance.door_shift.door_number
                ),
                "state": (
                    maintenance.door_shift.state
                ),
                "state_label": (
                    maintenance
                    .door_shift
                    .get_state_display()
                ),
                "notes": (
                    maintenance.door_shift.notes
                    or ""
                ),
                "updated_at": (
                    maintenance
                    .door_shift
                    .updated_at
                    .strftime(
                        "%Y-%m-%d %H:%M"
                    )
                    if maintenance.door_shift.updated_at
                    else None
                ),
            },
        }
    )


# =========================================================
# البلاغات التشغيلية
# =========================================================


@login_required
def incidents_view(request):
    """
    عرض البلاغات التشغيلية مع البحث والتصفية.
    """

    _require_ops_permission(request, PlatformPermissions.VIEW_DOORS)
    active_shift = _get_active_shift()

    status_filter = (
        request.GET.get(
            "status",
            "",
        )
        or ""
    ).strip()

    priority_filter = (
        request.GET.get(
            "priority",
            "",
        )
        or ""
    ).strip()

    section_filter = str(
        request.GET.get("section", "") or ""
    ).strip().lower()

    type_filter = (
        request.GET.get(
            "type",
            "",
        )
        or ""
    ).strip()

    query = (
        request.GET.get(
            "q",
            "",
        )
        or ""
    ).strip()

    door_filter = str(request.GET.get("door", "") or "").strip()
    assigned_filter = str(request.GET.get("assigned_to", "") or "").strip()
    escalation_filter = str(request.GET.get("escalation_level", "") or "").strip()
    maintenance_filter = str(request.GET.get("has_maintenance", "") or "").strip()

    incident_scope = (
        IncidentRoutingService.visible_incidents(Incident.objects, request.user)
        .select_related(
            "door",
            "door_shift",
            "door_shift__shift_plan",
            "shift_plan",
            "created_by",
            "closed_by",
            "assigned_to",
            "maintenance_request",
        )
        .prefetch_related(
            "routing_events", "routing_events__actor",
            Prefetch(
                "supervisory_actions",
                queryset=IncidentSupervisoryAction.objects.select_related(
                    "actor", "acting_for", "target_user", "parent",
                    "acknowledged_by", "completed_by",
                ).order_by("-created_at", "-pk")[:6],
                to_attr="latest_supervisory_actions",
            ),
            Prefetch(
                "supervisory_actions",
                queryset=IncidentSupervisoryAction.objects.filter(
                    status__in=(
                        IncidentSupervisoryAction.Status.OPEN,
                        IncidentSupervisoryAction.Status.ANSWERED,
                        IncidentSupervisoryAction.Status.ACKNOWLEDGED,
                    )
                ).select_related("actor", "target_user", "parent"),
                to_attr="open_supervisory_actions",
            ),
        )
    )
    incidents = incident_scope.order_by("-created_at")

    valid_statuses = {
        value
        for value, _label
        in Incident.Status.choices
    }

    valid_priorities = {
        value
        for value, _label
        in Incident.Priority.choices
    }

    valid_types = {
        value
        for value, _label
        in Incident.IncidentType.choices
    }

    if (
        status_filter
        and status_filter in valid_statuses
    ):
        incidents = incidents.filter(
            status=status_filter,
        )

    if (
        priority_filter
        and priority_filter in valid_priorities
    ):
        incidents = incidents.filter(
            priority=priority_filter,
        )

    if section_filter in {"male", "female"}:
        incidents = incidents.filter(section=section_filter)

    if door_filter.isdigit():
        incidents = incidents.filter(door_id=door_filter)
    if assigned_filter == "unassigned":
        incidents = incidents.filter(assigned_to__isnull=True)
    elif assigned_filter.isdigit():
        incidents = incidents.filter(assigned_to_id=assigned_filter)
    if escalation_filter in dict(Incident.EscalationLevel.choices):
        incidents = incidents.filter(escalation_level=escalation_filter)
    if maintenance_filter == "yes":
        incidents = incidents.filter(maintenance_request__isnull=False)
    elif maintenance_filter == "no":
        incidents = incidents.filter(maintenance_request__isnull=True)

    if (
        type_filter
        and type_filter in valid_types
    ):
        incidents = incidents.filter(
            incident_type=type_filter,
        )

    if query:
        incidents = incidents.filter(
            Q(
                incident_number__icontains=query
            )
            | Q(
                description__icontains=query
            )
            | Q(
                reported_by_name__icontains=query
            )
            | Q(
                assigned_to_name__icontains=query
            )
            | Q(
                door_shift__door_number__icontains=query
            )
            | Q(
                door__door_number__icontains=query
            )
            | Q(
                created_by__username__icontains=query
            )
        )

    doors = filter_doors_for_user(
        Door.objects.filter(is_active=True),
        request.user,
    ).order_by("sort_order", "door_number")
    engineering_door = None
    engineering_door_id = str(
        request.GET.get("engineering_door", "") or ""
    ).strip()
    if engineering_door_id:
        engineering_door = get_object_or_404(doors, pk=engineering_door_id)
    incident_leadership = leadership_for_shift(active_shift).get(
        ShiftOperationalLeadership.Responsibility.INCIDENT_SUPERVISOR
    )

    today = timezone.localdate()

    all_incidents = incident_scope

    closed_statuses = (
        Incident.Status.RESOLVED,
        Incident.Status.CLOSED,
    )

    context = {
        "active_shift": active_shift,
        "incidents": incidents,
        "doors": doors,
        "engineering_door": engineering_door,
        "incident_supervisor": incident_leadership,
        "current_section_label": (
            "جميع الأقسام" if request.user.is_superuser or has_institutional_scope(request.user)
            else "رجالي" if get_allowed_sections(request.user) == {"male"}
            else "نسائي" if get_allowed_sections(request.user) == {"female"}
            else "نطاق مشترك"
        ),

        "status_choices": (
            Incident.Status.choices
        ),
        "priority_choices": (
            Incident.Priority.choices
        ),
        "type_choices": (
            Incident.IncidentType.choices
        ),

        "selected_status": status_filter,
        "selected_priority": priority_filter,
        "selected_section": section_filter,
        "selected_type": type_filter,
        "q": query,
        "selected_door": door_filter,
        "selected_assigned_to": assigned_filter,
        "selected_escalation_level": escalation_filter,
        "selected_has_maintenance": maintenance_filter,
        "escalation_choices": Incident.EscalationLevel.choices,
        "can_assign_incident": user_has_permission(request.user, PlatformPermissions.ASSIGN_INCIDENT),
        "can_escalate_incident": user_has_permission(request.user, PlatformPermissions.ESCALATE_INCIDENT),
        "can_convert_incident": user_has_permission(request.user, PlatformPermissions.CONVERT_INCIDENT_TO_MAINTENANCE),
        "can_close_incident": user_has_permission(request.user, PlatformPermissions.CLOSE_INCIDENT),
        "can_update_incident": user_has_permission(request.user, PlatformPermissions.UPDATE_INCIDENT),
        "can_create_incident": user_has_permission(request.user, PlatformPermissions.CREATE_INCIDENT),

        "total_incidents": (
            all_incidents.count()
        ),

        "open_incidents": (
            all_incidents.exclude(
                status__in=closed_statuses,
            ).count()
        ),

        "critical_incidents": (
            all_incidents.filter(
                priority=(
                    Incident.Priority.CRITICAL
                )
            )
            .exclude(
                status__in=closed_statuses,
            )
            .count()
        ),

        "today_incidents": (
            all_incidents.filter(
                created_at__date=today,
            ).count()
        ),

        "closed_today": (
            all_incidents.filter(
                status__in=closed_statuses,
                closed_at__date=today,
            ).count()
        ),
        "in_progress_incidents": all_incidents.filter(
            status=Incident.Status.IN_PROGRESS
        ).count(),
        "new_incidents": all_incidents.filter(
            status=Incident.Status.NEW
        ).count(),
        "escalated_incidents": all_incidents.exclude(
            escalation_level=Incident.EscalationLevel.NONE
        ).exclude(status__in=closed_statuses).count(),
        "maintenance_incidents": all_incidents.filter(
            maintenance_request__isnull=False
        ).exclude(status__in=closed_statuses).count(),
        "unassigned_incidents": all_incidents.filter(
            assigned_to__isnull=True
        ).exclude(status__in=closed_statuses).count(),
        "awaiting_verification_incidents": all_incidents.filter(
            maintenance_request__status__in=(
                MaintenanceRequest.Status.DONE,
                MaintenanceRequest.Status.CLOSED,
            )
        ).exclude(status__in=closed_statuses).count(),
        "attention_incidents": all_incidents.filter(
            Q(status=Incident.Status.NEW)
            | Q(assigned_to__isnull=True)
            | ~Q(escalation_level=Incident.EscalationLevel.NONE)
            | Q(maintenance_request__status__in=(
                MaintenanceRequest.Status.DONE,
                MaintenanceRequest.Status.CLOSED,
            ))
        ).exclude(status__in=closed_statuses).order_by("created_at")[:4],
        "active_filters": any((
            status_filter, priority_filter, section_filter, type_filter, query,
            door_filter, assigned_filter, escalation_filter, maintenance_filter,
        )),
    }

    return render(
        request,
        "ops/incidents.html",
        context,
    )


@login_required
@require_POST
def create_incident_ajax(
    request,
    pk=None,
    engineering_door_pk=None,
):
    """
    إنشاء بلاغ تشغيلي جديد.
    """

    _require_ops_permission(request, PlatformPermissions.CREATE_INCIDENT)
    active_shift = _get_active_shift()

    door_shift = None
    door = None

    posted_door_id = str(
        request.POST.get("door_id", "") or ""
    ).strip()

    if engineering_door_pk is not None:
        if posted_door_id and posted_door_id != str(engineering_door_pk):
            return JsonResponse(
                {"success": False, "error": "سياق الباب المرسل لا يطابق بطاقة المركز الهندسي."},
                status=400,
            )
        door = get_object_or_404(
            filter_doors_for_user(
                Door.objects.filter(is_active=True),
                request.user,
            ),
            pk=engineering_door_pk,
        )
        door_shift = (
            DoorShift.objects.filter(
                shift_plan=active_shift,
                door_number=door.door_number,
                is_active=True,
            ).first()
        )

    elif posted_door_id:
        door = get_object_or_404(
            filter_doors_for_user(
                Door.objects.filter(is_active=True),
                request.user,
            ),
            pk=posted_door_id,
        )

        door_shift = (
            DoorShift.objects
            .filter(
                shift_plan=active_shift,
                door_number=door.door_number,
                is_active=True,
            )
            .first()
        )

    door_shift_id = (
        request.POST.get(
            "door_shift_id",
            "",
        )
        or ""
    ).strip()

    if engineering_door_pk is not None and door_shift_id:
        return JsonResponse(
            {"success": False, "error": "لا يُقبل تبديل سياق الباب في طلب المركز الهندسي."},
            status=400,
        )

    selected_door_id = (
        door_shift_id
        or pk
    )

    if selected_door_id:
        if active_shift is None:
            return JsonResponse(
                {"success": False, "error": "لا توجد وردية نشطة لهذا السجل."},
                status=400,
            )
        door_shift = get_object_or_404(
            _scoped_by_section(DoorShift.objects, request.user).select_related(
                "shift_plan",
            ),
            pk=selected_door_id,
            shift_plan=active_shift,
            is_active=True,
        )
        door = get_object_or_404(
            filter_doors_for_user(
                Door.objects.filter(is_active=True),
                request.user,
            ),
            door_number=door_shift.door_number,
        )

    description = (
        request.POST.get(
            "description",
            "",
        )
        or ""
    ).strip()

    incident_type = (
        request.POST.get(
            "incident_type",
            Incident.IncidentType.GENERAL,
        )
        or Incident.IncidentType.GENERAL
    ).strip()

    priority = (
        request.POST.get(
            "priority",
            Incident.Priority.MEDIUM,
        )
        or Incident.Priority.MEDIUM
    ).strip()

    reported_by_name = (
        request.POST.get(
            "reported_by_name",
            "",
        )
        or ""
    ).strip()

    section = str(
        request.POST.get("section", "") or ""
    ).strip().lower()
    if door and door.operational_section != Door.OperationalSection.SHARED:
        section = door.operational_section

    assigned_to_id = str(
        request.POST.get("assigned_to_id", "") or ""
    ).strip()
    if assigned_to_id:
        return JsonResponse(
            {"success": False, "error": "يتم تعيين مسؤول البلاغ تلقائيًا."},
            status=400,
        )
    assignment = None
    assignment_id = str(
        request.POST.get("assignment_id", "") or ""
    ).strip()

    if assignment_id.isdigit() and door_shift:
        assignment = get_object_or_404(
            _scoped_by_section(DoorAssignment.objects, request.user).select_related(
                "door",
            ),
            pk=int(assignment_id),
            shift_plan=door_shift.shift_plan,
            door__door_number=door_shift.door_number,
            is_active=True,
        )

    try:
        incident = IncidentService.create(
            request=request,
            active_shift=active_shift,
            door=door,
            door_shift=door_shift,
            assignment=assignment,
            section=section,
            description=description,
            incident_type=incident_type,
            priority=priority,
            reported_by_name=reported_by_name,
        )

    except ValidationError as error:
        return JsonResponse(
            {
                "success": False,
                "error": (
                    _validation_error_message(
                        error
                    )
                ),
            },
            status=400,
        )

    return JsonResponse(
        {
            "success": True,

            "incident": {
                "id": incident.id,
                "incident_number": (
                    incident.incident_number
                ),
                "description": (
                    incident.description
                ),
                "incident_type": (
                    incident.incident_type
                ),
                "incident_type_label": (
                    incident
                    .get_incident_type_display()
                ),
                "status": incident.status,
                "status_label": (
                    incident.get_status_display()
                ),
                "priority": incident.priority,
                "priority_label": (
                    incident.get_priority_display()
                ),
                "reported_by_name": (
                    incident.reported_by_name
                    or ""
                ),
                "assigned_to_name": (
                    incident.assigned_to_name
                    or ""
                ),
                "door_shift_id": (
                    incident.door_shift_id
                ),
                "door_number": (
                    incident.door.door_number
                    if incident.door_id
                    else (
                        incident.door_shift.door_number
                        if incident.door_shift_id
                        else None
                    )
                ),
                "section": incident.section,
                "created_at": (
                    incident.created_at.strftime(
                        "%Y-%m-%d %H:%M"
                    )
                    if incident.created_at
                    else None
                ),
            },
        }
    )


@login_required
@require_POST
def update_incident_status_ajax(
    request,
    pk,
):
    """
    تحديث حالة بلاغ تشغيلي.
    """

    requested_status = str(request.POST.get("status", "") or "").strip()
    required_permission = (
        PlatformPermissions.CLOSE_INCIDENT
        if requested_status == Incident.Status.CLOSED
        else PlatformPermissions.UPDATE_INCIDENT
    )
    _require_ops_permission(request, required_permission)
    incident = get_object_or_404(
        _scoped_by_section(Incident.objects, request.user).select_related(
            "door_shift",
            "door_shift__shift_plan",
            "shift_plan",
            "created_by",
            "closed_by",
        ),
        pk=pk,
    )

    new_status = requested_status

    closing_notes = (
        request.POST.get(
            "closing_notes",
            "",
        )
        or ""
    ).strip()

    try:
        incident, _changed = (
            IncidentService.change_status(
                request=request,
                incident=incident,
                new_status=new_status,
                closing_notes=closing_notes,
            )
        )

    except ValidationError as error:
        return JsonResponse(
            {
                "success": False,
                "error": (
                    _validation_error_message(
                        error
                    )
                ),
            },
            status=400,
        )

    door_data = None

    if incident.door_shift_id:
        incident.door_shift.refresh_from_db(
            fields=[
                "state",
                "notes",
                "updated_at",
            ]
        )

        door_data = {
            "id": incident.door_shift.id,
            "door_number": (
                incident.door_shift.door_number
            ),
            "state": (
                incident.door_shift.state
            ),
            "state_label": (
                incident
                .door_shift
                .get_state_display()
            ),
            "notes": (
                incident.door_shift.notes
                or ""
            ),
            "updated_at": (
                incident
                .door_shift
                .updated_at
                .strftime(
                    "%Y-%m-%d %H:%M"
                )
                if incident.door_shift.updated_at
                else None
            ),
        }

    return JsonResponse(
        {
            "success": True,

            "incident": {
                "id": incident.id,
                "incident_number": (
                    incident.incident_number
                ),
                "status": incident.status,
                "status_label": (
                    incident.get_status_display()
                ),
                "priority": incident.priority,
                "priority_label": (
                    incident.get_priority_display()
                ),
                "incident_type": (
                    incident.incident_type
                ),
                "incident_type_label": (
                    incident
                    .get_incident_type_display()
                ),
                "assigned_to_name": (
                    incident.assigned_to_name
                    or ""
                ),
                "closing_notes": (
                    incident.closing_notes
                    or ""
                ),
                "closed_by": (
                    incident.closed_by.username
                    if incident.closed_by
                    else ""
                ),
                "closed_at": (
                    incident.closed_at.strftime(
                        "%Y-%m-%d %H:%M"
                    )
                    if incident.closed_at
                    else None
                ),
            },

            "door": door_data,
        }
    )


@login_required
@require_POST
def escalate_incident_ajax(request, pk):
    _require_ops_permission(request, PlatformPermissions.ESCALATE_INCIDENT)
    incident = get_object_or_404(
        _scoped_by_section(Incident.objects, request.user), pk=pk
    )
    try:
        incident = IncidentRoutingService.escalate_incident(
            incident, request.user, request.POST.get("note", "")
        )
    except ValidationError as error:
        return JsonResponse(
            {"success": False, "error": _validation_error_message(error)}, status=400
        )
    return JsonResponse({
        "success": True,
        "escalation_level": incident.escalation_level,
        "escalation_label": incident.get_escalation_level_display(),
    })


@login_required
@require_POST
def convert_incident_to_maintenance_ajax(request, pk):
    _require_ops_permission(
        request, PlatformPermissions.CONVERT_INCIDENT_TO_MAINTENANCE
    )
    incident = get_object_or_404(
        _scoped_by_section(Incident.objects, request.user), pk=pk
    )
    planned_start_at = _parse_planned_datetime(request.POST.get("planned_start_at"))
    planned_end_at = _parse_planned_datetime(request.POST.get("planned_end_at"))
    try:
        maintenance = IncidentRoutingService.convert_to_maintenance(
            incident,
            request,
            planned_start_at,
            planned_end_at,
            actor=request.user,
        )
    except ValidationError as error:
        return JsonResponse(
            {"success": False, "error": _validation_error_message(error)}, status=400
        )
    return JsonResponse({
        "success": True,
        "request_number": maintenance.request_number,
        "status": maintenance.status,
    })


@login_required
@require_POST
def add_incident_shift_update_ajax(request, pk):
    _require_ops_permission(request, PlatformPermissions.UPDATE_INCIDENT)
    incident = get_object_or_404(
        _scoped_by_section(Incident.objects, request.user), pk=pk
    )
    try:
        event = IncidentRoutingService.add_shift_update(
            incident, request.user, request.POST.get("note", "")
        )
    except ValidationError as error:
        return JsonResponse(
            {"success": False, "error": _validation_error_message(error)}, status=400
        )
    return JsonResponse({
        "success": True,
        "event": {"id": event.pk, "note": event.note, "created_at": event.created_at.isoformat()},
    })


# =========================================================
# سجل العمليات المباشرة
# =========================================================


@login_required
def activity_log_view(request):
    """
    عرض سجل عمليات الأبواب والصيانة والبلاغات والتوزيع.
    """

    _require_ops_permission(request, PlatformPermissions.VIEW_SYSTEM_LOGS)

    query = (
        request.GET.get(
            "q",
            "",
        )
        or ""
    ).strip()

    module_filter = (
        request.GET.get(
            "module",
            "",
        )
        or ""
    ).strip()

    action_filter = (
        request.GET.get(
            "action",
            "",
        )
        or ""
    ).strip()

    user_filter = (
        request.GET.get(
            "user",
            "",
        )
        or ""
    ).strip()

    date_from = (
        request.GET.get(
            "date_from",
            "",
        )
        or ""
    ).strip()

    date_to = (
        request.GET.get(
            "date_to",
            "",
        )
        or ""
    ).strip()

    logs = (
        SystemActivityLog.objects
        .select_related("user")
        .filter(
            module__in=LIVE_OPERATION_MODULES,
        )
        .order_by("-created_at")
    )

    if query:
        logs = logs.filter(
            Q(
                description__icontains=query
            )
            | Q(
                module__icontains=query
            )
            | Q(
                user__username__icontains=query
            )
            | Q(
                user__first_name__icontains=query
            )
            | Q(
                user__last_name__icontains=query
            )
            | Q(
                ip_address__icontains=query
            )
        )

    if (
        module_filter
        and module_filter in LIVE_OPERATION_MODULES
    ):
        logs = logs.filter(
            module=module_filter,
        )

    valid_actions = {
        value
        for value, _label
        in SystemActivityLog.ActionType.choices
    }

    if (
        action_filter
        and action_filter in valid_actions
    ):
        logs = logs.filter(
            action=action_filter,
        )

    if user_filter.isdigit():
        logs = logs.filter(
            user_id=int(user_filter),
        )

    if date_from:
        logs = logs.filter(
            created_at__date__gte=date_from,
        )

    if date_to:
        logs = logs.filter(
            created_at__date__lte=date_to,
        )

    all_logs = (
        SystemActivityLog.objects
        .filter(
            module__in=LIVE_OPERATION_MODULES,
        )
    )

    operational_users = (
        all_logs
        .filter(
            user__isnull=False,
        )
        .values(
            "user_id",
            "user__username",
            "user__first_name",
            "user__last_name",
        )
        .distinct()
        .order_by(
            "user__first_name",
            "user__last_name",
            "user__username",
        )
    )

    paginator = Paginator(
        logs,
        25,
    )

    page_obj = paginator.get_page(
        request.GET.get("page")
    )

    today = timezone.localdate()
    filtered_logs_count = logs.count()

    context = {
        "logs": page_obj,
        "page_obj": page_obj,

        "q": query,
        "selected_module": module_filter,
        "selected_action": action_filter,
        "selected_user": user_filter,
        "date_from": date_from,
        "date_to": date_to,

        "module_choices": (
            LIVE_OPERATION_MODULES
        ),

        "action_choices": (
            SystemActivityLog
            .ActionType
            .choices
        ),

        "operational_users": (
            operational_users
        ),

        "total_logs": (
            all_logs.count()
        ),

        "today_logs": (
            all_logs.filter(
                created_at__date=today,
            ).count()
        ),
        "filtered_logs_count": filtered_logs_count,
        "active_users_today": all_logs.filter(
            created_at__date=today,
            user__isnull=False,
        ).values("user_id").distinct().count(),

        "door_logs": (
            all_logs.filter(
                module="الأبواب",
            ).count()
        ),

        "maintenance_logs": (
            all_logs.filter(
                module="الصيانة",
            ).count()
        ),

        "incident_logs": (
            all_logs.filter(
                module="البلاغات",
            ).count()
        ),

        "distribution_logs": (
            all_logs.filter(
                module="توزيع الأبواب",
            ).count()
        ),
    }

    return render(
        request,
        "ops/activity_log.html",
        context,
    )


@login_required
def export_activity_log_excel_view(request):
    """تصدير سجل العمليات المباشرة إلى Excel مع تطبيق الفلاتر الحالية."""
    _require_ops_permission(request, PlatformPermissions.VIEW_SYSTEM_LOGS)
    logs = SystemActivityLog.objects.select_related("user").filter(
        module__in=LIVE_OPERATION_MODULES
    ).order_by("-created_at")
    query = (request.GET.get("q") or "").strip()
    module = (request.GET.get("module") or "").strip()
    action = (request.GET.get("action") or "").strip()
    user_id = (request.GET.get("user") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()
    if query:
        logs = logs.filter(
            Q(description__icontains=query) | Q(module__icontains=query)
            | Q(user__username__icontains=query) | Q(ip_address__icontains=query)
        )
    if module in LIVE_OPERATION_MODULES:
        logs = logs.filter(module=module)
    if action in {value for value, _ in SystemActivityLog.ActionType.choices}:
        logs = logs.filter(action=action)
    if user_id.isdigit():
        logs = logs.filter(user_id=int(user_id))
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "سجل العمليات"
    sheet.sheet_view.rightToLeft = True
    sheet.merge_cells("A1:F2")
    sheet["A1"] = "منصة أبواب | سجل العمليات المباشرة"
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0B6B50")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    headers = ["التاريخ", "القسم", "الإجراء", "الوصف", "المستخدم", "عنوان IP"]
    sheet.append([])
    sheet.append(headers)
    for cell in sheet[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="B89A49")
        cell.alignment = Alignment(horizontal="center")
    for log in logs.iterator():
        user_name = "النظام"
        if log.user:
            user_name = log.user.get_full_name() or log.user.username
        sheet.append([
            timezone.localtime(log.created_at).strftime("%Y-%m-%d %H:%M:%S"),
            log.module, log.get_action_display(), log.description,
            user_name, str(log.ip_address or "—"),
        ])
    widths = [22, 20, 16, 65, 24, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:F{max(sheet.max_row, 4)}"
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="live_operations_log.xlsx"'
    return response

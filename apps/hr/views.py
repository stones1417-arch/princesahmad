from __future__ import annotations
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import IntegrityError
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from apps.accounts.role_permissions import (
    user_can_access_object,
    user_can_access_section,
)
from apps.roles.services.access_control import user_has_permission
from apps.roles.services.section_access import (
    can_view_section,
    filter_employees_for_user,
    has_institutional_scope,
)

from .forms import EmployeeForm
from .models import Employee
from .services import EmployeeService


def _require_employee_permission(request, permission_code: str):
    """Require an institutional permission and an assigned role scope."""
    if not user_has_permission(request.user, permission_code):
        raise PermissionDenied("لا تملك الصلاحية المطلوبة")

    if (
        not request.user.is_superuser
        and not has_institutional_scope(request.user)
    ):
        raise PermissionDenied("لا تملك نطاقًا تشغيليًا معتمدًا")


def _get_accessible_employee_or_404(request, pk):
    """Resolve an employee only after server-side scope authorization."""
    employee = get_object_or_404(
        Employee.objects.select_related("user"),
        pk=pk,
    )

    if has_institutional_scope(request.user):
        if not user_can_access_object(request.user, employee):
            raise PermissionDenied("لا تملك صلاحية الوصول إلى هذا الموظف")

    return employee


def _require_requested_employee_section(request, form):
    """Reject create or update requests outside the active role scope."""
    requested_section = form.cleaned_data["operational_section"]

    if not user_can_access_section(request.user, requested_section):
        raise PermissionDenied("لا تملك صلاحية إدارة هذا القسم التشغيلي")


@login_required
def employee_list_view(request):
    _require_employee_permission(
        request,
        "roles.view_employees",
    )

    q = (request.GET.get("q") or "").strip()
    job_title = (request.GET.get("job_title") or "").strip()
    work_status = (request.GET.get("work_status") or "").strip()
    selected_operational_section = (
        request.GET.get("operational_section") or ""
    ).strip()
    active_filter = (request.GET.get("active") or "").strip()
    two_factor_filter = (request.GET.get("two_factor_readiness") or "").strip()
    selected_sort = (request.GET.get("sort") or "").strip()
    selected_direction = (request.GET.get("direction") or "").strip().lower()

    can_view_male = can_view_section(
        request.user,
        Employee.OperationalSection.MALE,
    )
    can_view_female = can_view_section(
        request.user,
        Employee.OperationalSection.FEMALE,
    )

    employees = Employee.objects.select_related("user").all()

    employees = filter_employees_for_user(
        employees,
        request.user,
    )

    gender_query_map = {
        "رجالي": Employee.OperationalSection.MALE,
        "الرجالي": Employee.OperationalSection.MALE,
        "male": Employee.OperationalSection.MALE,
        "نسائي": Employee.OperationalSection.FEMALE,
        "النسائي": Employee.OperationalSection.FEMALE,
        "female": Employee.OperationalSection.FEMALE,
    }

    if q:
        q_lower = q.lower()
        q_filters = (
            Q(full_name__icontains=q)
            | Q(employee_number__icontains=q)
            | Q(phone_number__icontains=q)
            | Q(national_id__icontains=q)
            | Q(email__icontains=q)
            | Q(user__username__icontains=q)
        )

        if q_lower in gender_query_map:
            q_filters |= Q(
                operational_section=gender_query_map[q_lower]
            )

        employees = employees.filter(q_filters)

    if job_title:
        employees = employees.filter(job_title=job_title)

    if work_status:
        employees = employees.filter(work_status=work_status)

    if selected_operational_section in dict(Employee.OperationalSection.choices):
        employees = employees.filter(operational_section=selected_operational_section)

    if active_filter == "active":
        employees = employees.filter(is_active=True)
    elif active_filter == "inactive":
        employees = employees.filter(is_active=False)

    sort_fields = {
        "employee_number": "employee_number",
        "full_name": "full_name",
        "operational_section": "operational_section",
        "is_active": "is_active",
    }

    order_field = sort_fields.get(
        selected_sort,
        "employee_number",
    )

    if selected_direction == "desc":
        order_field = f"-{order_field}"

    employees = employees.order_by(order_field)

    if two_factor_filter in {"ready", "not-ready"}:
        matching_ids = [
            employee.pk
            for employee in employees
            if bool(employee.two_factor_readiness["channels"])
            == (two_factor_filter == "ready")
        ]
        employees = employees.filter(pk__in=matching_ids)

    accessible_employees = Employee.objects.all()
    accessible_employees = filter_employees_for_user(
        accessible_employees,
        request.user,
    )
    context = {
        "employees": employees,
        "q": q,
        "selected_job_title": job_title,
        "selected_work_status": work_status,
        "selected_operational_section": selected_operational_section,
        "selected_active": active_filter,
        "selected_two_factor_readiness": two_factor_filter,
        "selected_sort": selected_sort,
        "selected_direction": selected_direction,
        "sort_choices": (
            ("", "الترتيب الافتراضي"),
            ("employee_number", "الرقم الوظيفي"),
            ("full_name", "الاسم"),
            ("operational_section", "القسم التشغيلي"),
            ("is_active", "الحالة"),
        ),
        "job_title_choices": Employee.JobTitle.choices,
        "work_status_choices": Employee.WorkStatus.choices,
        "employee_operational_section_choices": [
            choice
            for choice in Employee.OperationalSection.choices
            if (
                choice[0] == Employee.OperationalSection.MALE
                and can_view_male
            )
            or (
                choice[0] == Employee.OperationalSection.FEMALE
                and can_view_female
            )
        ],
        "can_view_male": can_view_male,
        "can_view_female": can_view_female,
        "total_employees": accessible_employees.count(),
        "male_employees": accessible_employees.filter(
            operational_section=Employee.OperationalSection.MALE,
        ).count(),
        "female_employees": accessible_employees.filter(
            operational_section=Employee.OperationalSection.FEMALE,
        ).count(),
        "active_employees": accessible_employees.filter(
            is_active=True,
        ).count(),
        "inactive_employees": accessible_employees.filter(
            is_active=False,
        ).count(),
        "maintenance_members": accessible_employees.filter(
            can_execute_maintenance=True,
        ).count(),
        "door_assignable_members": accessible_employees.filter(
            is_active=True,
            can_work_on_doors=True,
        ).count(),
        "job_title_stats": (
            accessible_employees
            .values("job_title")
            .annotate(total=Count("id"))
            .order_by("-total")
        ),
    }

    return render(request, "hr/employee_list.html", context)


@login_required
def employee_create_view(request):
    _require_employee_permission(
        request,
        "roles.create_employee",
    )

    if request.method == "POST":
        form = EmployeeForm(request.POST)

        if form.is_valid():
            _require_requested_employee_section(request, form)
            try:
                employee = EmployeeService.create(
                    request=request,
                    form=form,
                )

                messages.success(
                    request,
                    f"تم إضافة الموظف {employee.full_name} بنجاح"
                )
                return redirect("hr:list")

            except IntegrityError:
                form.add_error("employee_number", "الرقم الوظيفي مستخدم مسبقًا")
                messages.error(request, "الرقم الوظيفي مستخدم مسبقًا")
        else:
            messages.error(request, "تحقق من الحقول المدخلة")
    else:
        form = EmployeeForm()

    return render(
        request,
        "hr/employee_form.html",
        {
            "title": "إضافة موظف",
            "form": form,
        }
    )


@login_required
def employee_update_view(request, pk):
    _require_employee_permission(
        request,
        "roles.update_employee",
    )

    employee = _get_accessible_employee_or_404(request, pk)

    if request.method == "POST":
        form = EmployeeForm(request.POST, instance=employee)

        if form.is_valid():
            _require_requested_employee_section(request, form)
            try:
                employee = EmployeeService.update(
                    request=request,
                    form=form,
                )

                messages.success(
                    request,
                    f"تم تحديث بيانات الموظف {employee.full_name}"
                )
                return redirect("hr:list")

            except IntegrityError:
                form.add_error("employee_number", "الرقم الوظيفي مستخدم مسبقًا")
                messages.error(request, "الرقم الوظيفي مستخدم مسبقًا")
        else:
            messages.error(request, "تحقق من الحقول المدخلة")
    else:
        form = EmployeeForm(instance=employee)

    return render(
        request,
        "hr/employee_form.html",
        {
            "title": "تعديل موظف",
            "form": form,
            "employee": employee,
            "two_factor_readiness": employee.two_factor_readiness,
        }
    )


@login_required
@require_POST
def employee_toggle_active_view(request, pk):
    _require_employee_permission(
        request,
        "roles.disable_employee",
    )

    employee = _get_accessible_employee_or_404(request, pk)
    was_active = employee.is_active

    employee = EmployeeService.toggle_active(
        request=request,
        employee=employee,
    )

    if was_active:
        messages.warning(request, f"تم تعطيل الموظف {employee.full_name}")
    else:
        messages.success(request, f"تم تفعيل الموظف {employee.full_name}")

    return redirect("hr:list")


@login_required
def employee_delete_view(request, pk):
    _require_employee_permission(
        request,
        "roles.disable_employee",
    )

    employee = _get_accessible_employee_or_404(request, pk)

    if request.method == "POST":
        employee = EmployeeService.safe_delete(
            request=request,
            employee=employee,
        )

        messages.warning(
            request,
            f"تم تعطيل الموظف {employee.full_name} بدل الحذف النهائي"
        )
        return redirect("hr:list")

    return render(
        request,
        "hr/employee_confirm_delete.html",
        {
            "employee": employee,
        }
    )


@login_required
def export_employees_excel(request):
    _require_employee_permission(
        request,
        "roles.export_report",
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "سجل الموظفين"

    headers = [
        "الرقم الوظيفي",
        "الاسم الكامل",
        "رقم الهوية",
        "رقم الجوال",
        "البريد الإلكتروني",
        "المسمى الوظيفي",
        "حالة الموظف",
        "نشط في النظام",
        "يمكن تسكينه على الأبواب",
        "يمكنه تنفيذ الصيانة",
        "تاريخ المباشرة",
        "ملاحظات",
        "تاريخ الإضافة",
        "آخر تحديث",
    ]

    for col, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = title
        cell.font = Font(bold=True)

    employees = filter_employees_for_user(
        Employee.objects.select_related("user"),
        request.user,
    ).order_by("employee_number")

    for row, employee in enumerate(employees, start=2):
        sheet.cell(row=row, column=1).value = employee.employee_number
        sheet.cell(row=row, column=2).value = employee.full_name
        sheet.cell(row=row, column=3).value = employee.national_id
        sheet.cell(row=row, column=4).value = employee.phone_number
        sheet.cell(row=row, column=5).value = employee.email
        sheet.cell(row=row, column=6).value = employee.get_job_title_display()
        sheet.cell(row=row, column=7).value = employee.get_work_status_display()
        sheet.cell(row=row, column=8).value = "نعم" if employee.is_active else "لا"
        sheet.cell(row=row, column=9).value = "نعم" if employee.can_work_on_doors else "لا"
        sheet.cell(row=row, column=10).value = "نعم" if employee.can_execute_maintenance else "لا"
        sheet.cell(row=row, column=11).value = (
            employee.hire_date.strftime("%Y-%m-%d")
            if employee.hire_date
            else ""
        )
        sheet.cell(row=row, column=12).value = employee.notes
        sheet.cell(row=row, column=13).value = employee.created_at.strftime("%Y-%m-%d %H:%M")
        sheet.cell(row=row, column=14).value = employee.updated_at.strftime("%Y-%m-%d %H:%M")

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        sheet.column_dimensions[column_letter].width = min(max_length + 4, 35)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="employees.xlsx"'

    workbook.save(response)

    return response


@login_required
@require_POST
def employee_toggle_active_ajax_view(request, pk):
    _require_employee_permission(
        request,
        "roles.disable_employee",
    )

    employee = _get_accessible_employee_or_404(request, pk)

    employee = EmployeeService.toggle_active(
        request=request,
        employee=employee,
    )

    return JsonResponse(
        {
            "success": True,
            "employee_id": employee.pk,
            "is_active": employee.is_active,
            "status_text": "نشط" if employee.is_active else "معطل",
            "message": (
                f"تم تفعيل الموظف {employee.full_name} بنجاح"
                if employee.is_active
                else f"تم تعطيل الموظف {employee.full_name} بنجاح"
            ),
        }
    )


@login_required
@require_POST
def employee_delete_ajax_view(request, pk):
    _require_employee_permission(
        request,
        "roles.disable_employee",
    )

    employee = _get_accessible_employee_or_404(request, pk)

    employee = EmployeeService.safe_delete(
        request=request,
        employee=employee,
    )

    return JsonResponse(
        {
            "success": True,
            "employee_id": employee.pk,
            "is_active": employee.is_active,
            "status_text": "معطل",
            "message": (
                f"تم تعطيل الموظف {employee.full_name} "
                "بدل الحذف النهائي"
            ),
        }
    )
    
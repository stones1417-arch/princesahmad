from io import BytesIO

from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.core.tests.factories import create_employee, create_user
from apps.hr.models import Employee
from apps.roles.models import Role
from apps.roles.services.role_manager import assign_role_to_user


class RolePermissionsTests(TestCase):
    def setUp(self):
        call_command("setup_roles")

    def assign_role(self, user, role_code, section=None):
        role = Role.objects.get(code=role_code)
        if section:
            role.operational_section = section
            role.save(update_fields=["operational_section", "updated_at"])
        assign_role_to_user(user=user, role_code=role_code)
        return role

    def employee_data(self, *, section, employee_number="94001"):
        return {
            "employee_number": employee_number,
            "full_name": "موظف اختبار الصلاحيات",
            "operational_section": section,
            "job_title": Employee.JobTitle.MONITOR,
            "work_status": Employee.WorkStatus.ACTIVE,
            "is_active": "on",
            "can_work_on_doors": "on",
        }

    def test_setup_roles_is_idempotent(self):
        first_count = Role.objects.filter(is_system_role=True).count()

        call_command("setup_roles")

        self.assertEqual(
            Role.objects.filter(is_system_role=True).count(),
            first_count,
        )
        self.assertEqual(Role.objects.filter(code="hr_manager").count(), 1)

    def test_setup_roles_preserves_existing_scope_and_permissions(self):
        role = Role.objects.get(code="hr_manager")
        role.operational_section = Role.OperationalSection.MALE
        role.save(update_fields=["operational_section", "updated_at"])
        original_permissions = set(
            role.group.permissions.values_list("codename", flat=True)
        )

        call_command("setup_roles")

        role.refresh_from_db()
        self.assertEqual(
            role.operational_section,
            Role.OperationalSection.MALE,
        )
        self.assertEqual(
            set(role.group.permissions.values_list("codename", flat=True)),
            original_permissions,
        )

    def test_system_admin_has_administrative_permissions(self):
        user = create_user(username="system-admin")
        self.assign_role(user, "system_admin")

        self.assertTrue(user.has_perm("roles.manage_users"))
        self.assertTrue(user.has_perm("roles.manage_roles"))
        self.assertFalse(user.is_staff)

    def test_employee_cannot_manage_system(self):
        user = create_user(username="employee")
        self.assign_role(user, "employee")

        self.assertFalse(user.has_perm("roles.manage_users"))
        self.assertFalse(user.has_perm("roles.manage_roles"))

    def test_shift_supervisor_cannot_manage_system(self):
        user = create_user(username="shift-supervisor")
        self.assign_role(user, "shift_supervisor")

        self.assertTrue(user.has_perm("roles.assign_employees"))
        self.assertFalse(user.has_perm("roles.manage_users"))

    def test_hr_role_enforces_section_scope_for_direct_employee_url(self):
        male_employee = create_employee(
            operational_section=Employee.OperationalSection.MALE,
        )
        female_employee = create_employee(
            operational_section=Employee.OperationalSection.FEMALE,
        )
        user = create_user(username="male-hr-manager")
        self.assign_role(user, "hr_manager", Role.OperationalSection.MALE)
        self.client.force_login(user)

        allowed_response = self.client.get(
            reverse("hr:update", args=[male_employee.pk])
        )
        denied_response = self.client.get(
            reverse("hr:update", args=[female_employee.pk])
        )

        self.assertEqual(allowed_response.status_code, 200)
        self.assertEqual(denied_response.status_code, 403)

    def test_user_without_role_is_denied_direct_employee_update(self):
        employee = create_employee(
            operational_section=Employee.OperationalSection.MALE,
        )
        user = create_user(username="unprivileged-user")
        self.client.force_login(user)

        response = self.client.get(reverse("hr:update", args=[employee.pk]))

        self.assertEqual(response.status_code, 403)

    def test_staff_without_platform_permission_cannot_update_or_disable(self):
        employee = create_employee(
            operational_section=Employee.OperationalSection.MALE,
        )
        user = create_user(username="staff-without-role", is_staff=True)
        self.client.force_login(user)

        update_response = self.client.get(
            reverse("hr:update", args=[employee.pk])
        )
        disable_response = self.client.post(
            reverse("hr:toggle-active", args=[employee.pk])
        )
        delete_response = self.client.post(
            reverse("hr:delete", args=[employee.pk])
        )

        self.assertEqual(update_response.status_code, 403)
        self.assertEqual(disable_response.status_code, 403)
        self.assertEqual(delete_response.status_code, 403)
        employee.refresh_from_db()
        self.assertTrue(employee.is_active)

    def test_male_role_cannot_create_female_employee(self):
        user = create_user(username="male-create")
        self.assign_role(user, "hr_manager", Role.OperationalSection.MALE)
        self.client.force_login(user)

        response = self.client.post(
            reverse("hr:create"),
            self.employee_data(
                section=Employee.OperationalSection.FEMALE,
                employee_number="94002",
            ),
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Employee.objects.filter(employee_number="94002").exists())

    def test_female_role_cannot_create_male_employee(self):
        user = create_user(username="female-create")
        self.assign_role(user, "hr_manager", Role.OperationalSection.FEMALE)
        self.client.force_login(user)

        response = self.client.post(
            reverse("hr:create"),
            self.employee_data(
                section=Employee.OperationalSection.MALE,
                employee_number="94003",
            ),
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Employee.objects.filter(employee_number="94003").exists())

    def test_male_role_cannot_move_male_employee_to_female(self):
        employee = create_employee(
            employee_number="94004",
            operational_section=Employee.OperationalSection.MALE,
        )
        user = create_user(username="male-update")
        self.assign_role(user, "hr_manager", Role.OperationalSection.MALE)
        self.client.force_login(user)

        response = self.client.post(
            reverse("hr:update", args=[employee.pk]),
            self.employee_data(
                section=Employee.OperationalSection.FEMALE,
                employee_number=employee.employee_number,
            ),
        )

        self.assertEqual(response.status_code, 403)
        employee.refresh_from_db()
        self.assertEqual(employee.operational_section, Employee.OperationalSection.MALE)

    def test_female_role_cannot_move_female_employee_to_male(self):
        employee = create_employee(
            employee_number="94005",
            operational_section=Employee.OperationalSection.FEMALE,
        )
        user = create_user(username="female-update")
        self.assign_role(user, "hr_manager", Role.OperationalSection.FEMALE)
        self.client.force_login(user)

        response = self.client.post(
            reverse("hr:update", args=[employee.pk]),
            self.employee_data(
                section=Employee.OperationalSection.MALE,
                employee_number=employee.employee_number,
            ),
        )

        self.assertEqual(response.status_code, 403)
        employee.refresh_from_db()
        self.assertEqual(employee.operational_section, Employee.OperationalSection.FEMALE)

    def test_all_scope_can_change_employee_section(self):
        employee = create_employee(
            employee_number="94006",
            operational_section=Employee.OperationalSection.MALE,
        )
        user = create_user(username="all-update")
        self.assign_role(user, "hr_manager", Role.OperationalSection.ALL)
        self.client.force_login(user)

        response = self.client.post(
            reverse("hr:update", args=[employee.pk]),
            self.employee_data(
                section=Employee.OperationalSection.FEMALE,
                employee_number=employee.employee_number,
            ),
        )

        self.assertRedirects(response, reverse("hr:list"))
        employee.refresh_from_db()
        self.assertEqual(employee.operational_section, Employee.OperationalSection.FEMALE)

    def test_export_is_filtered_by_section_scope(self):
        male_employee = create_employee(
            full_name="موظف للتصدير الرجالي",
            employee_number="94007",
            operational_section=Employee.OperationalSection.MALE,
        )
        female_employee = create_employee(
            full_name="موظفة للتصدير النسائي",
            employee_number="94008",
            operational_section=Employee.OperationalSection.FEMALE,
        )
        user = create_user(username="male-export")
        self.assign_role(user, "hr_manager", Role.OperationalSection.MALE)
        self.client.force_login(user)

        response = self.client.get(reverse("hr:export-excel"))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        names = {
            row[1]
            for row in workbook.active.iter_rows(
                min_row=2,
                values_only=True,
            )
        }
        self.assertIn(male_employee.full_name, names)
        self.assertNotIn(female_employee.full_name, names)

    def test_female_export_excludes_male_and_unprivileged_user_is_denied(self):
        male_employee = create_employee(
            full_name="موظف تصدير محجوب",
            employee_number="94009",
            operational_section=Employee.OperationalSection.MALE,
        )
        female_employee = create_employee(
            full_name="موظفة تصدير مسموحة",
            employee_number="94010",
            operational_section=Employee.OperationalSection.FEMALE,
        )
        user = create_user(username="female-export")
        self.assign_role(user, "hr_manager", Role.OperationalSection.FEMALE)
        self.client.force_login(user)

        response = self.client.get(reverse("hr:export-excel"))

        workbook = load_workbook(BytesIO(response.content))
        names = {
            row[1]
            for row in workbook.active.iter_rows(
                min_row=2,
                values_only=True,
            )
        }
        self.assertIn(female_employee.full_name, names)
        self.assertNotIn(male_employee.full_name, names)

        self.client.force_login(create_user(username="no-export"))
        denied_response = self.client.get(reverse("hr:export-excel"))
        self.assertEqual(denied_response.status_code, 403)
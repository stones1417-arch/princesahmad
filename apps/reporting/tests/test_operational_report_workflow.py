from __future__ import annotations

import tempfile
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.db import connection
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from apps.core.tests.factories import create_door, create_shift_plan
from apps.exports_center.models import ExportLog
from apps.locations.door_directions import OFFICIAL_DOOR_CODES
from apps.ops.models import DoorShift
from apps.reporting.models import ShiftReport
from apps.reporting.services import ReportService
from apps.roles.models import Role, UserRole


class OperationalReportWorkflowTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="report-workflow-admin"
        )
        self.client.force_login(self.user)
        self.shift = create_shift_plan(is_active=True, is_finished=False)
        for code in OFFICIAL_DOOR_CODES:
            create_door(door_number=code)
            DoorShift.objects.create(
                shift_plan=self.shift,
                door_number=code,
                state=DoorShift.DoorState.OPEN,
                is_active=True,
            )

    def _finish_shift(self):
        self.shift.is_active = False
        self.shift.is_finished = True
        self.shift.save(update_fields=["is_active", "is_finished"])
        self.shift.refresh_from_db()

    def test_active_shift_is_rejected_and_finished_shift_uses_one_snapshot(self):
        response = self.client.post(
            reverse("reporting:generate", args=[self.shift.pk])
        )
        self.assertEqual(
            response.status_code,
            302,
            getattr(response.context.get("form"), "errors", "")
            if response.context
            else "",
        )
        self.assertFalse(ShiftReport.objects.exists())

        self._finish_shift()
        response = self.client.post(
            reverse("reporting:create-operational"),
            {
                "report_type": ShiftReport.ReportType.OPERATIONAL,
                "shift_plan": self.shift.pk,
                "summary": "",
                "recommendations": "",
            },
        )
        self.assertEqual(
            response.status_code,
            302,
            getattr(response.context.get("form"), "errors", "")
            if response.context
            else "",
        )
        report = ShiftReport.objects.get()
        self.assertEqual(report.shift_plan, self.shift)
        self.assertEqual(report.status, ShiftReport.ReportStatus.FINAL)
        self.assertEqual(report.total_doors, 42)
        snapshot_codes = {
            item["door_number"] for item in report.snapshot_data["doors"]
        }
        self.assertEqual(snapshot_codes, set(OFFICIAL_DOOR_CODES))
        self.assertIn("6A", snapshot_codes)
        self.assertIn("6B", snapshot_codes)

        duplicate = self.client.post(
            reverse("reporting:generate", args=[self.shift.pk])
        )
        self.assertEqual(duplicate.status_code, 302)
        self.assertEqual(ShiftReport.objects.count(), 1)

    def test_view_approval_pdf_excel_and_export_log(self):
        self._finish_shift()
        report = ReportService.generate_shift_report(
            shift_plan=self.shift,
            user=self.user,
        )
        draft = ShiftReport.objects.create(
            report_type=ShiftReport.ReportType.MANUAL,
            summary="Draft report",
            created_by=self.user,
        )
        self.assertEqual(
            self.client.get(reverse("reporting:detail", args=[draft.pk])).status_code,
            200,
        )
        detail = self.client.get(reverse("reporting:detail", args=[report.pk]))
        self.assertEqual(detail.status_code, 200)

        approved = self.client.post(
            reverse("reporting:approve", args=[report.pk])
        )
        self.assertEqual(approved.status_code, 302)
        report.refresh_from_db()
        self.assertEqual(report.status, ShiftReport.ReportStatus.APPROVED)
        self.assertEqual(
            self.client.get(reverse("reporting:detail", args=[report.pk])).status_code,
            200,
        )

        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                pdf = self.client.get(
                    reverse("reporting:export-pdf", args=[report.pk])
                )
                self.assertEqual(pdf.status_code, 200)
                self.assertEqual(pdf["Content-Type"], "application/pdf")
                self.assertIn("attachment", pdf["Content-Disposition"])
                self.assertTrue(pdf.content.startswith(b"%PDF"))
                self.assertGreater(len(pdf.content), 100)

                excel = self.client.get(
                    reverse("reporting:export-excel", args=[report.pk])
                )
                self.assertEqual(excel.status_code, 200)
                self.assertGreater(len(excel.content), 100)
                workbook_path = Path(media_root) / "report.xlsx"
                workbook_path.write_bytes(excel.content)
                workbook = load_workbook(workbook_path, read_only=True)
                self.assertIn("الملخص التنفيذي", workbook.sheetnames)
                workbook.close()

                logs = ExportLog.objects.order_by("created_at")
                self.assertEqual(logs.count(), 2)
                self.assertTrue(all(log.status == "success" for log in logs))
                self.assertTrue(all(log.download_count == 1 for log in logs))
                self.assertTrue(all(log.file for log in logs))
                self.assertTrue(all(log.storage_path for log in logs))
                self.assertTrue(all(log.file.storage.exists(log.file.name) for log in logs))

    def test_create_and_approve_buttons_match_post_contract(self):
        self._finish_shift()
        list_page = self.client.get(reverse("reporting:list"))
        self.assertContains(list_page, reverse("reporting:create-operational"))
        self.assertNotContains(list_page, 'href="/reporting/generate-active/"')
        legacy = self.client.get(reverse("reporting:generate-active"))
        self.assertRedirects(legacy, reverse("reporting:create-operational"))

        report = ReportService.generate_shift_report(
            shift_plan=self.shift,
            user=self.user,
        )
        detail = self.client.get(reverse("reporting:detail", args=[report.pk]))
        approve_url = reverse("reporting:approve", args=[report.pk])
        self.assertContains(detail, f'action="{approve_url}"')
        self.assertNotContains(detail, f'href="{approve_url}"')
        self.assertEqual(self.client.get(approve_url).status_code, 405)

        draft = ShiftReport.objects.create(
            report_type=ShiftReport.ReportType.MANUAL,
            summary="Draft approval regression",
            created_by=self.user,
        )
        draft_response = self.client.post(
            reverse("reporting:approve", args=[draft.pk])
        )
        self.assertEqual(draft_response.status_code, 302)
        draft.refresh_from_db()
        self.assertEqual(draft.status, ShiftReport.ReportStatus.DRAFT)
        self.assertIsNone(draft.approved_by_id)
        self.assertIsNone(draft.approved_at)

    def test_platform_report_role_can_approve_without_legacy_model_permission(self):
        self._finish_shift()
        report = ReportService.generate_shift_report(
            shift_plan=self.shift,
            user=self.user,
        )
        group = Group.objects.create(name="institutional-report-approver")
        group.permissions.add(
            *Permission.objects.filter(
                content_type__app_label="roles",
                codename__in=("view_reports", "approve_report"),
            )
        )
        role = Role.objects.create(
            code="institutional-report-approver",
            name="Institutional report approver",
            group=group,
            operational_section=Role.OperationalSection.ALL,
        )
        approver = get_user_model().objects.create_user(
            username="institutional-report-approver"
        )
        UserRole.objects.create(user=approver, role=role)
        self.assertFalse(
            approver.has_perm("reporting.can_approve_shift_report")
        )
        self.client.force_login(approver)
        response = self.client.post(
            reverse("reporting:approve", args=[report.pk])
        )
        self.assertEqual(response.status_code, 302)
        report.refresh_from_db()
        self.assertEqual(report.status, ShiftReport.ReportStatus.APPROVED)
        self.assertEqual(report.approved_by, approver)
        self.assertIsNotNone(report.approved_at)
        repeated = self.client.post(
            reverse("reporting:approve", args=[report.pk])
        )
        self.assertEqual(repeated.status_code, 302)
        report.refresh_from_db()
        self.assertEqual(report.status, ShiftReport.ReportStatus.APPROVED)

    def test_approval_locks_only_report_row_without_nullable_joins(self):
        self._finish_shift()
        report = ReportService.generate_shift_report(
            shift_plan=self.shift,
            user=self.user,
        )
        self.assertEqual(report.status, ShiftReport.ReportStatus.FINAL)
        self.assertIsNone(report.approved_by_id)
        self.assertIsNone(report.approved_at)

        executed_sql = []

        def capture_sql(execute, sql, params, many, context):
            executed_sql.append(sql)
            return execute(sql, params, many, context)

        with connection.execute_wrapper(capture_sql):
            response = self.client.post(
                reverse("reporting:approve", args=[report.pk])
            )

        self.assertEqual(response.status_code, 302)
        locking_queries = [
            sql
            for sql in executed_sql
            if "FOR UPDATE" in sql.upper()
            and "reporting_shiftreport" in sql.lower()
        ]
        self.assertEqual(len(locking_queries), 1)
        self.assertNotIn(" JOIN ", locking_queries[0].upper())

        report.refresh_from_db()
        self.assertEqual(report.status, ShiftReport.ReportStatus.APPROVED)
        self.assertEqual(report.approved_by, self.user)
        self.assertIsNotNone(report.approved_at)

    def test_excel_uses_snapshot_and_accepts_missing_optional_values(self):
        self._finish_shift()
        report = ReportService.generate_shift_report(
            shift_plan=self.shift,
            user=self.user,
        )
        snapshot = report.snapshot_data
        snapshot["maintenance_requests"] = [{}]
        ShiftReport.objects.filter(pk=report.pk).update(
            summary="",
            recommendations="",
            snapshot_data=snapshot,
        )
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                response = self.client.get(
                    reverse("reporting:export-excel", args=[report.pk])
                )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        self.assertIn("attachment", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        door_sheet = workbook["حالة الأبواب"]
        door_codes = {
            str(door_sheet.cell(row=row, column=1).value)
            for row in range(6, door_sheet.max_row + 1)
        }
        self.assertEqual(report.total_doors, 42)
        self.assertIn("6A", door_codes)
        self.assertIn("6B", door_codes)
        workbook.close()

    def test_failed_excel_export_marks_export_log_failed(self):
        self._finish_shift()
        report = ReportService.generate_shift_report(
            shift_plan=self.shift,
            user=self.user,
        )
        with tempfile.TemporaryDirectory() as media_root:
            with override_settings(MEDIA_ROOT=media_root):
                with patch(
                    "apps.reporting.views.Workbook.save",
                    side_effect=RuntimeError("excel renderer failed"),
                ):
                    with self.assertRaises(RuntimeError):
                        self.client.get(
                            reverse(
                                "reporting:export-excel",
                                args=[report.pk],
                            )
                        )
        export_log = ExportLog.objects.get(
            report_key="operational_shift_report",
            export_format=ExportLog.ExportFormat.EXCEL,
        )
        self.assertEqual(export_log.status, ExportLog.ExportStatus.FAILED)

    def test_storage_failure_returns_service_unavailable_without_partial_success(self):
        self._finish_shift()
        report = ReportService.generate_shift_report(
            shift_plan=self.shift,
            user=self.user,
        )
        with patch(
            "django.core.files.storage.FileSystemStorage._save",
            side_effect=OSError("storage unavailable"),
        ):
            response = self.client.get(
                reverse("reporting:export-excel", args=[report.pk])
            )
        self.assertEqual(response.status_code, 503)
        export_log = ExportLog.objects.get(
            report_key="operational_shift_report",
            export_format=ExportLog.ExportFormat.EXCEL,
        )
        self.assertEqual(export_log.status, ExportLog.ExportStatus.FAILED)
        self.assertFalse(export_log.file)
        self.assertEqual(export_log.download_count, 0)

    def test_pdf_storage_failure_uses_same_archival_contract(self):
        self._finish_shift()
        report = ReportService.generate_shift_report(
            shift_plan=self.shift,
            user=self.user,
        )
        with patch(
            "django.core.files.storage.FileSystemStorage._save",
            side_effect=OSError("storage unavailable"),
        ):
            response = self.client.get(
                reverse("reporting:export-pdf", args=[report.pk])
            )
        self.assertEqual(response.status_code, 503)
        export_log = ExportLog.objects.get(
            report_key="operational_shift_report",
            export_format=ExportLog.ExportFormat.PDF,
        )
        self.assertEqual(export_log.status, ExportLog.ExportStatus.FAILED)
        self.assertFalse(export_log.file)
        self.assertEqual(export_log.download_count, 0)

    def test_unauthenticated_and_unauthorized_access_is_denied(self):
        self._finish_shift()
        report = ReportService.generate_shift_report(
            shift_plan=self.shift,
            user=self.user,
        )
        urls = (
            reverse("reporting:detail", args=[report.pk]),
            reverse("reporting:export-pdf", args=[report.pk]),
            reverse("reporting:export-excel", args=[report.pk]),
        )
        self.client.logout()
        for url in urls:
            self.assertEqual(self.client.get(url).status_code, 302)

        unauthorized = get_user_model().objects.create_user(
            username="report-unauthorized"
        )
        self.client.force_login(unauthorized)
        for url in urls:
            self.assertEqual(self.client.get(url).status_code, 403)
        self.assertEqual(
            self.client.post(
                reverse("reporting:approve", args=[report.pk])
            ).status_code,
            403,
        )

    def test_section_scoped_user_cannot_view_or_export_all_scope_report(self):
        self._finish_shift()
        report = ReportService.generate_shift_report(
            shift_plan=self.shift,
            user=self.user,
        )
        group = Group.objects.create(name="male-report-scope")
        group.permissions.add(
            *Permission.objects.filter(
                content_type__app_label="roles",
                codename__in=("view_reports", "export_report"),
            )
        )
        role = Role.objects.create(
            code="male-report-scope",
            name="Male report scope",
            group=group,
            operational_section=Role.OperationalSection.MALE,
        )
        user = get_user_model().objects.create_user(username="male-reporter")
        UserRole.objects.create(user=user, role=role)
        self.client.force_login(user)
        self.assertEqual(
            self.client.get(reverse("reporting:detail", args=[report.pk])).status_code,
            404,
        )
        self.assertEqual(
            self.client.get(
                reverse("reporting:export-pdf", args=[report.pk])
            ).status_code,
            404,
        )

    def test_generation_rolls_back_when_summary_generation_fails(self):
        self._finish_shift()
        from unittest.mock import patch

        with patch(
            "apps.reporting.services.build_executive_summary",
            side_effect=RuntimeError("summary failed"),
        ):
            with self.assertRaises(RuntimeError):
                ReportService.generate_shift_report(
                    shift_plan=self.shift,
                    user=self.user,
                )
        self.assertFalse(ShiftReport.objects.exists())

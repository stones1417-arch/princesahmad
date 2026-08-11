from __future__ import annotations

import os
import tempfile
import hashlib
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import (
    get_object_or_404,
    redirect,
    render,
)
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.http import require_POST

from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from apps.distribution.models import DoorAssignment
from apps.ops.models import (
    DoorShift,
    MaintenanceRequest,
)
from apps.roles.decorators import permission_required
from apps.roles.services.permission_registry import (
    PlatformPermissions,
)
from apps.roles.services.section_access import get_allowed_sections, has_institutional_scope
from apps.scheduling.models import ShiftPlan

from .ai_summary import (
    build_executive_summary,
    build_recommendations,
)
from .forms import ShiftReportForm
from .models import ShiftReport
from .services import ReportService


# ==================================================
# أسماء الأشهر العربية
# ==================================================

ARABIC_MONTHS = {
    1: "يناير",
    2: "فبراير",
    3: "مارس",
    4: "أبريل",
    5: "مايو",
    6: "يونيو",
    7: "يوليو",
    8: "أغسطس",
    9: "سبتمبر",
    10: "أكتوبر",
    11: "نوفمبر",
    12: "ديسمبر",
}


def _electronic_approval_data(report: ShiftReport) -> dict:
    """Build a reproducible verification fingerprint for an approved report."""
    if not (
        report.status == ShiftReport.ReportStatus.APPROVED
        and report.approved_by_id
        and report.approved_at
    ):
        return {"is_signed": False}

    approved_at = timezone.localtime(report.approved_at)
    payload = "|".join(
        [
            str(report.pk),
            report.report_number or "",
            str(report.approved_by_id),
            approved_at.isoformat(),
            str(report.total_doors),
            str(report.total_employees),
            report.status,
            settings.SECRET_KEY,
        ]
    )
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest().upper()
    verification_code = "-".join(digest[index:index + 4] for index in range(0, 20, 4))
    approver_name = report.approved_by.get_full_name() or report.approved_by.get_username()
    job_title = "مسؤول اعتماد التقارير"
    employee = getattr(report.approved_by, "employee", None)
    if employee:
        job_title = employee.get_job_title_display()
    return {
        "is_signed": True,
        "approver_name": approver_name,
        "job_title": job_title,
        "approved_at": approved_at,
        "verification_code": verification_code,
        "initials": "".join(part[0] for part in approver_name.split()[:2]) or "✓",
    }


# ==================================================
# تحديث مؤشرات التقرير
# ==================================================

def _refresh_report_metrics(
    report: ShiftReport,
) -> ShiftReport:
    """
    تحديث مؤشرات التقرير من بيانات الوردية المرتبطة به.
    """
    if not report.shift_plan:
        return report

    door_shifts = DoorShift.objects.filter(
        shift_plan=report.shift_plan,
    )

    assignments = DoorAssignment.objects.filter(
        shift_plan=report.shift_plan,
        is_active=True,
    )

    maintenance_requests = (
        MaintenanceRequest.objects.filter(
            door_shift__shift_plan=report.shift_plan,
        )
    )

    report.total_doors = (
        door_shifts.count()
    )

    report.open_doors = (
        door_shifts.filter(
            state=DoorShift.DoorState.OPEN,
        ).count()
    )

    report.closed_doors = (
        door_shifts.filter(
            state=DoorShift.DoorState.CLOSED,
        ).count()
    )

    report.maintenance_doors = (
        door_shifts.filter(
            state=DoorShift.DoorState.MAINTENANCE,
        ).count()
    )

    report.total_employees = (
        assignments.count()
    )

    report.total_maintenance_requests = (
        maintenance_requests.count()
    )

    completed_statuses = []

    for status_name in (
        "DONE",
        "FIXED",
        "CLOSED",
    ):
        if hasattr(
            MaintenanceRequest.Status,
            status_name,
        ):
            completed_statuses.append(
                getattr(
                    MaintenanceRequest.Status,
                    status_name,
                )
            )

    if completed_statuses:
        report.completed_maintenance_requests = (
            maintenance_requests.filter(
                status__in=completed_statuses,
            ).count()
        )
    else:
        report.completed_maintenance_requests = 0

    return report


# ==================================================
# التحقق من صلاحية إنشاء التقارير للواجهة
# ==================================================

def _can_create_report(
    user,
) -> bool:
    """
    تستخدم لإظهار أو إخفاء أزرار إنشاء التقرير داخل القالب.

    الحماية الفعلية للدوال تتم بواسطة permission_required.
    """
    if not user or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    return user.has_perm(
        PlatformPermissions.CREATE_REPORT
    )


# ==================================================
# التحقق من صلاحية اعتماد التقارير للواجهة
# ==================================================

def _can_approve_report(
    user,
) -> bool:
    """
    تستخدم لإظهار أو إخفاء زر اعتماد التقرير داخل القالب.

    الحماية الفعلية للدالة تتم بواسطة permission_required.
    """
    if not user or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    return user.has_perm(
        PlatformPermissions.APPROVE_REPORT
    )


def _scoped_reports(user):
    queryset = ShiftReport.objects
    if user.is_superuser:
        return queryset
    if not has_institutional_scope(user):
        return queryset.none()
    allowed_sections = get_allowed_sections(user)
    if allowed_sections == {"male", "female"}:
        return queryset
    return queryset.filter(operational_section__in=allowed_sections)


# ==================================================
# تحريك الشهر السابق أو التالي
# ==================================================

def _shift_month(
    year: int,
    month: int,
    offset: int,
) -> tuple[int, int]:
    """
    إرجاع السنة والشهر بعد تطبيق الإزاحة المطلوبة.

    offset=-1 للشهر السابق.
    offset=1 للشهر التالي.
    """
    month_index = (
        year * 12
        + month
        - 1
        + offset
    )

    new_year = (
        month_index // 12
    )

    new_month = (
        month_index % 12
        + 1
    )

    return (
        new_year,
        new_month,
    )
# ==================================================
# قائمة التقارير
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.VIEW_REPORTS,
    message=(
        "ليس لديك صلاحية "
        "عرض التقارير التشغيلية."
    ),
)
def report_list_view(
    request,
):
    """
    عرض قائمة التقارير التشغيلية
    مع إحصاءات حالات التقارير والوردية النشطة.
    """
    all_reports = (
        _scoped_reports(request.user)
        .select_related(
            "shift_plan",
            "shift_plan__shift_type",
            "created_by",
            "approved_by",
        )
        .order_by(
            "-created_at"
        )
    )

    reports = all_reports

    report_type = request.GET.get("type", "").strip()
    status = request.GET.get("status", "").strip()
    query = request.GET.get("q", "").strip()
    if report_type in {value for value, _ in ShiftReport.ReportType.choices}:
        reports = reports.filter(report_type=report_type)
    if status in {value for value, _ in ShiftReport.ReportStatus.choices}:
        reports = reports.filter(status=status)
    if query:
        reports = reports.filter(
            Q(report_number__icontains=query)
            | Q(shift_plan__shift_type__name__icontains=query)
            | Q(created_by__username__icontains=query)
        )

    active_shift = (
        ShiftPlan.objects
        .select_related(
            "shift_type"
        )
        .filter(
            is_active=True
        )
        .first()
    )

    approved_count = all_reports.filter(
        status=(
            ShiftReport
            .ReportStatus
            .APPROVED
        )
    ).count()

    final_count = all_reports.filter(
        status=(
            ShiftReport
            .ReportStatus
            .FINAL
        )
    ).count()

    draft_count = all_reports.filter(
        status=(
            ShiftReport
            .ReportStatus
            .DRAFT
        )
    ).count()

    total_reports_count = all_reports.count()
    approval_rate = round(
        approved_count / total_reports_count * 100,
        1,
    ) if total_reports_count else 0

    context = {
        "reports": reports,
        "approved_count": approved_count,
        "final_count": final_count,
        "draft_count": draft_count,
        "total_reports_count": total_reports_count,
        "filtered_reports_count": reports.count(),
        "approval_rate": approval_rate,
        "active_shift": active_shift,
        "selected_type": report_type,
        "selected_status": status,
        "search_query": query,
        "can_create": _can_create_report(
            request.user
        ),
        "can_approve": _can_approve_report(
            request.user
        ),
    }

    return render(
        request,
        "reporting/report_list.html",
        context,
    )


# ==================================================
# لوحة المؤشرات التنفيذية الشهرية
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.VIEW_REPORTS,
    message=(
        "ليس لديك صلاحية "
        "عرض لوحة المؤشرات التنفيذية."
    ),
)
def executive_monthly_dashboard_view(
    request,
):
    """
    عرض مؤشرات الورديات والتقارير
    والأبواب والصيانة حسب الشهر المحدد.
    """
    today = timezone.localdate()

    # ------------------------------
    # السنة المحددة
    # ------------------------------

    try:
        selected_year = int(
            request.GET.get(
                "year",
                today.year,
            )
        )

    except (
        TypeError,
        ValueError,
    ):
        selected_year = today.year

    # ------------------------------
    # الشهر المحدد
    # ------------------------------

    try:
        selected_month = int(
            request.GET.get(
                "month",
                today.month,
            )
        )

    except (
        TypeError,
        ValueError,
    ):
        selected_month = today.month

    if selected_month not in range(
        1,
        13,
    ):
        selected_month = today.month

    if (
        selected_year < 2020
        or selected_year > 2100
    ):
        selected_year = today.year

    # ------------------------------
    # ورديات الشهر
    # ------------------------------

    month_shifts = (
        ShiftPlan.objects
        .select_related(
            "shift_type"
        )
        .filter(
            date__year=selected_year,
            date__month=selected_month,
        )
    )

    # ------------------------------
    # تقارير الشهر
    # ------------------------------

    month_reports = (
        ShiftReport.objects
        .select_related(
            "shift_plan",
            "shift_plan__shift_type",
            "created_by",
            "approved_by",
        )
        .filter(
            shift_plan__date__year=(
                selected_year
            ),
            shift_plan__date__month=(
                selected_month
            ),
        )
    )

    total_shifts = (
        month_shifts.count()
    )

    active_shifts = (
        month_shifts.filter(
            is_active=True
        ).count()
    )

    # ------------------------------
    # حالات الورديات
    # ------------------------------
    #
    # نموذج ShiftPlan لا يحتوي على حقل status.
    # لذلك يتم الاعتماد على الحقلين الموجودين فعليًا:
    # is_active و is_finished.
    #
    # الأولوية:
    # 1) الوردية المنتهية: is_finished=True
    # 2) الوردية النشطة: is_active=True و is_finished=False
    # 3) الوردية المجدولة/المعلقة: غير نشطة وغير منتهية
    # ------------------------------

    finished_shifts = (
        month_shifts.filter(
            is_finished=True,
        ).count()
    )

    pending_shifts = (
        month_shifts.filter(
            is_active=False,
            is_finished=False,
        ).count()
    )

    shift_status_labels = {
        "active": "وردية نشطة",
        "finished": "وردية منتهية",
        "scheduled": "وردية مجدولة",
    }

    shift_status_counts = {
        "active": active_shifts,
        "finished": finished_shifts,
        "scheduled": pending_shifts,
    }

    # ------------------------------
    # حالات التقارير
    # ------------------------------

    approved_reports = (
        month_reports.filter(
            status=(
                ShiftReport
                .ReportStatus
                .APPROVED
            )
        ).count()
    )

    final_reports = (
        month_reports.filter(
            status=(
                ShiftReport
                .ReportStatus
                .FINAL
            )
        ).count()
    )

    draft_reports = (
        month_reports.filter(
            status=(
                ShiftReport
                .ReportStatus
                .DRAFT
            )
        ).count()
    )

    total_reports = (
        month_reports.count()
    )

    approval_rate = (
        round(
            (
                approved_reports
                / total_reports
            ) * 100,
            2,
        )
        if total_reports
        else 0
    )

    # ------------------------------
    # مؤشرات الأبواب
    # ------------------------------

    door_totals = (
        month_reports.aggregate(
            total_doors_sum=Sum(
                "total_doors"
            ),
            open_doors_sum=Sum(
                "open_doors"
            ),
            closed_doors_sum=Sum(
                "closed_doors"
            ),
            maintenance_doors_sum=Sum(
                "maintenance_doors"
            ),
        )
    )

    total_doors_sum = (
        door_totals.get(
            "total_doors_sum"
        )
        or 0
    )

    open_doors_sum = (
        door_totals.get(
            "open_doors_sum"
        )
        or 0
    )

    closed_doors_sum = (
        door_totals.get(
            "closed_doors_sum"
        )
        or 0
    )

    maintenance_doors_sum = (
        door_totals.get(
            "maintenance_doors_sum"
        )
        or 0
    )

    avg_open_rate = (
        round(
            (
                open_doors_sum
                / total_doors_sum
            ) * 100,
            2,
        )
        if total_doors_sum
        else 0
    )

    # ------------------------------
    # مؤشرات الصيانة
    # ------------------------------

    maintenance_totals = (
        month_reports.aggregate(
            total_requests_sum=Sum(
                "total_maintenance_requests"
            ),
            completed_requests_sum=Sum(
                "completed_maintenance_requests"
            ),
        )
    )

    total_maintenance_requests = (
        maintenance_totals.get(
            "total_requests_sum"
        )
        or 0
    )

    completed_maintenance_requests = (
        maintenance_totals.get(
            "completed_requests_sum"
        )
        or 0
    )

    avg_maintenance_rate = (
        round(
            (
                completed_maintenance_requests
                / total_maintenance_requests
            ) * 100,
            2,
        )
        if total_maintenance_requests
        else 0
    )

    # ------------------------------
    # الشهر السابق والتالي
    # ------------------------------

    (
        previous_year,
        previous_month,
    ) = _shift_month(
        selected_year,
        selected_month,
        -1,
    )

    (
        next_year,
        next_month,
    ) = _shift_month(
        selected_year,
        selected_month,
        1,
    )

    # ------------------------------
    # قائمة الأشهر
    # ------------------------------

    months = [
        {
            "number": month_number,
            "name": ARABIC_MONTHS[
                month_number
            ],
            "active": (
                month_number
                == selected_month
            ),
        }
        for month_number
        in range(1, 13)
    ]

    context = {
        "selected_year": (
            selected_year
        ),
        "selected_month": (
            selected_month
        ),
        "selected_month_name": (
            ARABIC_MONTHS[
                selected_month
            ]
        ),
        "month_label": (
            f"{ARABIC_MONTHS[selected_month]} "
            f"{selected_year}"
        ),

        "previous_year": (
            previous_year
        ),
        "previous_month": (
            previous_month
        ),
        "next_year": (
            next_year
        ),
        "next_month": (
            next_month
        ),
        "months": months,

        "total_shifts": (
            total_shifts
        ),
        "active_shifts": (
            active_shifts
        ),
        "finished_shifts": (
            finished_shifts
        ),
        "pending_shifts": (
            pending_shifts
        ),
        "shift_status_labels": (
            shift_status_labels
        ),
        "shift_status_counts": (
            shift_status_counts
        ),

        "total_reports": (
            total_reports
        ),
        "approved_reports": (
            approved_reports
        ),
        "final_reports": (
            final_reports
        ),
        "draft_reports": (
            draft_reports
        ),
        "approval_rate": (
            approval_rate
        ),

        "total_doors_sum": (
            total_doors_sum
        ),
        "open_doors_sum": (
            open_doors_sum
        ),
        "closed_doors_sum": (
            closed_doors_sum
        ),
        "maintenance_doors_sum": (
            maintenance_doors_sum
        ),
        "avg_open_rate": (
            avg_open_rate
        ),

        "total_maintenance_requests": (
            total_maintenance_requests
        ),
        "completed_maintenance_requests": (
            completed_maintenance_requests
        ),
        "avg_maintenance_rate": (
            avg_maintenance_rate
        ),

        "is_current_month": (
            selected_year
            == today.year
            and selected_month
            == today.month
        ),

        "can_create": (
            _can_create_report(
                request.user
            )
        ),
        "can_approve": (
            _can_approve_report(
                request.user
            )
        ),
    }

    return render(
        request,
        (
            "reporting/"
            "executive_monthly_dashboard.html"
        ),
        context,
    )
# ==================================================
# التقرير التشغيلي المباشر
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.VIEW_REPORTS,
    message=(
        "ليس لديك صلاحية "
        "عرض التقرير التشغيلي."
    ),
)
def operational_report_view(
    request,
):
    """
    عرض التقرير التشغيلي للوردية النشطة،
    شاملاً الأبواب والتوزيع والصيانة.
    """
    active_shift = (
        ShiftPlan.objects
        .select_related(
            "shift_type"
        )
        .filter(
            is_active=True
        )
        .first()
    )

    door_shifts = (
        DoorShift.objects.none()
    )

    assignments = (
        DoorAssignment.objects.none()
    )

    maintenance_queryset = (
        MaintenanceRequest.objects.none()
    )

    if active_shift:
        door_shifts = (
            DoorShift.objects
            .filter(
                shift_plan=active_shift,
                is_active=True,
            )
            .select_related(
                "shift_plan",
                "shift_plan__shift_type",
            )
            .order_by(
                "door_number"
            )
        )

        assignments = (
            DoorAssignment.objects
            .select_related(
                "employee",
                "door",
                "shift_plan",
                "shift_plan__shift_type",
            )
            .filter(
                shift_plan=active_shift,
                is_active=True,
            )
            .order_by(
                "door__door_number",
                "role",
                "employee__employee_number",
            )
        )

        maintenance_queryset = (
            MaintenanceRequest.objects
            .select_related(
                "door_shift",
                "door_shift__shift_plan",
                "door_shift__shift_plan__shift_type",
                "created_by",
                "technician",
            )
            .filter(
                door_shift__shift_plan=active_shift,
            )
            .order_by(
                "-created_at"
            )
        )

    # ------------------------------
    # حالات الصيانة المكتملة
    # ------------------------------

    completed_statuses = []

    for status_name in (
        "DONE",
        "FIXED",
        "CLOSED",
    ):
        if hasattr(
            MaintenanceRequest.Status,
            status_name,
        ):
            completed_statuses.append(
                getattr(
                    MaintenanceRequest.Status,
                    status_name,
                )
            )

    # ------------------------------
    # مؤشرات الصيانة
    # ------------------------------

    maintenance_total = (
        maintenance_queryset.count()
    )

    if completed_statuses:
        maintenance_completed = (
            maintenance_queryset.filter(
                status__in=completed_statuses
            ).count()
        )

    else:
        maintenance_completed = 0

    maintenance_pending = max(
        maintenance_total
        - maintenance_completed,
        0,
    )

    maintenance_rate = (
        round(
            (
                maintenance_completed
                / maintenance_total
            ) * 100,
            2,
        )
        if maintenance_total
        else 0
    )

    # ------------------------------
    # مؤشرات الأبواب
    # ------------------------------

    total_doors = (
        door_shifts.count()
    )

    open_doors = (
        door_shifts.filter(
            state=(
                DoorShift
                .DoorState
                .OPEN
            )
        ).count()
    )

    closed_doors = (
        door_shifts.filter(
            state=(
                DoorShift
                .DoorState
                .CLOSED
            )
        ).count()
    )

    maintenance_doors = (
        door_shifts.filter(
            state=(
                DoorShift
                .DoorState
                .MAINTENANCE
            )
        ).count()
    )

    secured_doors = door_shifts.filter(
        state=DoorShift.DoorState.SECURED,
    ).count()

    readiness_rate = (
        round(
            (
                open_doors
                / total_doors
            ) * 100,
            2,
        )
        if total_doors
        else 0
    )

    # ------------------------------
    # مؤشرات توزيع الموظفين
    # ------------------------------

    supervisors_count = (
        assignments.filter(
            role=(
                DoorAssignment
                .Role
                .SUPERVISOR
            )
        ).count()
    )

    monitors_count = (
        assignments.filter(
            role=(
                DoorAssignment
                .Role
                .MONITOR
            )
        ).count()
    )

    technicians_count = (
        assignments.filter(
            role=(
                DoorAssignment
                .Role
                .TECHNICIAN
            )
        ).count()
    )

    support_count = assignments.filter(
        role=DoorAssignment.Role.SUPPORT,
    ).count()

    total_assignments = (
        assignments.count()
    )

    employees_count = assignments.values(
        "employee_id"
    ).distinct().count()

    # يجب تنفيذ عمليات filter وcount
    # قبل قص QuerySet.
    maintenance_requests = (
        maintenance_queryset[:20]
    )

    context = {
        "active_shift": (
            active_shift
        ),
        "door_shifts": (
            door_shifts
        ),
        "assignments": (
            assignments
        ),
        "maintenance_requests": (
            maintenance_requests
        ),

        "total_doors": (
            total_doors
        ),
        "open_doors": (
            open_doors
        ),
        "closed_doors": (
            closed_doors
        ),
        "maintenance_doors": (
            maintenance_doors
        ),
        "secured_doors": secured_doors,
        "employees_count": employees_count,
        "readiness_rate": (
            readiness_rate
        ),

        "total_assignments": (
            total_assignments
        ),
        "supervisors_count": (
            supervisors_count
        ),
        "monitors_count": (
            monitors_count
        ),
        "technicians_count": (
            technicians_count
        ),
        "support_count": support_count,
        "assignment_role_choices": DoorAssignment.Role.choices,

        "maintenance_total": (
            maintenance_total
        ),
        "maintenance_completed": (
            maintenance_completed
        ),
        "maintenance_pending": (
            maintenance_pending
        ),
        "maintenance_rate": (
            maintenance_rate
        ),

        "can_create": (
            _can_create_report(
                request.user
            )
        ),
        "can_approve": (
            _can_approve_report(
                request.user
            )
        ),
    }

    return render(
        request,
        "reporting/operational_report.html",
        context,
    )
# ==================================================
# إنشاء تقرير جديد
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.CREATE_REPORT,
    message=(
        "ليس لديك صلاحية "
        "إنشاء التقارير."
    ),
)
def report_create_view(
    request,
    default_report_type=None,
):
    """
    إنشاء تقرير جديد يدوي أو تشغيلي.
    """
    initial_data = {}

    allowed_report_types = {
        ShiftReport.ReportType.OPERATIONAL,
        ShiftReport.ReportType.MANUAL,
    }

    if default_report_type in allowed_report_types:
        initial_data[
            "report_type"
        ] = default_report_type

    if request.method == "POST":
        form = ShiftReportForm(
            request.POST
        )

        if form.is_valid():
            report = form.save(
                commit=False
            )

            report.created_by = (
                request.user
            )

            report.status = (
                ShiftReport
                .ReportStatus
                .DRAFT
            )

            if (
                report.report_type
                == ShiftReport
                .ReportType
                .OPERATIONAL
            ):
                _refresh_report_metrics(
                    report
                )

                report.summary = (
                    build_executive_summary(
                        report
                    )
                )

                report.recommendations = (
                    "\n".join(
                        build_recommendations(
                            report
                        )
                    )
                )

            elif (
                report.report_type
                == ShiftReport
                .ReportType
                .MANUAL
            ):
                report.total_doors = 0
                report.open_doors = 0
                report.closed_doors = 0
                report.maintenance_doors = 0
                report.total_employees = 0
                report.total_maintenance_requests = 0
                report.completed_maintenance_requests = 0

            report.save()

            messages.success(
                request,
                "تم إنشاء التقرير بنجاح",
            )

            return redirect(
                "reporting:detail",
                pk=report.pk,
            )

        messages.error(
            request,
            "تحقق من الحقول المدخلة",
        )

    else:
        form = ShiftReportForm(
            initial=initial_data
        )

    return render(
        request,
        "reporting/report_form.html",
        {
            "form": form,
            "default_report_type": (
                default_report_type
            ),
        },
    )


# ==================================================
# توليد تقرير لوردية محددة
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.CREATE_REPORT,
    message=(
        "ليس لديك صلاحية "
        "إنشاء تقارير الورديات."
    ),
)
@require_POST
def generate_report_view(
    request,
    pk,
):
    """
    إنشاء تقرير تلقائي لوردية محددة.
    """
    shift = get_object_or_404(
        ShiftPlan.objects
        .select_related(
            "shift_type"
        ),
        pk=pk,
    )

    try:
        report = (
            ReportService
            .generate_shift_report(
                request=request,
                shift_plan=shift,
                user=request.user,
            )
        )

        messages.success(
            request,
            "تم إنشاء تقرير الوردية تلقائيًا",
        )

        return redirect(
            "reporting:detail",
            pk=report.pk,
        )

    except ValidationError as error:
        messages.error(
            request,
            str(error),
        )

        return redirect(
            "reporting:list"
        )


# ==================================================
# توليد تقرير للوردية النشطة
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.CREATE_REPORT,
    message=(
        "ليس لديك صلاحية "
        "إنشاء تقرير الوردية النشطة."
    ),
)
@require_POST
def generate_active_shift_report_view(
    request,
):
    """
    إنشاء تقرير تلقائي للوردية النشطة الحالية.
    """
    active_shift = (
        ShiftPlan.objects
        .select_related(
            "shift_type"
        )
        .filter(
            is_active=True
        )
        .first()
    )

    if not active_shift:
        messages.error(
            request,
            (
                "لا توجد وردية نشطة "
                "لإنشاء تقرير"
            ),
        )

        return redirect(
            "reporting:list"
        )

    try:
        report = (
            ReportService
            .generate_shift_report(
                request=request,
                shift_plan=active_shift,
                user=request.user,
            )
        )

        messages.success(
            request,
            (
                "تم إنشاء تقرير "
                "الوردية النشطة تلقائيًا"
            ),
        )

        return redirect(
            "reporting:detail",
            pk=report.pk,
        )

    except ValidationError as error:
        messages.error(
            request,
            str(error),
        )

        return redirect(
            "reporting:list"
        )
    # ==================================================
# تفاصيل التقرير
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.VIEW_REPORTS,
    message=(
        "ليس لديك صلاحية "
        "عرض تفاصيل التقرير."
    ),
)
def report_detail_view(
    request,
    pk,
):
    """
    عرض لوحة تفاصيل التقرير،
    شاملاً الأبواب والتوزيع والصيانة والمؤشرات.
    """
    report = get_object_or_404(
        _scoped_reports(request.user).select_related(
            "shift_plan",
            "shift_plan__shift_type",
            "created_by",
            "approved_by",
            "approved_by__employee",
        ),
        pk=pk,
    )

    door_shifts = (
        DoorShift.objects.none()
    )

    assignments = (
        DoorAssignment.objects.none()
    )

    maintenance_requests = (
        MaintenanceRequest.objects.none()
    )

    supervisors_count = 0
    monitors_count = 0
    technicians_count = 0
    readiness_rate = 0
    maintenance_rate = 0

    if report.shift_plan:
        door_shifts = (
            DoorShift.objects
            .filter(
                shift_plan=report.shift_plan,
                is_active=True,
            )
            .order_by(
                "door_number"
            )
        )

        assignments = (
            DoorAssignment.objects
            .select_related(
                "door",
                "employee",
            )
            .filter(
                shift_plan=report.shift_plan,
                is_active=True,
            )
            .order_by(
                "door__door_number",
                "role",
            )
        )

        maintenance_requests = (
            MaintenanceRequest.objects
            .select_related(
                "door_shift",
                "created_by",
                "technician",
            )
            .filter(
                door_shift__shift_plan=(
                    report.shift_plan
                )
            )
            .order_by(
                "-created_at"
            )
        )

        supervisors_count = (
            assignments.filter(
                role=(
                    DoorAssignment
                    .Role
                    .SUPERVISOR
                )
            ).count()
        )

        monitors_count = (
            assignments.filter(
                role=(
                    DoorAssignment
                    .Role
                    .MONITOR
                )
            ).count()
        )

        technicians_count = (
            assignments.filter(
                role=(
                    DoorAssignment
                    .Role
                    .TECHNICIAN
                )
            ).count()
        )

    if report.total_doors:
        readiness_rate = round(
            (
                report.open_doors
                / report.total_doors
            ) * 100,
            2,
        )

    if report.total_maintenance_requests:
        maintenance_rate = round(
            (
                report.completed_maintenance_requests
                / report.total_maintenance_requests
            ) * 100,
            2,
        )

    context = {
        "report": (
            report
        ),
        "door_shifts": (
            door_shifts
        ),
        "assignments": (
            assignments
        ),
        "maintenance_requests": (
            maintenance_requests
        ),

        "total_assignments": (
            assignments.count()
        ),
        "supervisors_count": (
            supervisors_count
        ),
        "monitors_count": (
            monitors_count
        ),
        "technicians_count": (
            technicians_count
        ),

        "readiness_rate": (
            readiness_rate
        ),
        "maintenance_rate": (
            maintenance_rate
        ),

        "can_approve": (
            _can_approve_report(
                request.user
            )
        ),
        "can_create": (
            _can_create_report(
                request.user
            )
        ),
    }

    return render(
        request,
        "reporting/report_dashboard.html",
        context,
    )


# ==================================================
# اعتماد التقرير
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.APPROVE_REPORT,
    message=(
        "ليس لديك صلاحية "
        "اعتماد التقارير."
    ),
)
@require_POST
def approve_report_view(
    request,
    pk,
):
    """
    اعتماد التقرير رسميًا.
    """
    report = get_object_or_404(
        _scoped_reports(request.user),
        pk=pk,
    )

    try:
        ReportService.approve_report(
            request=request,
            report=report,
            user=request.user,
        )

        messages.success(
            request,
            "تم اعتماد التقرير رسميًا",
        )

    except ValidationError as error:
        messages.error(
            request,
            str(error),
        )

    return redirect(
        "reporting:detail",
        pk=pk,
    )


# ==================================================
# تحديث مؤشرات التقرير والملخص
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.CREATE_REPORT,
    message=(
        "ليس لديك صلاحية "
        "تحديث ملخص التقرير."
    ),
)
@require_POST
def regenerate_report_summary_view(
    request,
    pk,
):
    """
    تحديث مؤشرات التقرير،
    ثم إعادة إنشاء الملخص والتوصيات.
    """
    report = get_object_or_404(
        _scoped_reports(request.user),
        pk=pk,
    )

    _refresh_report_metrics(
        report
    )

    report.save(
        update_fields=[
            "total_doors",
            "open_doors",
            "closed_doors",
            "maintenance_doors",
            "total_employees",
            "total_maintenance_requests",
            "completed_maintenance_requests",
        ]
    )

    ReportService.regenerate_summary(
        request=request,
        report=report,
    )

    messages.success(
        request,
        (
            "تم تحديث مؤشرات التقرير "
            "والملخص والتوصيات."
        ),
    )

    return redirect(
        "reporting:detail",
        pk=report.pk,
    )
# ==================================================
# العثور على شعار الحرمين
# ==================================================

def _get_haramain_logo_path():
    """
    البحث عن شعار الحرمين داخل media أو static.
    """
    candidates = [
        os.path.join(
            settings.BASE_DIR,
            "media",
            "aharamain_logo.png",
        ),
        os.path.join(
            settings.BASE_DIR,
            "media",
            "aharamaian_logo.png",
        ),
        os.path.join(
            settings.BASE_DIR,
            "media",
            "alharamain_logo.png",
        ),
        os.path.join(
            settings.BASE_DIR,
            "media",
            "haramain_logo.png",
        ),
        os.path.join(
            settings.BASE_DIR,
            "media",
            "شعار الحرمين.png",
        ),
        os.path.join(
            settings.BASE_DIR,
            "static",
            "img",
            "aharamain_logo.png",
        ),
        os.path.join(
            settings.BASE_DIR,
            "static",
            "img",
            "aharamaian_logo.png",
        ),
        os.path.join(
            settings.BASE_DIR,
            "static",
            "img",
            "alharamain_logo.png",
        ),
        os.path.join(
            settings.BASE_DIR,
            "static",
            "img",
            "haramain_logo.png",
        ),
        os.path.join(
            settings.BASE_DIR,
            "static",
            "img",
            "شعار الحرمين.png",
        ),
    ]

    for path in candidates:
        if os.path.exists(path):
            return path

    return None


# ==================================================
# تصدير التقرير إلى PDF
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.EXPORT_REPORT,
    message=(
        "ليس لديك صلاحية "
        "تصدير التقرير إلى PDF."
    ),
)
def export_report_pdf_view(
    request,
    pk,
):
    """
    تصدير تقرير تشغيلي واحد إلى PDF.
    """
    report = get_object_or_404(
        _scoped_reports(request.user).select_related(
            "shift_plan",
            "shift_plan__shift_type",
            "created_by",
            "approved_by",
            "approved_by__employee",
        ),
        pk=pk,
    )

    door_shifts = (
        DoorShift.objects.none()
    )

    assignments = (
        DoorAssignment.objects.none()
    )

    maintenance_requests = (
        MaintenanceRequest.objects.none()
    )

    if report.shift_plan:
        door_shifts = (
            DoorShift.objects
            .filter(
                shift_plan=report.shift_plan,
                is_active=True,
            )
            .order_by(
                "door_number"
            )
        )

        assignments = (
            DoorAssignment.objects
            .select_related(
                "door",
                "employee",
            )
            .filter(
                shift_plan=report.shift_plan,
                is_active=True,
            )
            .order_by(
                "door__door_number",
                "role",
            )
        )

        maintenance_requests = (
            MaintenanceRequest.objects
            .select_related(
                "door_shift",
                "created_by",
            )
            .filter(
                door_shift__shift_plan=(
                    report.shift_plan
                )
            )
            .order_by(
                "-created_at"
            )[:8]
        )

    total_assignments = (
        assignments.count()
    )

    supervisors_count = (
        assignments.filter(
            role=(
                DoorAssignment
                .Role
                .SUPERVISOR
            )
        ).count()
    )

    monitors_count = (
        assignments.filter(
            role=(
                DoorAssignment
                .Role
                .MONITOR
            )
        ).count()
    )

    technicians_count = (
        assignments.filter(
            role=(
                DoorAssignment
                .Role
                .TECHNICIAN
            )
        ).count()
    )

    readiness_rate = 0

    if report.total_doors:
        readiness_rate = round(
            (
                report.open_doors
                / report.total_doors
            ) * 100,
            2,
        )

    total_maintenance = (
        report.total_maintenance_requests
        or 0
    )

    completed_maintenance = (
        report.completed_maintenance_requests
        or 0
    )

    maintenance_rate = 0

    if total_maintenance:
        maintenance_rate = round(
            (
                completed_maintenance
                / total_maintenance
            ) * 100,
            2,
        )

    html_string = render_to_string(
        "reporting/pdf/shift_report_pdf.html",
        {
            "report": report,
            "door_shifts": door_shifts,
            "assignments": assignments,
            "maintenance_requests": (
                maintenance_requests
            ),
            "total_assignments": (
                total_assignments
            ),
            "supervisors_count": (
                supervisors_count
            ),
            "monitors_count": (
                monitors_count
            ),
            "technicians_count": (
                technicians_count
            ),
            "readiness_rate": (
                readiness_rate
            ),
            "maintenance_rate": (
                maintenance_rate
            ),
            "logo_path": (
                _get_haramain_logo_path()
            ),
            "electronic_approval": _electronic_approval_data(report),
        },
    )

    html_path = None
    pdf_path = None
    pdf_data = b""

    try:
        with tempfile.NamedTemporaryFile(
            suffix=".html",
            delete=False,
            mode="w",
            encoding="utf-8",
        ) as html_file:
            html_file.write(
                html_string
            )

            html_path = (
                html_file.name
            )

        pdf_path = (
            html_path.replace(
                ".html",
                ".pdf",
            )
        )

        with sync_playwright() as playwright:
            browser = (
                playwright.chromium.launch(
                    headless=True
                )
            )

            try:
                page = browser.new_page()

                file_url = (
                    "file:///"
                    + html_path.replace(
                        os.sep,
                        "/",
                    )
                )

                page.goto(
                    file_url,
                    wait_until="networkidle",
                )

                page.pdf(
                    path=pdf_path,
                    format="A4",
                    print_background=True,
                    margin={
                        "top": "0mm",
                        "right": "0mm",
                        "bottom": "0mm",
                        "left": "0mm",
                    },
                )

            finally:
                browser.close()

        with open(
            pdf_path,
            "rb",
        ) as pdf_file:
            pdf_data = (
                pdf_file.read()
            )

    finally:
        if html_path:
            try:
                os.remove(
                    html_path
                )

            except OSError:
                pass

        if pdf_path:
            try:
                os.remove(
                    pdf_path
                )

            except OSError:
                pass

    response = HttpResponse(
        pdf_data,
        content_type="application/pdf",
    )

    response[
        "Content-Disposition"
    ] = (
        f'attachment; '
        f'filename="operational_report_'
        f'{report.report_number}.pdf"'
    )

    return response


@login_required
@permission_required(
    PlatformPermissions.EXPORT_REPORT,
    message="ليس لديك صلاحية تصدير التقارير.",
)
def export_report_excel_view(request, pk):
    """تصدير تقرير تشغيلي متعدد الأوراق بصيغة Excel مؤسسية."""
    report = get_object_or_404(
        _scoped_reports(request.user).select_related(
            "shift_plan", "shift_plan__shift_type", "created_by", "approved_by"
        ),
        pk=pk,
    )
    door_shifts = DoorShift.objects.none()
    assignments = DoorAssignment.objects.none()
    maintenance_requests = MaintenanceRequest.objects.none()
    if report.shift_plan:
        door_shifts = DoorShift.objects.filter(
            shift_plan=report.shift_plan,
        ).select_related("supervisor").order_by("door_number")
        assignments = DoorAssignment.objects.filter(
            shift_plan=report.shift_plan,
        ).select_related("door", "door__zone", "employee").order_by(
            "door__door_number", "role", "employee__full_name"
        )
        maintenance_requests = MaintenanceRequest.objects.filter(
            door_shift__shift_plan=report.shift_plan,
        ).select_related("door_shift", "technician").order_by("-created_at")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "الملخص التنفيذي"
    green, deep_green, gold = "0F6B50", "084C3B", "C9A548"
    pale_green, pale_gold, pale_gray = "EAF6F1", "FFF7DF", "F4F6F5"
    white, ink, muted = "FFFFFF", "17211E", "69756F"
    thin = Side(style="thin", color="DDE5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def display_user(user):
        if not user:
            return "—"
        return user.get_full_name() or user.get_username()

    def configure(ws, widths, landscape=True):
        ws.sheet_view.rightToLeft = True
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape" if landscape else "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.outlinePr.summaryBelow = True
        ws.oddFooter.center.text = "منصة أبواب | تقرير مؤسسي"
        ws.oddFooter.right.text = "صفحة &P من &N"
        ws.oddFooter.left.text = f"{report.report_number or report.pk}"
        for column, width in widths.items():
            ws.column_dimensions[column].width = width

    def title_band(ws, title, subtitle, end_column):
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_column)
        cell = ws.cell(1, 1, title)
        cell.font = Font(name="Arial", size=20, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=deep_green)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_column)
        cell = ws.cell(3, 1, subtitle)
        cell.font = Font(name="Arial", size=10, bold=True, color=muted)
        cell.fill = PatternFill("solid", fgColor="F8FAF9")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 24

    def table_sheet(name, title, headers, data, widths, tab_color):
        ws = workbook.create_sheet(name)
        configure(ws, widths)
        title_band(ws, title, f"رقم التقرير: {report.report_number or '—'}", len(headers))
        header_row = 5
        for column, label in enumerate(headers, 1):
            cell = ws.cell(header_row, column, label)
            cell.font = Font(name="Arial", bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=green)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row_index, values in enumerate(data, header_row + 1):
            for column, value in enumerate(values, 1):
                cell = ws.cell(row_index, column, value)
                cell.font = Font(name="Arial", size=10, color=ink)
                cell.fill = PatternFill("solid", fgColor=white if row_index % 2 else pale_gray)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            ws.row_dimensions[row_index].height = 24
        last_row = max(header_row + 1, header_row + len(data))
        if not data:
            ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=len(headers))
            ws.cell(header_row + 1, 1, "لا توجد بيانات مرتبطة بهذا التقرير")
            ws.cell(header_row + 1, 1).alignment = Alignment(horizontal="center")
        else:
            table = Table(displayName=f"ReportTable{len(workbook.worksheets)}", ref=f"A{header_row}:{chr(64 + len(headers))}{last_row}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
            ws.add_table(table)
        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = f"A{header_row}:{chr(64 + len(headers))}{last_row}"
        ws.sheet_properties.tabColor = tab_color
        ws.print_title_rows = f"1:{header_row}"
        ws.print_area = f"A1:{chr(64 + len(headers))}{last_row}"
        return ws

    configure(summary, {"A": 22, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22})
    title_band(summary, "منصة أبواب | التقرير التشغيلي المؤسسي", f"تم التصدير في {timezone.localtime().strftime('%Y-%m-%d %H:%M')}", 6)
    summary.sheet_properties.tabColor = green
    shift_name = "—"
    if report.shift_plan:
        shift_name = f"{report.shift_plan.shift_type.name} | {report.shift_plan.date:%Y-%m-%d}"
    electronic_approval = _electronic_approval_data(report)
    approval_state = "معتمد إلكترونيًا" if electronic_approval["is_signed"] else "بانتظار الاعتماد"
    approval_code = electronic_approval.get("verification_code", "—")
    metadata = [
        ("رقم التقرير", report.report_number or "—", "نوع التقرير", report.get_report_type_display(), "الحالة", report.get_status_display()),
        ("الوردية", shift_name, "منشئ التقرير", display_user(report.created_by), "المعتمد", display_user(report.approved_by)),
        (
            "حالة التوقيع",
            approval_state,
            "وقت الاعتماد",
            timezone.localtime(report.approved_at).strftime("%Y-%m-%d %H:%M") if report.approved_at else "—",
            "بصمة التحقق",
            approval_code,
        ),
    ]
    for row_index, row in enumerate(metadata, 5):
        for column, value in enumerate(row, 1):
            cell = summary.cell(row_index, column, value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if column % 2:
                cell.font = Font(name="Arial", bold=True, color=white)
                cell.fill = PatternFill("solid", fgColor=gold)
            else:
                cell.font = Font(name="Arial", bold=True, color=ink)
                cell.fill = PatternFill("solid", fgColor=white)
        summary.row_dimensions[row_index].height = 28

    metrics = [
        ("إجمالي الأبواب", report.total_doors or 0), ("الأبواب المفتوحة", report.open_doors or 0),
        ("الأبواب المغلقة", report.closed_doors or 0), ("تحت الصيانة", report.maintenance_doors or 0),
        ("إجمالي الموظفين", report.total_employees or 0), ("طلبات الصيانة", report.total_maintenance_requests or 0),
        ("الطلبات المنجزة", report.completed_maintenance_requests or 0), ("نسبة إنجاز الصيانة", (report.maintenance_completion_rate or 0) / 100),
    ]
    summary.merge_cells("A8:F8")
    summary["A8"] = "مؤشرات الأداء الرئيسية"
    summary["A8"].font = Font(name="Arial", size=13, bold=True, color=white)
    summary["A8"].fill = PatternFill("solid", fgColor=green)
    summary["A8"].alignment = Alignment(horizontal="center")
    for index, (label, value) in enumerate(metrics):
        row = 10 + (index // 4) * 3
        column = 1 + (index % 4)
        summary.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column)
        summary.cell(row, column, label)
        summary.cell(row + 1, column, value)
        for target in (summary.cell(row, column), summary.cell(row + 1, column)):
            target.border = border
            target.alignment = Alignment(horizontal="center", vertical="center")
        summary.cell(row, column).font = Font(name="Arial", bold=True, color=muted)
        summary.cell(row, column).fill = PatternFill("solid", fgColor=pale_green)
        summary.cell(row + 1, column).font = Font(name="Arial", size=17, bold=True, color=green)
        if "نسبة" in label:
            summary.cell(row + 1, column).number_format = "0.0%"
    # Data source and chart for door state overview.
    secured_doors = max(
        0,
        (report.total_doors or 0)
        - (report.open_doors or 0)
        - (report.closed_doors or 0)
        - (report.maintenance_doors or 0),
    )
    chart_data = [
        ("مفتوحة", report.open_doors or 0),
        ("مغلقة", report.closed_doors or 0),
        ("صيانة", report.maintenance_doors or 0),
        ("مؤمنة", secured_doors),
    ]
    for row_index, values in enumerate([("الحالة", "العدد"), *chart_data], 18):
        summary.cell(row_index, 1, values[0]); summary.cell(row_index, 2, values[1])
    chart = DoughnutChart()
    chart.title = "توزيع حالات الأبواب"
    chart.add_data(Reference(summary, min_col=2, min_row=18, max_row=22), titles_from_data=True)
    chart.set_categories(Reference(summary, min_col=1, min_row=19, max_row=22))
    chart.height, chart.width, chart.holeSize = 7, 11, 55
    summary.add_chart(chart, "D18")
    summary.freeze_panes = "A5"
    summary.print_area = "A1:F32"

    door_data = [(d.door_number, d.get_state_display(), "نعم" if d.is_active else "لا", display_user(d.supervisor), d.notes or "—", timezone.localtime(d.updated_at).strftime("%Y-%m-%d %H:%M")) for d in door_shifts]
    table_sheet("حالة الأبواب", "سجل حالة الأبواب", ["رقم الباب", "الحالة", "نشط", "المشرف", "ملاحظات", "آخر تحديث"], door_data, {"A": 14, "B": 18, "C": 12, "D": 25, "E": 38, "F": 22}, "2563EB")
    assignment_data = [(a.door.door_number, a.employee.employee_number, a.employee.full_name, a.get_role_display(), a.employee.phone_number or "—", "نشط" if a.is_active else "غير نشط", a.notes or "—") for a in assignments]
    table_sheet("توزيع الموظفين", "التوزيع التشغيلي للموظفين", ["الباب", "الرقم الوظيفي", "الموظف", "الدور", "الجوال", "الحالة", "ملاحظات"], assignment_data, {"A": 12, "B": 18, "C": 28, "D": 22, "E": 18, "F": 14, "G": 34}, "0F6B50")
    maintenance_data = [(m.request_number or "—", m.door_shift.door_number, m.get_priority_display(), m.get_status_display(), m.description, display_user(m.technician), timezone.localtime(m.created_at).strftime("%Y-%m-%d %H:%M")) for m in maintenance_requests]
    table_sheet("طلبات الصيانة", "طلبات الصيانة المرتبطة بالوردية", ["رقم الطلب", "الباب", "الأولوية", "الحالة", "وصف المشكلة", "الفني", "تاريخ الإنشاء"], maintenance_data, {"A": 22, "B": 11, "C": 15, "D": 20, "E": 44, "F": 24, "G": 22}, "C9A548")

    content = workbook.create_sheet("الملخص والتوصيات")
    configure(content, {"A": 55, "B": 55}, landscape=True)
    title_band(content, "الملخص التنفيذي والتوصيات", f"التقرير {report.report_number or '—'}", 2)
    content.sheet_properties.tabColor = gold
    for column, label in enumerate(("الملخص التنفيذي", "التوصيات"), 1):
        cell = content.cell(5, column, label)
        cell.font = Font(name="Arial", size=12, bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=green)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        body = content.cell(6, column, (report.summary if column == 1 else report.recommendations) or "لا توجد بيانات مضافة")
        body.font = Font(name="Arial", size=11, color=ink)
        body.fill = PatternFill("solid", fgColor=white)
        body.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        body.border = border
    content.row_dimensions[5].height = 28
    content.row_dimensions[6].height = 220
    content.freeze_panes = "A5"
    content.print_area = "A1:B6"
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="report_{report.report_number or report.pk}.xlsx"'
    return response


# ==================================================
# تحديث التقرير ثم تصديره إلى PDF
# ==================================================

@login_required
@permission_required(
    PlatformPermissions.EXPORT_REPORT,
    message=(
        "ليس لديك صلاحية "
        "تحديث التقرير وتصديره."
    ),
)
def refresh_and_export_report_pdf_view(
    request,
    pk,
):
    """
    تحديث مؤشرات التقرير وملخصه،
    ثم تصديره إلى PDF.
    """
    report = get_object_or_404(
        _scoped_reports(request.user),
        pk=pk,
    )

    _refresh_report_metrics(
        report
    )

    report.save(
        update_fields=[
            "total_doors",
            "open_doors",
            "closed_doors",
            "maintenance_doors",
            "total_employees",
            "total_maintenance_requests",
            "completed_maintenance_requests",
        ]
    )

    ReportService.regenerate_summary(
        request=request,
        report=report,
    )

    return export_report_pdf_view(
        request,
        report.pk,
    )
